from multiprocessing import context
from urllib import request
import csv
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.models import Group, Permission
from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.urls import reverse
from django.views.decorators.cache import never_cache
from django.utils.http import url_has_allowed_host_and_scheme
from .avatar_presets import AVATAR_COLORS, get_avatar_options
from .models import (
    Produkt, Objednavka, Stroj, VyrobnyZaznam, KontrolaKvality, 
    HlasenieVyroby, Operacia, Kontrakt, Material, VyrobnaDavka,
    SkladHotovychDielov, PrijemkaHotovychDielov, VydajkaHotovychDielov,
    PrijemkaNaSklad, VydajkaZoSkladu, OperaciaVyroby, OperatorNaOperacii,
    KontrolnyParameter, MeraniePriKontrole, MaterialAINavrh,
    DocumentAuditLog, ProduktDokument,
    ProduktKartaNastavenia, Kaliber, InventurnyZaznamMaterialu,
    DochadzkovyToken, DochadzkovyZaznam, DovolenkaZiadost,
)
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.template.loader import render_to_string
from django.db import transaction
from django.db import connection
from django.db.models import Q, Sum, Count, F, Prefetch, OuterRef, Subquery, IntegerField
from django.db.models.functions import TruncDate, Coalesce
from django.core.cache import cache
from django.core.files.base import ContentFile
import json
import io
import os
import hashlib
import unicodedata
import re
import shutil
from pathlib import Path
from .pdf_generator import generate_sprievodka_pdf
from datetime import datetime, timedelta, date
from math import ceil
from urllib.parse import quote, urlparse, urlencode
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict


logger = logging.getLogger(__name__)


PREDEFINED_OPERATIONS = [
    'Výroba',
    'Omielanie',
    'Balenie',
    'Čistenie - ofukovanie',
    'Meranie',
    'Opracovanie',
]

PACKAGING_RULE_TYPES = [
    'Krabička papierová - voľne uložené',
    'Krabička papierová - Tyčka D3',
    'Krabička papierová - Tyčka D4',
    'Krabička papierová - Tyčka D5',
    'Krabička papierová - Tyčka D6',
    'Krabička papierová - diely zabalene v streč folii',
    'Krabička papierová - volne uložene s papierovou preložkou',
    'Kovova prepravka - Hengstler',
    'Plastova bednička - Hengstler',
    'Plastova bednička s plastovými prepravkami - Hengstler',
    'Papierova krabica 500x400x300',
    'Papierova krabica 500x400x300 volne uložene s papierovou preložkou',
    'Papierova krabica  volne uložene s papierovou preložkou',
    'Papierova krabica  volne uložene v sačkoch',
]


def _is_packaging_operation(operation_name):
    normalized = unicodedata.normalize('NFD', str(operation_name or ''))
    normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return 'balenie' in normalized.lower()


def _api_ok(message, **extra):
    payload = {'status': 'ok', 'message': message}
    if extra:
        payload.update(extra)
    return JsonResponse(payload)


def _api_error(message, **extra):
    payload = {'status': 'error', 'message': message}
    if extra:
        payload.update(extra)
    return JsonResponse(payload)


def _is_operator_user(user):
    return bool(
        user and user.is_authenticated and not user.is_staff and user.groups.filter(name='Operatori').exists()
    )


def _can_import_orders(user):
    return bool(user and user.is_authenticated and (user.is_staff or user.has_perm('core.add_objednavka')))


def _default_post_login_redirect(user):
    if _is_operator_user(user):
        return reverse('operator_dashboard')
    if user.has_perm('core.view_objednavka'):
        return reverse('home')
    if user.has_perm('core.view_produkt'):
        return reverse('zoznam_produktov')
    return reverse('operator_dashboard')


def _resolve_post_login_redirect(request, requested_url=None):
    target = (requested_url or '').strip()
    if not target:
        return _default_post_login_redirect(request.user)
    if _is_operator_user(request.user):
        home_path = reverse('home')
        if target in {home_path, '/'}:
            return reverse('operator_dashboard')
    if url_has_allowed_host_and_scheme(
        url=target,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return target
    return _default_post_login_redirect(request.user)


@login_required
@require_POST
@never_cache
def admin_restart_server(request):
    from django.contrib import messages

    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'Nemáte oprávnenie na reštart servera.')
    else:
        messages.info(request, 'Reštart servera spustite cez deployment skript.')

    return redirect(request.META.get('HTTP_REFERER') or reverse('home'))


def _build_user_avatar_url(user, size=200):
    try:
        profile = user.profile
        if profile.avatar:
            return profile.avatar.url
    except Exception:
        pass

    display_name = user.get_full_name().strip() or user.username
    initials = ''.join(part[0] for part in display_name.split()[:2]).upper() or user.username[:2].upper()
    email = (user.email or '').strip().lower()
    avatar_hash = hashlib.md5(email.encode('utf-8')).hexdigest() if email else None
    if avatar_hash:
        return f"https://www.gravatar.com/avatar/{avatar_hash}?d=identicon&s={size}"
    return f"https://ui-avatars.com/api/?name={initials}&background=0d6efd&color=fff&size={size}"
def _extract_user_initials(user):
    display_name = user.get_full_name().strip() or user.username
    initials = ''.join(part[0] for part in display_name.split()[:2]).upper() if display_name else ''
    return initials or user.username[:2].upper()


def _create_svg_avatar(initials, color_hex):
    safe_initials = (str(initials or '').strip().upper() or '?')[:3]
    safe_color = str(color_hex or '').strip().lower().replace('#', '')
    if not re.fullmatch(r'[0-9a-f]{6}', safe_color):
        safe_color = '0d6efd'

    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" width="200" height="200" viewBox="0 0 200 200" role="img" aria-label="avatar">'
        f'<rect width="200" height="200" rx="100" fill="#{safe_color}"/>'
        '<text x="100" y="112" text-anchor="middle" fill="#ffffff" font-family="Arial, sans-serif" font-size="76" font-weight="700">'
        f'{safe_initials}'
        '</text>'
        '</svg>'
    )
    return svg.encode('utf-8')


def _save_avatar_preset_for_user(user, color_hex):
    profile = getattr(user, 'profile', None)
    if profile is None:
        return False

    initials = _extract_user_initials(user)
    avatar_bytes = _create_svg_avatar(initials, color_hex)
    file_name = f"avatar_{user.id}_{color_hex.lower().replace('#', '')}.svg"
    profile.avatar.save(file_name, ContentFile(avatar_bytes), save=False)
    profile.save(update_fields=['avatar', 'updated_at'])
    return True


def _sync_product_operations_to_open_orders(produkt):
    templates = list(produkt.operacie.select_related('stroj').order_by('poradie', 'id'))
    synced_orders = 0
    skipped_orders = 0
    synced_orders_list = []

    otvorene_zakazky = (
        Objednavka.objects
        .exclude(stav='hotovo')
        .filter(produkt=produkt)
        .prefetch_related('operacie')
    )

    for objednavka in otvorene_zakazky:
        existing_operations = list(objednavka.operacie.all().order_by('poradie', 'id'))
        can_replace = True

        for operation in existing_operations:
            if operation.stav != 'caka':
                can_replace = False
                break
            if operation.operator_id or operation.datum_zaciatku or operation.datum_ukoncenia:
                can_replace = False
                break
            if any([
                operation.kusy_na_vstupe,
                operation.vyrobene_kusy,
                operation.nepodarky,
                operation.kusy_na_vystupe,
            ]):
                can_replace = False
                break

        if not can_replace:
            skipped_orders += 1
            continue

        objednavka.operacie.all().delete()
        if templates:
            OperaciaVyroby.objects.bulk_create([
                OperaciaVyroby(
                    objednavka=objednavka,
                    vyrobna_davka=getattr(objednavka, 'vyrobna_davka', None),
                    operacia_sablona=template,
                    stroj=template.stroj,
                    poradie=template.poradie,
                    nazov_operacie=template.nazov_operacie,
                    typ_balenia='',
                    cas_pripravy=template.cas_pripravy,
                    cas_kus=template.cas_kus,
                    stav='caka',
                )
                for template in templates
            ])
        synced_orders += 1
        synced_orders_list.append({
            'id': objednavka.id,
            'cislo_objednavky': objednavka.cislo_objednavky,
            'zakaznik': objednavka.zakaznik,
            'stav': objednavka.stav,
            'stav_label': objednavka.get_stav_display(),
            'detail_url': reverse('detail_zakazky', kwargs={'pk': objednavka.id}),
            'datum_pozadovane': objednavka.datum_pozadovane.strftime('%d.%m.%Y') if objednavka.datum_pozadovane else '',
        })

    return synced_orders, skipped_orders, synced_orders_list


def _build_xdatabase_product_folder(produkt):
    """Return target XDatabase folder for a product, creating a stable 3-digit prefix."""
    xdb_root = Path(settings.BASE_DIR) / 'XDatabase'
    xdb_root.mkdir(parents=True, exist_ok=True)

    part_raw = (produkt.cislo_dielu or 'unknown').replace('/', '-').replace('\\', '-')
    part = re.sub(r'\s+', '', part_raw)
    if not part:
        part = 'unknown'

    if produkt.poradove_cislo is not None:
        return xdb_root / f"{produkt.poradove_cislo:03d} - {part}"

    candidates = sorted(xdb_root.glob(f"* - {part}"))
    for candidate in candidates:
        if candidate.is_dir():
            return candidate

    return xdb_root / f"000 - {part}"


def _mirror_doc_to_xdatabase(doc):
    """Copy saved media file into XDatabase/<seq - part>/<kategoria>/... and return path."""
    if not doc.subor:
        return None

    src_path = Path(doc.subor.path)
    product_folder = _build_xdatabase_product_folder(doc.produkt)
    category_folder = product_folder / doc.kategoria
    if doc.subcesta:
        category_folder = category_folder / doc.subcesta

    category_folder.mkdir(parents=True, exist_ok=True)
    target_path = category_folder / doc.nazov_suboru
    shutil.copy2(src_path, target_path)
    return target_path


def _delete_doc_from_xdatabase(doc):
    """Delete mirrored XDatabase file for a document if it exists."""
    xdb_root = (Path(settings.BASE_DIR) / 'XDatabase').resolve()

    target_path = None
    if doc.povodna_cesta:
        try:
            candidate = Path(doc.povodna_cesta).resolve()
            if str(candidate).startswith(str(xdb_root)):
                target_path = candidate
        except Exception:
            target_path = None

    if target_path is None:
        candidate = _build_xdatabase_product_folder(doc.produkt) / doc.kategoria
        if doc.subcesta:
            candidate = candidate / doc.subcesta
        target_path = (candidate / doc.nazov_suboru).resolve()

    if target_path.exists() and target_path.is_file():
        target_path.unlink()
        return target_path
    return None


def _load_excel_kalibre_rows():
    """Load calibration rows from workbook tab 'Zoznam kalibrov' with short cache."""
    workbook_path = Path(settings.BASE_DIR) / 'Fc 2026 Strojmacher 22.xlsx'
    if not workbook_path.exists():
        return []

    cache_key = f"kalibre_excel_rows::{workbook_path.stat().st_mtime_ns}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    rows = []
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        if 'Zoznam kalibrov' not in wb.sheetnames:
            return []

        ws = wb['Zoznam kalibrov']
        for r in range(5, ws.max_row + 1):
            cislo = ws.cell(r, 1).value
            zakaznik = ws.cell(r, 2).value
            suciastka = ws.cell(r, 3).value
            rozmer_ok = ws.cell(r, 4).value
            rozmer_nok = ws.cell(r, 5).value
            datum_kontroly = ws.cell(r, 6).value
            vykonal = ws.cell(r, 7).value

            if not any([cislo, zakaznik, suciastka, rozmer_ok, rozmer_nok, datum_kontroly, vykonal]):
                continue

            rows.append(
                {
                    'cislo': str(cislo).strip() if cislo is not None else '',
                    'zakaznik': str(zakaznik).strip() if zakaznik is not None else '',
                    'suciastka': str(suciastka).strip() if suciastka is not None else '',
                    'rozmer_ok': str(rozmer_ok).strip() if rozmer_ok is not None else '',
                    'rozmer_nok': str(rozmer_nok).strip() if rozmer_nok is not None else '',
                    'datum_kontroly': str(datum_kontroly).strip() if datum_kontroly is not None else '',
                    'vykonal': str(vykonal).strip() if vykonal is not None else '',
                }
            )
    except Exception:
        logger.exception('Nepodarilo sa nacitat kalibre z Excel suboru.')
        rows = []

    cache.set(cache_key, rows, 120)
    return rows


def _check_operator_rate_limit(request, action, limit=20, window_seconds=60):
    if request.user.is_authenticated:
        actor = f'user:{request.user.id}'
    else:
        forwarded_for = (request.META.get('HTTP_X_FORWARDED_FOR') or '').split(',')[0].strip()
        remote_addr = (forwarded_for or request.META.get('REMOTE_ADDR') or 'anonymous').strip() or 'anonymous'
        actor = f'ip:{remote_addr}'

    cache_key = f'operator-rate-limit:{action}:{actor}'
    current = cache.get(cache_key)

    if current is None:
        cache.set(cache_key, 1, timeout=window_seconds)
        return None

    try:
        current = cache.incr(cache_key)
    except Exception:
        current = int(current) + 1
        cache.set(cache_key, current, timeout=window_seconds)

    if current > limit:
        return _api_error(
            'Príliš veľa požiadaviek za krátky čas. Skúste to znova o chvíľu.',
            retry_after_seconds=window_seconds,
        )

    return None


def _safe_decimal(value, default='0'):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _normalize_dochadzka_identifikator(value):
    return str(value or '').strip().upper()


def _generate_dochadzka_identifikator(user, prefix='EMP'):
    base_prefix = _normalize_dochadzka_identifikator(prefix) or 'EMP'
    base = f'{base_prefix}-{user.id:04d}'
    candidate = base
    suffix = 2

    while DochadzkovyToken.objects.filter(identifikator=candidate).exists():
        candidate = f'{base}-{suffix}'
        suffix += 1

    return candidate


def _format_dochadzka_duration(delta):
    if not delta:
        return ''

    total_seconds = max(int(delta.total_seconds()), 0)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f'{hours:02d}:{minutes:02d}'


def _build_dochadzka_summary_rows(queryset):
    rows = {}

    for zaznam in queryset.select_related('user', 'token').order_by(
        'user__first_name', 'user__last_name', 'user__username', 'cas_udalosti', 'id'
    ):
        local_dt = timezone.localtime(zaznam.cas_udalosti)
        datum = local_dt.date()
        key = (zaznam.user_id, datum)
        row = rows.setdefault(key, {
            'user_id': zaznam.user_id,
            'datum': datum,
            'meno': zaznam.user.get_full_name().strip() or zaznam.user.username,
            'prichod': None,
            'odchod': None,
            'prichod_manual': False,
            'odchod_manual': False,
            'manual_adjusted': False,
            'stav': 'Bez príchodu',
            'trvanie': '',
            'pocet_udalosti': 0,
            'udalosti': [],
            'poznamky': [],
        })
        row['pocet_udalosti'] += 1
        row['udalosti'].append(f"{local_dt:%H:%M} {zaznam.get_typ_udalosti_display()}")

        is_manual = zaznam.zdroj == 'ADMIN'
        if is_manual:
            row['manual_adjusted'] = True
        if zaznam.poznamka:
            row['poznamky'].append(zaznam.poznamka.strip())

        if zaznam.typ_udalosti == 'IN' and row['prichod'] is None:
            row['prichod'] = local_dt
            row['prichod_manual'] = is_manual
        if zaznam.typ_udalosti == 'OUT':
            row['odchod'] = local_dt
            row['odchod_manual'] = is_manual

    summary_rows = []
    for row in rows.values():
        if row['prichod'] and row['odchod']:
            row['stav'] = 'Uzavreté'
            row['trvanie'] = _format_dochadzka_duration(row['odchod'] - row['prichod'])
        elif row['prichod']:
            row['stav'] = 'Otvorená dochádzka'

        row['prichod_text'] = row['prichod'].strftime('%H:%M') if row['prichod'] else '-'
        row['odchod_text'] = row['odchod'].strftime('%H:%M') if row['odchod'] else '-'
        row['prichod_time_value'] = row['prichod'].strftime('%H:%M') if row['prichod'] else ''
        row['odchod_time_value'] = row['odchod'].strftime('%H:%M') if row['odchod'] else ''
        row['udalosti_text'] = ' | '.join(row['udalosti'])
        unique_notes = []
        for note in row.get('poznamky', []):
            if note and note not in unique_notes:
                unique_notes.append(note)
        row['poznamka'] = unique_notes[0] if unique_notes else ''
        row['poznamka_text'] = ' | '.join(unique_notes)
        summary_rows.append(row)

    return sorted(summary_rows, key=lambda item: (item['datum'], item['meno']))


def _parse_iso_date(raw_value, fallback):
    value = str(raw_value or '').strip()
    if not value:
        return fallback

    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return fallback


def _parse_iso_month(raw_value, fallback):
    value = str(raw_value or '').strip()
    if not value:
        return fallback.year, fallback.month

    try:
        parsed = datetime.strptime(value, '%Y-%m')
        return parsed.year, parsed.month
    except ValueError:
        return fallback.year, fallback.month


def _attendance_range_for_day(selected_date):
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(selected_date, datetime.min.time()), tz)
    end = start + timedelta(days=1)
    return start, end


def _attendance_range_for_month(year, month):
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime(year, month, 1), tz)
    if month == 12:
        end = timezone.make_aware(datetime(year + 1, 1, 1), tz)
    else:
        end = timezone.make_aware(datetime(year, month + 1, 1), tz)
    return start, end


def _serialize_kiosk_rows(summary_rows):
    serialized = []
    for row in summary_rows:
        serialized.append({
            'date_label': row['datum'].strftime('%d.%m.%Y') if row.get('datum') else '-',
            'arrival': row.get('prichod_text') or '-',
            'departure': row.get('odchod_text') or '-',
            'duration': row.get('trvanie') or '-',
            'status': row.get('stav') or '-',
        })
    return serialized


def _kiosk_month_rows_for_user(user, reference_date=None):
    ref_date = reference_date or timezone.localdate()
    start, end = _attendance_range_for_month(ref_date.year, ref_date.month)
    queryset = DochadzkovyZaznam.objects.filter(
        user=user,
        cas_udalosti__gte=start,
        cas_udalosti__lt=end,
    )
    summary_rows = _build_dochadzka_summary_rows(queryset)
    return _serialize_kiosk_rows(summary_rows), ref_date.strftime('%m/%Y')


def _validate_attendance_sequence(user, typ_udalosti, cas_udalosti):
    latest_record = (
        DochadzkovyZaznam.objects
        .filter(user=user)
        .order_by('-cas_udalosti', '-id')
        .first()
    )

    if latest_record and cas_udalosti < latest_record.cas_udalosti:
        latest_local = timezone.localtime(latest_record.cas_udalosti)
        return (
            False,
            f'Ručný záznam musí byť po poslednej udalosti ({latest_local:%d.%m.%Y %H:%M}).',
        )

    if typ_udalosti == 'OUT':
        if not latest_record or latest_record.typ_udalosti != 'IN':
            return False, 'Nie je možné uložiť odchod bez predchádzajúceho príchodu.'
        return True, ''

    if typ_udalosti == 'IN':
        if latest_record and latest_record.typ_udalosti == 'IN':
            return False, 'Nie je možné uložiť ďalší príchod bez ukončenia predchádzajúceho.'
        return True, ''

    return False, 'Neplatný typ dochádzkovej udalosti.'


def _get_allowed_material_ai_domains():
    configured = getattr(settings, 'AI_MATERIAL_ALLOWED_DOMAINS', [])
    domains = [str(item).strip().lower() for item in configured if str(item).strip()]
    if domains:
        return domains
    return ['ferona.sk', 'profimetal.sk', 'mersteel.eu']


def _is_ai_material_enabled():
    return bool(getattr(settings, 'AI_MATERIAL_ENABLED', False))


def _is_allowed_material_ai_url(source_url):
    if not source_url:
        return True, ''
    try:
        parsed = urlparse(source_url)
    except ValueError:
        return False, ''

    if parsed.scheme not in {'http', 'https'} or not parsed.netloc:
        return False, ''

    host = parsed.netloc.lower().split(':')[0]
    allowed = _get_allowed_material_ai_domains()
    is_allowed = any(host == domain or host.endswith(f'.{domain}') for domain in allowed)
    return is_allowed, host


def _extract_first_json_object(raw_text):
    if not raw_text:
        return None
    start = raw_text.find('{')
    end = raw_text.rfind('}')
    if start < 0 or end < 0 or end <= start:
        return None
    candidate = raw_text[start:end + 1]
    try:
        return json.loads(candidate)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None


def _generate_material_ai_response(query, source_url=''):
    api_key = getattr(settings, 'OPENAI_API_KEY', '')
    if not api_key:
        raise ValueError('Chýba OPENAI_API_KEY v nastaveniach servera.')

    try:
        from openai import OpenAI
    except Exception as exc:
        raise ValueError('Knižnica openai nie je nainštalovaná. Doplň ju do requirements.') from exc

    model_name = str(getattr(settings, 'OPENAI_MATERIAL_MODEL', 'gpt-4.1-mini') or 'gpt-4.1-mini').strip()
    timeout_seconds = float(getattr(settings, 'OPENAI_TIMEOUT_SECONDS', 20))
    client = OpenAI(api_key=api_key, timeout=timeout_seconds, max_retries=0)

    source_hint = source_url.strip() if source_url else 'Bez konkrétnej URL, hľadaj na dôveryhodných slovenských weboch.'
    allowed_domains = ', '.join(_get_allowed_material_ai_domains())

    prompt = (
        'Si asistent pre nákup materiálu v strojárskej výrobe. '\
        'Vráť STRICTNE iba JSON objekt bez markdownu. '\
        'Použi verejne dostupné údaje a keď si nie si istý, nechaj hodnotu null. '\
        f'Povolené domény: {allowed_domains}. '\
        f'Vstupný dopyt: {query}. '\
        f'Zdroj: {source_hint}. '\
        'JSON schema: '
        '{'
        '"nazov":"",'
        '"kod":"",'
        '"typ":"SUROVINA|POLOTOVAR|KOMPONENT",'
        '"jednotka":"kg|ks|tyc",'
        '"minimalna_zasoba":0,'
        '"cena_za_jednotku":0,'
        '"priemer_mm":0,'
        '"tyc_dlzka_m":0,'
        '"kg_na_meter":0,'
        '"aktualna_zasoba":0,'
        '"confidence":0.0,'
        '"poznamka":"stručné zhrnutie zdroja a neistôt"'
        '}'
    )

    try:
        response = client.responses.create(
            model=model_name,
            input=prompt,
        )
    except Exception as exc:
        message = str(exc).lower()
        if 'timed out' in message or 'timeout' in message:
            raise ValueError('AI služba neodpovedala v časovom limite. Skontroluj internet na NASe a skús to znova.') from exc
        if 'model' in message and ('not found' in message or 'does not exist' in message):
            raise ValueError(f'AI model "{model_name}" nie je dostupný. Skontroluj OPENAI_MATERIAL_MODEL v .env.') from exc
        if '429' in message or 'quota' in message or 'billing' in message:
            raise ValueError('OpenAI účet nemá dostupný kredit/quota. Skontroluj billing na platform.openai.com.') from exc
        raise

    raw_text = getattr(response, 'output_text', '') or ''
    parsed = _extract_first_json_object(raw_text)
    if not parsed:
        raise ValueError('AI vrátila neplatný formát. Skús upresniť dopyt alebo URL.')

    return {
        'model': model_name,
        'raw_text': raw_text,
        'data': parsed,
    }


def _serialize_material_ai_navrh(navrh):
    data = navrh.navrh_data or {}
    return {
        'id': navrh.pk,
        'stav': navrh.stav,
        'query': navrh.query,
        'source_url': navrh.source_url,
        'source_domain': navrh.source_domain,
        'ai_model': navrh.ai_model,
        'confidence': float(navrh.confidence) if navrh.confidence is not None else None,
        'navrh': {
            'nazov': data.get('nazov') or '',
            'kod': data.get('kod') or '',
            'typ': data.get('typ') or 'SUROVINA',
            'jednotka': data.get('jednotka') or 'kg',
            'minimalna_zasoba': str(data.get('minimalna_zasoba') if data.get('minimalna_zasoba') is not None else 0),
            'cena_za_jednotku': str(data.get('cena_za_jednotku') if data.get('cena_za_jednotku') is not None else 0),
            'priemer_mm': str(data.get('priemer_mm') if data.get('priemer_mm') is not None else 0),
            'tyc_dlzka_m': str(data.get('tyc_dlzka_m') if data.get('tyc_dlzka_m') is not None else 0),
            'kg_na_meter': str(data.get('kg_na_meter') if data.get('kg_na_meter') is not None else 0),
            'aktualna_zasoba': str(data.get('aktualna_zasoba') if data.get('aktualna_zasoba') is not None else 0),
            'poznamka': data.get('poznamka') or '',
        },
    }


def _normalize_unit(value):
    return unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii').strip().lower()


def _is_kg_unit(value):
    return _normalize_unit(value) == 'kg'


def _is_bar_unit(value):
    normalized = _normalize_unit(value)
    return normalized in {'tyc', 'tyce', 'ks'}


def _calculate_material_requirement(produkt, mnozstvo):
    material = getattr(produkt, 'material_ref', None)
    if not material or not mnozstvo or mnozstvo <= 0:
        return None

    dlzka_na_kus_mm = float(produkt.dlzka_na_kus_mm or 0)
    if dlzka_na_kus_mm <= 0:
        return None

    zasoba = float(material.aktualna_zasoba or 0)
    jednotka = material.jednotka or ''
    normalized_unit = _normalize_unit(jednotka)
    dlzka_m = (float(mnozstvo) * dlzka_na_kus_mm) / 1000.0

    if _is_kg_unit(normalized_unit):
        kg_na_meter = float(material.kg_na_meter or 0)
        if kg_na_meter <= 0:
            return None
        potreba = dlzka_m * kg_na_meter
        return {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'potreba': potreba,
            'zasoba': zasoba,
            'jednotka': 'kg',
            'display_decimals': 2,
        }

    if _is_bar_unit(normalized_unit):
        tyc_dlzka_m = float(material.tyc_dlzka_m or 0)
        if tyc_dlzka_m <= 0:
            return None
        potreba_tyce = ceil(dlzka_m / tyc_dlzka_m)
        return {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'potreba': float(potreba_tyce),
            'zasoba': zasoba,
            'jednotka': jednotka or 'ks',
            'display_decimals': 0,
        }

    return None


def _build_material_shortage_message(produkt, mnozstvo):
    data = _calculate_material_requirement(produkt, mnozstvo)
    if not data:
        return None

    nedostatok = max(0.0, data['potreba'] - data['zasoba'])
    if nedostatok <= 0:
        return None

    decimals = data['display_decimals']
    needed_display = f"{nedostatok:.{decimals}f}"
    return (
        f"Nedostatok materiálu: chýba {needed_display} {data['jednotka']} "
        f"pre materiál {data['material_nazov']} ({data['material_kod']})."
    )


def _user_has_operator_access(user, objednavka):
    return user in objednavka.priradeni_operatori.all() or objednavka.zaznamy.filter(operator=user).exists()


def _get_json_body(request):
    if not request.body:
        return {}
    try:
        data = json.loads(request.body)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _get_active_operator_session(operacia):
    return operacia.operatori.filter(cas_konca__isnull=True).select_related('operator').order_by('-cas_zaciatku').first()


def _close_other_open_operator_sessions(operacia, current_user=None):
    now = timezone.now()
    queryset = operacia.operatori.filter(cas_konca__isnull=True)
    if current_user is not None:
        queryset = queryset.exclude(operator=current_user)
    queryset.update(cas_konca=now)


def _get_or_create_open_operator_session(operacia, operator_user):
    session = operacia.operatori.filter(
        operator=operator_user,
        cas_konca__isnull=True,
    ).order_by('-cas_zaciatku').first()
    if session:
        return session
    return OperatorNaOperacii.objects.create(
        operacia=operacia,
        operator=operator_user,
        cas_zaciatku=timezone.now(),
    )


def _finish_operation_batch(operacia, operator_user, vyrobene, nepodarky):
    _close_other_open_operator_sessions(operacia, operator_user)
    operator_zaznam = _get_or_create_open_operator_session(operacia, operator_user)

    operacia.ukonci_davku(vyrobene, nepodarky)

    operator_zaznam.cas_konca = timezone.now()
    operator_zaznam.vyrobene_kusy += vyrobene
    operator_zaznam.save()

    zostava = operacia.get_dostupne_kusy_na_vstupe()
    if operacia.stav == 'hotova':
        return (
            f'Operácia ÚPLNE UKONČENÁ! '
            f'Celkovo vyrobené: {operacia.vyrobene_kusy} ks, Nepodarky: {operacia.nepodarky} ks'
        )

    return f'Dávka ukončená! Vyrobené: {vyrobene} ks, Nepodarky: {nepodarky} ks. Zostáva ešte: {zostava} ks'


def _close_order_with_packaging_photo(zakazka, operator_user, fotka_balenia, poznamka_balenia=''):
    if not fotka_balenia:
        raise ValueError('Pri finálnom uzavretí je povinná fotka balenia.')

    KontrolaKvality.objects.create(
        objednavka=zakazka,
        operator=operator_user,
        typ_kontroly='FINALNA',
        pocet_ok_kusov=zakazka.celkom_ok_kusy,
        pocet_nok_kusov=zakazka.celkom_nok_kusy,
        namerana_hodnota='Finálna kontrola balenia',
        vysledok_ok=True,
        fotka_balenia=fotka_balenia,
        poznamka=poznamka_balenia,
    )

    zakazka.uzavri_zakazku()
    return f'Zakázka #{zakazka.cislo_objednavky} bola uzavretá a hotové diely boli naskladnené!'


def _book_finished_goods_delta(objednavka, operator_user=None, note=''):
    uz_naskladnene = PrijemkaHotovychDielov.objects.filter(
        objednavka=objednavka
    ).aggregate(total=Sum('mnozstvo'))['total'] or 0

    aktualne_ok = max(0, int(objednavka.celkom_ok_kusy or 0))
    delta = max(0, aktualne_ok - int(uz_naskladnene))

    if delta <= 0:
        return 0

    sklad, _ = SkladHotovychDielov.objects.get_or_create(
        produkt=objednavka.produkt,
        defaults={
            'mnozstvo': 0,
            'minimalna_zasoba': 10,
            'optimalna_zasoba': 100,
        }
    )

    base_note = note.strip() if note else f'Priebežné naskladnenie zo zákazky #{objednavka.cislo_objednavky}'
    PrijemkaHotovychDielov.objects.create(
        sklad=sklad,
        objednavka=objednavka,
        mnozstvo=delta,
        datum=timezone.now(),
        operator=operator_user,
        poznamka=base_note,
    )
    return delta


@login_required
@require_POST
@never_cache
def quick_logout(request):
    logout(request)
    response = redirect('/accounts/login/')
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
@never_cache
def admin_users_view(request):
    if not request.user.is_staff or not request.user.has_perm('auth.view_user'):
        return render(request, 'core/no_access.html')

    can_add_tokens = request.user.has_perm('core.add_dochadzkovytoken')
    can_change_tokens = request.user.has_perm('core.change_dochadzkovytoken')

    if request.method == 'POST':
        from django.contrib import messages

        action = (request.POST.get('action') or 'create-token').strip()

        if action == 'toggle-token':
            if not can_change_tokens:
                messages.error(request, '❌ Nemáte oprávnenie meniť stav dochádzkových tokenov.')
                return redirect('admin_users_view')

            token_id_raw = (request.POST.get('token_id') or '').strip()
            if not token_id_raw.isdigit():
                messages.error(request, '❌ Vyberte platný token.')
                return redirect('admin_users_view')

            token = get_object_or_404(DochadzkovyToken.objects.select_related('user'), pk=int(token_id_raw))
            token.aktivny = not token.aktivny
            token.save(update_fields=['aktivny', 'updated_at'])

            stav = 'aktivovaný' if token.aktivny else 'deaktivovaný'
            messages.success(request, f'✅ Token {token.identifikator} bol {stav}.')
            return redirect('admin_users_view')

        if not can_add_tokens:
            messages.error(request, '❌ Nemáte oprávnenie vytvárať dochádzkové tokeny.')
            return redirect('admin_users_view')

        user_id_raw = (request.POST.get('user_id') or '').strip()
        token_type = _normalize_dochadzka_identifikator(request.POST.get('typ') or 'MANUAL')
        identifikator = _normalize_dochadzka_identifikator(request.POST.get('identifikator'))
        nazov = str(request.POST.get('nazov') or '').strip()

        if not user_id_raw.isdigit():
            messages.error(request, '❌ Vyberte platného používateľa.')
            return redirect('admin_users_view')

        if token_type not in dict(DochadzkovyToken.TYP_CHOICES):
            messages.error(request, '❌ Neplatný typ tokenu.')
            return redirect('admin_users_view')

        UserModel = get_user_model()
        user_obj = get_object_or_404(UserModel.objects.select_related('profile'), pk=int(user_id_raw))
        if not identifikator:
            identifikator = _generate_dochadzka_identifikator(user_obj)

        if DochadzkovyToken.objects.filter(identifikator=identifikator).exists():
            messages.error(request, f'❌ Identifikátor {identifikator} už existuje.')
            return redirect('admin_users_view')

        DochadzkovyToken.objects.create(
            user=user_obj,
            identifikator=identifikator,
            nazov=nazov,
            typ=token_type,
        )

        if getattr(user_obj, 'profile', None) and not user_obj.profile.has_pin:
            messages.warning(request, f'⚠️ Token {identifikator} bol vytvorený, ale používateľ {user_obj.username} ešte nemá PIN.')
        else:
            messages.success(request, f'✅ Token {identifikator} bol vytvorený pre používateľa {user_obj.username}.')

        return redirect('admin_users_view')

    query = (request.GET.get('q') or '').strip()
    UserModel = get_user_model()
    users_qs = UserModel.objects.select_related('profile').prefetch_related('groups', 'dochadzkove_tokeny').order_by('username')

    if query:
        users_qs = users_qs.filter(
            Q(username__icontains=query)
            | Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(email__icontains=query)
        )

    users = []
    for user_obj in users_qs:
        full_name = user_obj.get_full_name().strip()
        try:
            profile = user_obj.profile
            has_pin = profile.has_pin
        except Exception:
            has_pin = False

        users.append({
            'obj': user_obj,
            'display_name': full_name or user_obj.username,
            'avatar_url': _build_user_avatar_url(user_obj, size=96),
            'group_names': [group.name for group in user_obj.groups.all()],
            'has_pin': has_pin,
            'attendance_tokens': list(user_obj.dochadzkove_tokeny.all().order_by('created_at')),
        })

    return render(request, 'core/admin/users.html', {
        'users': users,
        'query': query,
        'total_users': users_qs.count(),
        'can_add_tokens': can_add_tokens,
        'can_change_tokens': can_change_tokens,
        'attendance_token_types': DochadzkovyToken.TYP_CHOICES,
    })


@login_required
@never_cache
def admin_groups_view(request):
    if not request.user.is_staff or not request.user.has_perm('auth.view_group'):
        return render(request, 'core/no_access.html')

    query = (request.GET.get('q') or '').strip()
    groups_qs = Group.objects.prefetch_related('permissions', 'user_set').order_by('name')
    if query:
        groups_qs = groups_qs.filter(name__icontains=query)

    groups_data = []
    for group in groups_qs:
        perms = list(group.permissions.all())
        groups_data.append({
            'obj': group,
            'members_count': group.user_set.count(),
            'permissions_count': len(perms),
            'permissions_preview': perms[:6],
        })

    return render(request, 'core/admin/groups.html', {
        'groups': groups_data,
        'query': query,
        'total_groups': groups_qs.count(),
    })


@login_required
@never_cache
def admin_permissions_view(request):
    if not request.user.is_staff or not request.user.has_perm('auth.view_permission'):
        return render(request, 'core/no_access.html')

    app_filter = (request.GET.get('app') or '').strip()
    query = (request.GET.get('q') or '').strip()

    permissions_qs = Permission.objects.select_related('content_type').order_by(
        'content_type__app_label', 'content_type__model', 'codename'
    )

    if app_filter:
        permissions_qs = permissions_qs.filter(content_type__app_label=app_filter)
    if query:
        permissions_qs = permissions_qs.filter(
            Q(name__icontains=query)
            | Q(codename__icontains=query)
            | Q(content_type__model__icontains=query)
        )

    app_labels = list(
        Permission.objects
        .order_by('content_type__app_label')
        .values_list('content_type__app_label', flat=True)
        .distinct()
    )

    return render(request, 'core/admin/permissions.html', {
        'permissions': permissions_qs[:400],
        'query': query,
        'app_filter': app_filter,
        'app_labels': app_labels,
        'total_permissions': permissions_qs.count(),
    })


def offline_page(request):
    return render(request, 'core/offline.html')


def service_worker(request):
    cache_version = getattr(settings, 'APP_BUILD_TS', timezone.now().strftime('%Y%m%d%H%M%S')).replace('.', '').replace(':', '').replace(' ', '-')
    content = """const CACHE_NAME = 'erp-pwa-__CACHE_VERSION__';
const OFFLINE_URL = '/offline/';
const PRECACHE_URLS = [
    OFFLINE_URL,
    '/manifest.webmanifest',
    '/pwa-icon.svg'
];

self.addEventListener('install', (event) => {
    event.waitUntil(
        caches.open(CACHE_NAME).then((cache) => cache.addAll(PRECACHE_URLS))
    );
    self.skipWaiting();
        content = content.replace('__CACHE_VERSION__', cache_version)
});

self.addEventListener('activate', (event) => {
    event.waitUntil(
        caches.keys().then((keys) =>
            Promise.all(
                keys
                    .filter((key) => key !== CACHE_NAME)
                    .map((key) => caches.delete(key))
            )
        )
    );
    self.clients.claim();
});

self.addEventListener('fetch', (event) => {
    if (event.request.method !== 'GET') return;

    if (event.request.mode === 'navigate') {
        event.respondWith(
            fetch(event.request)
                .then((response) => {
                    const copy = response.clone();
                    caches.open(CACHE_NAME).then((cache) => cache.put(event.request, copy));
                    return response;
                })
                .catch(() => caches.match(OFFLINE_URL))
        );
        return;
    }

    const url = new URL(event.request.url);
    if (url.origin !== self.location.origin) return;

    event.respondWith(
        caches.match(event.request).then((cached) => {
            const networkFetch = fetch(event.request)
                .then((response) => {
                    const copy = response.clone();
                    caches.open(CACHE_NAME).then((cache) => cache.put(event.request, copy));
                    return response;
                })
                .catch(() => cached);

            return cached || networkFetch;
        })
    );
}});
"""
    response = HttpResponse(content, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def web_manifest(request):
    manifest = {
        'name': 'ERP Systém - Kovovýroba',
        'short_name': 'ERP Kov',
        'start_url': '/',
        'scope': '/',
        'display': 'standalone',
        'background_color': '#f8f9fa',
        'theme_color': '#0d6efd',
        'lang': 'sk',
        'icons': [
            {
                'src': '/pwa-icon.svg',
                'sizes': '512x512',
                'type': 'image/svg+xml',
                'purpose': 'any'
            }
        ]
    }
    response = HttpResponse(
        json.dumps(manifest, ensure_ascii=False),
        content_type='application/manifest+json'
    )
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def pwa_icon(request):
    content = render_to_string('core/pwa/icon.svg')
    return HttpResponse(content, content_type='image/svg+xml')


@never_cache
def healthz(request):
    token = (os.environ.get('HEALTHCHECK_TOKEN') or '').strip()
    provided_token = (
        request.headers.get('X-Health-Token')
        or request.GET.get('token')
        or ''
    ).strip()

    if token and provided_token != token:
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)

    db_ok = True
    db_error = ''
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT 1')
    except Exception as exc:
        db_ok = False
        db_error = str(exc)

    payload = {
        'status': 'ok' if db_ok else 'degraded',
        'app_version': getattr(settings, 'APP_VERSION', 'n/a'),
        'git_commit': getattr(settings, 'APP_GIT_COMMIT', 'unknown'),
        'build_ts': getattr(settings, 'APP_BUILD_TS', 'n/a'),
        'timestamp': timezone.now().isoformat(),
        'host': request.get_host(),
        'scheme': request.scheme,
        'source_dir': getattr(settings, 'APP_SOURCE_DIR', ''),
        'checks': {
            'db': 'ok' if db_ok else 'error',
        },
    }
    if db_error:
        payload['checks']['db_error'] = db_error

    return JsonResponse(payload, status=200 if db_ok else 503)


@never_cache
def runtime_info(request):
    payload = {
        'status': 'ok',
        'app_version': getattr(settings, 'APP_VERSION', 'n/a'),
        'git_commit': getattr(settings, 'APP_GIT_COMMIT', 'unknown'),
        'build_ts': getattr(settings, 'APP_BUILD_TS', 'n/a'),
        'server_time': timezone.now().isoformat(),
        'host': request.get_host(),
        'scheme': request.scheme,
        'path': request.path,
        'source_dir': getattr(settings, 'APP_SOURCE_DIR', ''),
    }
    response = JsonResponse(payload)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

# ========================================
# ZÁKLADNÉ VIEWS (Pre adminov/technikov)
# ========================================

@login_required
@permission_required("core.view_produkt", raise_exception=True)
def zoznam_produktov(request):
    q = (request.GET.get('q') or '').strip()
    material_filter = (request.GET.get('material') or '').strip()
    docs_filter = (request.GET.get('docs') or '').strip()

    produkty = Produkt.objects.all().prefetch_related('dokumenty').order_by('cislo_dielu')

    if q:
        produkty = produkty.filter(
            Q(cislo_dielu__icontains=q)
            | Q(nazov__icontains=q)
            | Q(cislo_vykresu__icontains=q)
            | Q(material__icontains=q)
        )

    if material_filter:
        produkty = produkty.filter(material__iexact=material_filter)

    materialy = sorted({m for m in produkty.values_list('material', flat=True) if m})

    vz_pattern = re.compile(r'\bVZ\d{2,4}-\d+\b', re.IGNORECASE)
    produkty_list = list(produkty)

    filtered_produkty = []
    for produkt in produkty_list:
        docs = list(produkt.dokumenty.all())
        docs_by_cat = {}
        for d in docs:
            docs_by_cat[d.kategoria] = docs_by_cat.get(d.kategoria, 0) + 1

        produkt.docs_total = len(docs)
        produkt.docs_by_cat = docs_by_cat
        # Drawing indicator now follows XDatabase category only.
        produkt.has_vykres = docs_by_cat.get('02_Drawing', 0) > 0
        produkt.has_balenie = bool(produkt.baliaci_predpis_pdf) or docs_by_cat.get('05_Manual_Balenie', 0) > 0
        produkt.has_cena = docs_by_cat.get('06_Cena', 0) > 0

        produkt.vzorkovanie_kod = ''
        for d in docs:
            if d.kategoria != '01_Vzorkovanie':
                continue
            match = vz_pattern.search(d.nazov_suboru or '')
            if match:
                produkt.vzorkovanie_kod = match.group(0).upper()
                break

        if docs_filter == 'with_docs' and produkt.docs_total == 0:
            continue
        if docs_filter == 'without_drawing' and produkt.has_vykres:
            continue
        if docs_filter == 'missing_price' and produkt.has_cena:
            continue

        filtered_produkty.append(produkt)

    filtered_produkty.sort(
        key=lambda p: (
            p.poradove_cislo is None,
            p.poradove_cislo if p.poradove_cislo is not None else 10**9,
            (p.cislo_dielu or ''),
        )
    )

    return render(
        request,
        "core/zoznam.html",
        {
            "produkty": filtered_produkty,
            "q": q,
            "material_filter": material_filter,
            "docs_filter": docs_filter,
            "materialy": materialy,
        },
    )

@login_required
@permission_required("core.view_produkt", raise_exception=True)
def detail_produkt(request, pk):
    from .docs_utils import is_docs_admin
    produkt = get_object_or_404(Produkt, pk=pk)
    karta_nastavenia, _ = ProduktKartaNastavenia.objects.get_or_create(produkt=produkt)
    kalibre_excel_rows = _load_excel_kalibre_rows()
    docs_qs = produkt.dokumenty.all().order_by('kategoria', 'subcesta', 'nazov_suboru')
    docs_by_category = {}
    for doc in docs_qs:
        docs_by_category.setdefault(doc.kategoria, []).append(doc)

    otvorene_zakazky = list(
        Objednavka.objects
        .exclude(stav='hotovo')
        .filter(produkt=produkt)
        .order_by('datum_pozadovane', 'id')
        .only('id', 'cislo_objednavky', 'zakaznik', 'stav', 'datum_pozadovane', 'mnozstvo')
    )
    sklad_hotovych = SkladHotovychDielov.objects.filter(produkt=produkt).first()
    material_sklad = getattr(produkt, 'material_ref', None)

    return render(
        request,
        "core/detail.html",
        {
            "produkt": produkt,
            "karta_nastavenia": karta_nastavenia,
            "docs_by_category": docs_by_category,
            "docs_total": docs_qs.count(),
            "materialy": Material.objects.order_by('nazov')[:500],
            "stroje": Stroj.objects.order_by('nazov'),
            "operacie": produkt.operacie.select_related('stroj').all().order_by('poradie'),
            "preddefinovane_operacie": PREDEFINED_OPERATIONS,
            "typy_baliaceho_predpisu": PACKAGING_RULE_TYPES,
            "otvorene_zakazky": otvorene_zakazky,
            "sklad_hotovych": sklad_hotovych,
            "material_sklad": material_sklad,
            "kontrolne_parametre": produkt.kontrolne_parametre.all().order_by('poradie'),
            "kalibre_excel_rows": kalibre_excel_rows,
            "kalibre_excel_preview": [r.get('cislo', '') for r in kalibre_excel_rows[:4] if r.get('cislo')],
            "user_is_docs_admin": is_docs_admin(request.user),
        },
    )


@login_required
@permission_required("core.change_kaliber", raise_exception=True)
@require_POST
def uloz_kaliber(request):
    kaliber_id = (request.POST.get('kaliber_id') or '').strip()
    cislo = (request.POST.get('cislo') or '').strip()
    if not cislo:
        return _api_error('Pole "Číslo" je povinné.')

    kontrola_platna_do_raw = (request.POST.get('kontrola_platna_do') or '').strip()
    kontrola_platna_do = None
    if kontrola_platna_do_raw:
        try:
            kontrola_platna_do = datetime.strptime(kontrola_platna_do_raw, '%Y-%m-%d').date()
        except ValueError:
            return _api_error('Dátum "Platnosť kontroly do" má nesprávny formát.')

    payload = {
        'cislo': cislo,
        'zakaznik': (request.POST.get('zakaznik') or '').strip(),
        'cislo_suciastky': (request.POST.get('cislo_suciastky') or '').strip(),
        'rozmer_ok': (request.POST.get('rozmer_ok') or '').strip(),
        'rozmer_nok': (request.POST.get('rozmer_nok') or '').strip(),
        'datum_kontroly': (request.POST.get('datum_kontroly') or '').strip(),
        'kontrola_platna_do': kontrola_platna_do,
        'vykonal': (request.POST.get('vykonal') or '').strip(),
        'poznamka': (request.POST.get('poznamka') or '').strip(),
    }

    if kaliber_id:
        kaliber = get_object_or_404(Kaliber, pk=kaliber_id)
        for k, v in payload.items():
            setattr(kaliber, k, v)
        if request.FILES.get('fotka'):
            kaliber.fotka = request.FILES['fotka']
        kaliber.save()
        action = 'updated'
    else:
        if request.FILES.get('fotka'):
            payload['fotka'] = request.FILES['fotka']
        kaliber = Kaliber.objects.create(**payload)
        action = 'created'

    dnes = timezone.localdate()
    if kaliber.kontrola_platna_do:
        zostava_dni = (kaliber.kontrola_platna_do - dnes).days
        if zostava_dni < 0:
            platnost_stav = 'expired'
            platnost_text = 'Po termíne'
        elif zostava_dni <= 30:
            platnost_stav = 'warning'
            platnost_text = f'Končí o {zostava_dni} dní'
        else:
            platnost_stav = 'ok'
            platnost_text = f'Platný ({zostava_dni} dní)'
    else:
        platnost_stav = 'unknown'
        platnost_text = 'Bez termínu platnosti'

    return _api_ok(
        'Kaliber bol uložený.',
        action=action,
        kaliber={
            'id': kaliber.id,
            'cislo': kaliber.cislo,
            'zakaznik': kaliber.zakaznik,
            'cislo_suciastky': kaliber.cislo_suciastky,
            'rozmer_ok': kaliber.rozmer_ok,
            'rozmer_nok': kaliber.rozmer_nok,
            'datum_kontroly': kaliber.datum_kontroly,
            'kontrola_platna_do': kaliber.kontrola_platna_do.isoformat() if kaliber.kontrola_platna_do else '',
            'platnost_stav': platnost_stav,
            'platnost_text': platnost_text,
            'vykonal': kaliber.vykonal,
            'poznamka': kaliber.poznamka,
            'fotka_url': kaliber.fotka.url if kaliber.fotka else '',
        },
    )


@login_required
@permission_required("core.delete_kaliber", raise_exception=True)
@require_POST
def zmaz_kaliber(request, kaliber_id):
    kaliber = get_object_or_404(Kaliber, pk=kaliber_id)
    kaliber.delete()
    return _api_ok('Kaliber bol zmazaný.')


@login_required
@permission_required("core.change_produkt", raise_exception=True)
@require_POST
def uloz_kartu_produktu_etapa_a(request, pk):
    produkt = get_object_or_404(Produkt, pk=pk)
    karta, _ = ProduktKartaNastavenia.objects.get_or_create(produkt=produkt)

    changed_produkt = []
    changed_karta = []

    material_ref_raw = (request.POST.get('material_ref_id') or '').strip()
    material_text = (request.POST.get('material') or '').strip()
    rozmer = (request.POST.get('rozmer_polotovaru') or '').strip()
    spotreba_raw = (request.POST.get('spotreba_ks') or '').strip()
    norma_raw = (request.POST.get('norma_hod') or '').strip()

    upozornenie_operator = (request.POST.get('upozornenie_operator') or '').strip()
    reklamacie_poznamky = (request.POST.get('reklamacie_poznamky') or '').strip()
    baliaci_predpis_text = (request.POST.get('baliaci_predpis_text') or '').strip()
    typ_baliaceho_predpisu = (request.POST.get('typ_baliaceho_predpisu') or '').strip()
    kalibre_poznamky = (request.POST.get('kalibre_poznamky') or '').strip()
    vyrobny_postup_poznamky = (request.POST.get('vyrobny_postup_poznamky') or '').strip()
    operacie_raw = (request.POST.get('operacie_json') or '').strip()

    if typ_baliaceho_predpisu and typ_baliaceho_predpisu not in PACKAGING_RULE_TYPES:
        return _api_error('Vybraný typ baliaceho predpisu nie je platný.')

    material_ref = None
    if material_ref_raw:
        material_ref = Material.objects.filter(pk=material_ref_raw).first()
        if not material_ref:
            return _api_error('Vybraný materiál neexistuje.')

    if produkt.material_ref != material_ref:
        produkt.material_ref = material_ref
        changed_produkt.append('material_ref')

    if material_text != produkt.material:
        produkt.material = material_text
        changed_produkt.append('material')

    if rozmer != produkt.rozmer_polotovaru:
        produkt.rozmer_polotovaru = rozmer
        changed_produkt.append('rozmer_polotovaru')

    try:
        spotreba_val = Decimal(spotreba_raw.replace(',', '.')) if spotreba_raw else Decimal('0')
    except (InvalidOperation, ValueError):
        return _api_error('Spotreba na kus musí byť číslo.')

    if spotreba_val != produkt.spotreba_ks:
        produkt.spotreba_ks = spotreba_val
        changed_produkt.append('spotreba_ks')

    try:
        norma_val = int(norma_raw) if norma_raw else 0
    except ValueError:
        return _api_error('Norma (ks/hod) musí byť celé číslo.')

    if norma_val < 0:
        return _api_error('Norma (ks/hod) nemôže byť záporná.')

    if norma_val != produkt.norma_hod:
        produkt.norma_hod = norma_val
        changed_produkt.append('norma_hod')

    if upozornenie_operator != karta.upozornenie_operator:
        karta.upozornenie_operator = upozornenie_operator
        changed_karta.append('upozornenie_operator')

    if reklamacie_poznamky != karta.reklamacie_poznamky:
        karta.reklamacie_poznamky = reklamacie_poznamky
        changed_karta.append('reklamacie_poznamky')

    if baliaci_predpis_text != karta.baliaci_predpis_text:
        karta.baliaci_predpis_text = baliaci_predpis_text
        changed_karta.append('baliaci_predpis_text')

    if typ_baliaceho_predpisu != karta.typ_baliaceho_predpisu:
        karta.typ_baliaceho_predpisu = typ_baliaceho_predpisu
        changed_karta.append('typ_baliaceho_predpisu')

    if kalibre_poznamky != karta.kalibre_poznamky:
        karta.kalibre_poznamky = kalibre_poznamky
        changed_karta.append('kalibre_poznamky')

    if vyrobny_postup_poznamky != karta.vyrobny_postup_poznamky:
        karta.vyrobny_postup_poznamky = vyrobny_postup_poznamky
        changed_karta.append('vyrobny_postup_poznamky')

    operacie_payload = []
    if operacie_raw:
        try:
            parsed_operacie = json.loads(operacie_raw)
        except json.JSONDecodeError:
            return _api_error('Operácie majú neplatný formát.')

        if not isinstance(parsed_operacie, list):
            return _api_error('Operácie musia byť zoznam.')

        for index, item in enumerate(parsed_operacie, start=1):
            if not isinstance(item, dict):
                return _api_error(f'Operácia #{index} má neplatný formát.')

            nazov_operacie = str(item.get('nazov_operacie') or '').strip()
            stroj_id = str(item.get('stroj_id') or '').strip()
            poradie_raw = str(item.get('poradie') or '').strip()
            cas_pripravy_raw = str(item.get('cas_pripravy') or '').strip()
            cas_kus_raw = str(item.get('cas_kus') or '').strip()

            if not any([nazov_operacie, stroj_id, poradie_raw, cas_pripravy_raw, cas_kus_raw]):
                continue

            if not nazov_operacie:
                return _api_error(f'Operácia #{index} musí mať názov.')
            is_packaging = _is_packaging_operation(nazov_operacie)
            if not stroj_id and not is_packaging:
                return _api_error(f'Operácia #{index} musí mať vybraný stroj.')

            try:
                poradie = int(poradie_raw) if poradie_raw else index
            except ValueError:
                return _api_error(f'Poradie operácie #{index} musí byť celé číslo.')

            try:
                cas_pripravy = int(cas_pripravy_raw) if cas_pripravy_raw else 0
            except ValueError:
                return _api_error(f'Čas prípravy operácie #{index} musí byť celé číslo.')

            try:
                cas_kus = Decimal(cas_kus_raw.replace(',', '.')) if cas_kus_raw else Decimal('0')
            except (InvalidOperation, ValueError):
                return _api_error(f'Čas na kus pri operácii #{index} musí byť číslo.')

            if poradie < 1:
                return _api_error(f'Poradie operácie #{index} musí byť aspoň 1.')
            if is_packaging:
                cas_pripravy = 0
            if cas_pripravy < 0:
                return _api_error(f'Čas prípravy operácie #{index} nemôže byť záporný.')
            if cas_kus < 0:
                return _api_error(f'Čas na kus pri operácii #{index} nemôže byť záporný.')

            stroj = Stroj.objects.filter(pk=stroj_id).first() if stroj_id else None
            if not stroj and is_packaging:
                stroj = Stroj.objects.order_by('id').first()
            if not stroj:
                return _api_error(f'Vybraný stroj pri operácii #{index} neexistuje.')

            operacie_payload.append({
                'stroj': stroj,
                'poradie': poradie,
                'nazov_operacie': nazov_operacie,
                'typ_balenia': '',
                'cas_pripravy': cas_pripravy,
                'cas_kus': cas_kus,
            })

    synced_orders = 0
    skipped_orders = 0
    synced_orders_list = []

    with transaction.atomic():
        if changed_produkt:
            produkt.save(update_fields=changed_produkt)

        if changed_karta:
            karta.save(update_fields=changed_karta + ['updated_at'])

        if operacie_raw:
            existing_templates = list(produkt.operacie.order_by('poradie', 'id'))

            for index, payload in enumerate(operacie_payload):
                if index < len(existing_templates):
                    template = existing_templates[index]
                    changed_fields_template = []
                    for field_name, new_value in payload.items():
                        if getattr(template, field_name) != new_value:
                            setattr(template, field_name, new_value)
                            changed_fields_template.append(field_name)
                    if changed_fields_template:
                        template.save(update_fields=changed_fields_template)
                else:
                    Operacia.objects.create(produkt=produkt, **payload)

            for template in existing_templates[len(operacie_payload):]:
                if OperaciaVyroby.objects.filter(operacia_sablona=template).exists():
                    return _api_error(
                        'Nie je možné odobrať operáciu, ktorá sa už používa v existujúcej zákazke. '
                        'Najprv ju dokončite alebo ponechajte v zozname.'
                    )
                template.delete()

            synced_orders, skipped_orders, synced_orders_list = _sync_product_operations_to_open_orders(produkt)

    changed_fields = []
    if changed_produkt:
        changed_fields.extend(changed_produkt)
    if changed_karta:
        changed_fields.extend(changed_karta)
    if operacie_raw:
        changed_fields.append('operacie')

    message = 'Etapa A karta produktu bola uložená.'
    if operacie_raw:
        if skipped_orders:
            message = (
                f'Etapa A karta produktu bola uložená. '
                f'Operácie sa synchronizovali do {synced_orders} otvorených zákaziek, '
                f'{skipped_orders} rozpracované zákazky ostali bez zmeny.'
            )
        else:
            message = f'Etapa A karta produktu bola uložená. Operácie sa synchronizovali do {synced_orders} otvorených zákaziek.'

    return _api_ok(
        message,
        changed_fields=changed_fields,
        synced_orders=synced_orders,
        skipped_orders=skipped_orders,
        synced_orders_list=synced_orders_list,
    )

@login_required
@permission_required("core.add_produktdokument", raise_exception=True)
@require_POST
def upload_vzorkovanie(request, pk):
    """Upload PDF do sekcie 01_Vzorkovanie pre daný produkt."""
    produkt = get_object_or_404(Produkt, pk=pk)
    uploaded = request.FILES.get('subor')
    if not uploaded:
        return _api_error('Súbor nebol odoslaný.')

    ext = uploaded.name.rsplit('.', 1)[-1].lower() if '.' in uploaded.name else ''
    if ext not in {'pdf', 'xlsx', 'xls', 'png', 'jpg', 'jpeg'}:
        return _api_error('Povolené formáty: PDF, XLSX, XLS, PNG, JPG.')

    from django.utils.text import get_valid_filename
    safe_name = get_valid_filename(uploaded.name)
    kategoria = '01_Vzorkovanie'

    existing = ProduktDokument.objects.filter(
        produkt=produkt,
        kategoria=kategoria,
        nazov_suboru=safe_name,
    ).first()
    if existing:
        existing.subor.delete(save=False)
        existing.delete()

    doc = ProduktDokument(
        produkt=produkt,
        kategoria=kategoria,
        subcesta='',
        nazov_suboru=safe_name,
    )
    doc.subor.save(safe_name, uploaded, save=False)
    doc.save()
    xdb_path = _mirror_doc_to_xdatabase(doc)
    if xdb_path:
        doc.povodna_cesta = str(xdb_path)
        doc.save(update_fields=['povodna_cesta'])
    return _api_ok(
        'Súbor bol nahraný.',
        doc_id=doc.pk,
        url=doc.subor.url,
        nazov=safe_name,
        xdb_path=str(xdb_path) if xdb_path else '',
    )


@login_required
@permission_required("core.change_produkt", raise_exception=True)
@require_POST
def uloz_vzorkovanie_meta(request, pk):
    """Uloží číslo a dátum vzorkovania pre produkt."""
    produkt = get_object_or_404(Produkt, pk=pk)
    cislo = (request.POST.get('vzorkovanie') or '').strip()
    datum_raw = (request.POST.get('datum_vzorkovania') or '').strip()

    changed = []
    if cislo != produkt.vzorkovanie:
        produkt.vzorkovanie = cislo
        changed.append('vzorkovanie')

    if datum_raw:
        try:
            from datetime import date as _date
            new_datum = _date.fromisoformat(datum_raw)
            if new_datum != produkt.datum_vzorkovania:
                produkt.datum_vzorkovania = new_datum
                changed.append('datum_vzorkovania')
        except ValueError:
            return _api_error('Neplatný formát dátumu. Očakáva sa YYYY-MM-DD.')
    elif produkt.datum_vzorkovania is not None:
        produkt.datum_vzorkovania = None
        changed.append('datum_vzorkovania')

    if changed:
        produkt.save(update_fields=changed)

    datum_display = produkt.datum_vzorkovania.strftime('%d.%m.%Y') if produkt.datum_vzorkovania else ''
    return _api_ok('Uložené.', vzorkovanie=produkt.vzorkovanie, datum=datum_display)


@login_required
@permission_required("core.change_produkt", raise_exception=True)
@require_POST
def upload_vykres(request, pk):
    """Nahrá výkres produktu do kategórie 02_Drawing."""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        produkt = get_object_or_404(Produkt, pk=pk)
        file = request.FILES.get('subor')
        logger.info(f"upload_vykres: produkt={pk}, file={file}")
        
        if not file:
            logger.warning("No file provided")
            return _api_error('Súbor nie je vybratý.')
        
        safe_name = file.name
        logger.info(f"Creating document with name: {safe_name}")
        
        doc = ProduktDokument(
            produkt=produkt,
            kategoria='02_Drawing',
            subcesta='',
            nazov_suboru=safe_name,
        )
        logger.info(f"Document object created: {doc}")
        
        doc.subor.save(safe_name, file, save=False)
        logger.info(f"File saved to {doc.subor.path if doc.subor else 'NO PATH'}")
        
        doc.save()
        logger.info(f"Document saved: {doc.pk}")
        xdb_path = _mirror_doc_to_xdatabase(doc)
        if xdb_path:
            doc.povodna_cesta = str(xdb_path)
            doc.save(update_fields=['povodna_cesta'])
            logger.info(f"Mirrored to XDatabase: {xdb_path}")
        
        return _api_ok(
            'Výkres bol nahraný.',
            doc_id=doc.pk,
            url=doc.subor.url,
            nazov=safe_name,
            xdb_path=str(xdb_path) if xdb_path else '',
        )
    except Exception as e:
        logger.exception(f"Error in upload_vykres: {e}")
        return _api_error(f"Chyba: {str(e)}")


@login_required
@permission_required("core.change_produkt", raise_exception=True)
@require_POST
def delete_vykres(request, pk, doc_id):
    """Zmaže výkres produktu z DB, media a zrkadla v XDatabase."""
    produkt = get_object_or_404(Produkt, pk=pk)
    doc = get_object_or_404(
        ProduktDokument,
        pk=doc_id,
        produkt=produkt,
        kategoria='02_Drawing',
    )

    deleted_name = doc.nazov_suboru

    # Delete media file first.
    if doc.subor:
        try:
            doc.subor.delete(save=False)
        except Exception:
            logger.exception("Nepodarilo sa zmazať media súbor pre doc_id=%s", doc_id)

    # Delete mirrored file in XDatabase.
    try:
        _delete_doc_from_xdatabase(doc)
    except Exception:
        logger.exception("Nepodarilo sa zmazať XDatabase súbor pre doc_id=%s", doc_id)

    doc.delete()
    return _api_ok('Výkres bol zmazaný.', doc_id=doc_id, nazov=deleted_name)


@login_required
@permission_required("core.change_produkt", raise_exception=True)
@require_POST
def upload_balenie(request, pk):
    """Nahrá balenie produktu do kategórie 05_Manual_Balenie."""
    produkt = get_object_or_404(Produkt, pk=pk)
    file = request.FILES.get('subor')
    if not file:
        return _api_error('Súbor nie je vybratý.')
    
    safe_name = file.name
    
    doc = ProduktDokument(
        produkt=produkt,
        kategoria='05_Manual_Balenie',
        subcesta='',
        nazov_suboru=safe_name,
    )
    doc.subor.save(safe_name, file, save=False)
    doc.save()
    xdb_path = _mirror_doc_to_xdatabase(doc)
    if xdb_path:
        doc.povodna_cesta = str(xdb_path)
        doc.save(update_fields=['povodna_cesta'])
    return _api_ok(
        'Balenie bolo nahrane.',
        doc_id=doc.pk,
        url=doc.subor.url,
        nazov=safe_name,
        xdb_path=str(xdb_path) if xdb_path else '',
    )


@login_required
@permission_required("core.view_objednavka", raise_exception=True)
def plan_vyroby(request):
    """Plán výroby s filtrovaním a vyhľadávaním"""
    search = request.GET.get('search', '')
    stav_filter = request.GET.get('stav', '')
    zakaznik_filter = request.GET.get('zakaznik', '')
    sort_key = (request.GET.get('sort') or 'termin').strip()
    sort_direction = (request.GET.get('direction') or 'asc').strip().lower()

    dnes = timezone.now().date()
    kontrakty = Kontrakt.objects.select_related('produkt')
    if not search:
        kontrakty = kontrakty.filter(
            Q(datum_do__gte=dnes) |
            Q(datum_do__lt=dnes, zostavajuce_mnozstvo__gt=0)
        )
    zakazky = Objednavka.objects.select_related('produkt')
    
    if search:
        kontrakty = kontrakty.filter(
            Q(cislo_kontraktu__icontains=search) | Q(zakaznik__icontains=search) |
            Q(produkt__nazov__icontains=search) | Q(produkt__cislo_dielu__icontains=search)
        )
        zakazky = zakazky.filter(
            Q(cislo_objednavky__icontains=search) | Q(zakaznik__icontains=search) |
            Q(produkt__nazov__icontains=search) | Q(produkt__cislo_dielu__icontains=search)
        )
    
    if stav_filter:
        zakazky = zakazky.filter(stav=stav_filter)
    else:
        zakazky = zakazky.exclude(stav='hotovo')
    
    if zakaznik_filter:
        kontrakty = kontrakty.filter(zakaznik__icontains=zakaznik_filter)
        zakazky = zakazky.filter(zakaznik__icontains=zakaznik_filter)
    
    _sklad_subq = SkladHotovychDielov.objects.filter(
        produkt=OuterRef('produkt')
    ).values('mnozstvo')[:1]
    kontrakty = kontrakty.annotate(
        ks_sklad=Coalesce(Subquery(_sklad_subq, output_field=IntegerField()), 0),
        ks_dodane=F('pocet_kusov_celkovo') - F('zostavajuce_mnozstvo'),
    )
    kontrakty = kontrakty.order_by('datum_do')

    sort_map = {
        'termin': 'datum_pozadovane',
        'zakazka': 'cislo_objednavky',
        'zakaznik': 'zakaznik',
        'produkt': 'produkt__nazov',
        'mnozstvo': 'mnozstvo',
        'vyrobene': 'vyrobene_mnozstvo',
        'stav': 'stav',
    }
    if sort_key not in sort_map:
        sort_key = 'termin'
    if sort_direction not in {'asc', 'desc'}:
        sort_direction = 'asc'

    order_field = sort_map[sort_key]
    if sort_direction == 'desc':
        order_field = f'-{order_field}'
    zakazky = zakazky.order_by(order_field, 'cislo_objednavky')
    
    zakaznici_objednavky = Objednavka.objects.values_list('zakaznik', flat=True).distinct()
    zakaznici_kontrakty = Kontrakt.objects.values_list('zakaznik', flat=True).distinct()
    zakaznici = sorted(set(list(zakaznici_objednavky) + list(zakaznici_kontrakty)))
    expirovane_neukoncene_count = kontrakty.filter(datum_do__lt=dnes, zostavajuce_mnozstvo__gt=0).count()
    
    return render(request, "core/plan.html", {
        "kontrakty": kontrakty, "zakazky": zakazky, "search": search,
        "stav_filter": stav_filter, "zakaznik_filter": zakaznik_filter,
        "sort_key": sort_key, "sort_direction": sort_direction,
        "zakaznici": zakaznici, "dnes": dnes,
        "expirovane_neukoncene_count": expirovane_neukoncene_count,
    })


@login_required
def plan_odovzdania_timeline(request):
    """Graficky timeline prehlad zakazok podla terminu odovzdania."""
    if not request.user.is_authenticated:
        return redirect('admin:login')
    if not (request.user.is_staff or request.user.has_perm('core.view_objednavka')):
        return render(request, 'core/no_access.html')

    search = (request.GET.get('search') or '').strip()
    zakaznik_filter = (request.GET.get('zakaznik') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    only_late = (request.GET.get('only_late') or '').strip() in {'1', 'true', 'on'}

    try:
        days = int(request.GET.get('days') or 30)
    except (TypeError, ValueError):
        days = 30
    days = max(14, min(days, 120))

    dnes = timezone.now().date()
    half_window = days // 2
    window_start = dnes - timedelta(days=half_window)
    window_end = window_start + timedelta(days=days - 1)
    range_days = max((window_end - window_start).days + 1, 1)

    zakazky_qs = (
        Objednavka.objects
        .select_related('produkt')
        .filter(datum_pozadovane__isnull=False)
    )

    if search:
        zakazky_qs = zakazky_qs.filter(
            Q(cislo_objednavky__icontains=search)
            | Q(cislo_objednavky_zakaznika__icontains=search)
            | Q(zakaznik__icontains=search)
            | Q(produkt__nazov__icontains=search)
            | Q(produkt__cislo_dielu__icontains=search)
        )

    if zakaznik_filter:
        zakazky_qs = zakazky_qs.filter(zakaznik__icontains=zakaznik_filter)

    if only_late:
        zakazky_qs = zakazky_qs.exclude(stav='hotovo').filter(datum_pozadovane__lt=dnes)
    else:
        zakazky_qs = zakazky_qs.filter(
            Q(datum_pozadovane__range=(window_start, window_end))
            | Q(datum_pozadovane__lt=dnes, stav__in=['nova', 'vyroba', 'pozastavene'])
        )

    zakazky_qs = zakazky_qs.order_by('datum_pozadovane', 'cislo_objednavky')

    zakaznici = sorted(
        z for z in Objednavka.objects.values_list('zakaznik', flat=True).distinct() if z
    )

    status_priority = {'meska': 0, 'riziko': 1, 'nacas': 2, 'dokoncene': 3}
    timeline_rows = []

    for obj in zakazky_qs:
        ok_kusy = max(int(obj.celkom_ok_kusy or 0), 0)
        ciel_kusy = max(int(obj.mnozstvo or 0), 0)
        zostava_kusy = max(ciel_kusy - ok_kusy, 0)

        if ciel_kusy > 0:
            progres_pct = int(min(100, round((ok_kusy / ciel_kusy) * 100)))
        else:
            progres_pct = 0

        days_to_due = (obj.datum_pozadovane - dnes).days
        start_date = obj.datum_zadania or obj.datum_pozadovane
        total_span = max((obj.datum_pozadovane - start_date).days, 1)
        elapsed = min(max((dnes - start_date).days, 0), total_span)
        expected_pct = int(round((elapsed / total_span) * 100))
        progress_gap = expected_pct - progres_pct

        if obj.stav == 'hotovo' or zostava_kusy == 0:
            timeline_status = 'dokoncene'
            status_label = 'Dokon�ene'
            status_class = 'success'
        elif days_to_due < 0:
            timeline_status = 'meska'
            status_label = 'Me�k�'
            status_class = 'danger'
        elif progress_gap >= 15 or (days_to_due <= 3 and progres_pct < 70):
            timeline_status = 'riziko'
            status_label = 'Riziko'
            status_class = 'warning'
        else:
            timeline_status = 'nacas'
            status_label = 'Na�as'
            status_class = 'primary'

        if status_filter and status_filter != timeline_status:
            continue

        visible_start = max(start_date, window_start)
        visible_end = min(obj.datum_pozadovane, window_end)
        if visible_start > visible_end:
            visible_start = visible_end

        start_offset = max((visible_start - window_start).days, 0)
        end_offset = min(max((visible_end - window_start).days, 0), range_days - 1)

        left_pct = (start_offset / range_days) * 100
        width_pct = max(((end_offset - start_offset + 1) / range_days) * 100, 1.5)

        due_offset = (obj.datum_pozadovane - window_start).days
        due_in_window = 0 <= due_offset <= (range_days - 1)
        due_pct = (due_offset / range_days) * 100 if due_in_window else 0

        timeline_rows.append({
            'obj': obj,
            'ok_kusy': ok_kusy,
            'zostava_kusy': zostava_kusy,
            'progres_pct': progres_pct,
            'days_to_due': days_to_due,
            'status': timeline_status,
            'status_label': status_label,
            'status_class': status_class,
            'left_pct': round(left_pct, 2),
            'width_pct': round(width_pct, 2),
            'due_pct': round(due_pct, 2),
            'due_in_window': due_in_window,
            'expected_pct': expected_pct,
        })

    timeline_rows.sort(
        key=lambda row: (
            status_priority.get(row['status'], 9),
            row['obj'].datum_pozadovane,
            row['obj'].cislo_objednavky,
        )
    )

    summary = {
        'spolu': len(timeline_rows),
        'meska': sum(1 for row in timeline_rows if row['status'] == 'meska'),
        'riziko': sum(1 for row in timeline_rows if row['status'] == 'riziko'),
        'nacas': sum(1 for row in timeline_rows if row['status'] == 'nacas'),
        'dokoncene': sum(1 for row in timeline_rows if row['status'] == 'dokoncene'),
    }

    today_offset = (dnes - window_start).days
    today_pct = (today_offset / range_days) * 100 if 0 <= today_offset <= (range_days - 1) else None

    return render(request, 'core/plan_odovzdania_timeline.html', {
        'timeline_rows': timeline_rows,
        'summary': summary,
        'search': search,
        'zakaznik_filter': zakaznik_filter,
        'status_filter': status_filter,
        'only_late': only_late,
        'days': days,
        'window_start': window_start,
        'window_end': window_end,
        'today': dnes,
        'today_pct': round(today_pct, 2) if today_pct is not None else None,
        'zakaznici': zakaznici,
    })

@login_required
@permission_required("core.view_objednavka", raise_exception=True)
def export_plan_excel(request):
    """Export plánu výroby do Excel"""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Kontrakty"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers1 = ['Číslo kontraktu', 'Zákazník', 'Produkt', 'Číslo dielu', 
                'Celkovo kusov', 'Zostáva dodať', 'Platnosť od', 'Platnosť do', 'Skladom']
    ws1.append(headers1)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    dnes = timezone.now().date()
    kontrakty = Kontrakt.objects.filter(
        Q(datum_do__gte=dnes) |
        Q(datum_do__lt=dnes, zostavajuce_mnozstvo__gt=0)
    ).select_related('produkt').order_by('datum_do')
    
    for kontrakt in kontrakty:
        ws1.append([
            kontrakt.cislo_kontraktu, kontrakt.zakaznik, kontrakt.produkt.nazov,
            kontrakt.produkt.cislo_dielu, kontrakt.pocet_kusov_celkovo,
            kontrakt.zostavajuce_mnozstvo, kontrakt.datum_od.strftime('%d.%m.%Y'),
            kontrakt.datum_do.strftime('%d.%m.%Y'), 'Áno' if kontrakt.je_skladom else 'Nie'
        ])
    
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    ws2 = wb.create_sheet("Zakázky")
    header_fill2 = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    headers2 = ['Číslo zakázky', 'Číslo objednávky zákazníka', 'Zákazník', 'Produkt', 'Číslo dielu', 
                'Množstvo', 'Vyrobené', 'Zostáva', 'Termín', 'Stav', 'Poznámka']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.fill = header_fill2
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    zakazky = Objednavka.objects.exclude(stav="hotovo").select_related('produkt').order_by('datum_pozadovane')
    
    for zakazka in zakazky:
        ws2.append([
            zakazka.cislo_objednavky, zakazka.cislo_objednavky_zakaznika, zakazka.zakaznik, zakazka.produkt.nazov,
            zakazka.produkt.cislo_dielu, zakazka.mnozstvo, zakazka.vyrobene_mnozstvo,
            zakazka.zostava_vyroba(), zakazka.datum_pozadovane.strftime('%d.%m.%Y'),
            zakazka.get_stav_display(), zakazka.poznamka
        ])
    
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="plan_vyroby_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
@permission_required("core.view_objednavka", raise_exception=True)
def detail_zakazky(request, pk):
    """Read-only manažérsky prehľad zakázky."""
    zakazka = get_object_or_404(Objednavka, pk=pk)

    vyrobne_davky = []
    if hasattr(zakazka, 'vyrobna_davka') and zakazka.vyrobna_davka:
        vyrobna_davka = zakazka.vyrobna_davka
        if not vyrobna_davka.operacie.exists():
            vyrobna_davka.vytvor_operacie()
        vyrobne_davky.append(vyrobna_davka)

    operacie = (
        zakazka.operacie
        .select_related('stroj', 'operator')
        .prefetch_related('operatori__operator')
        .order_by('poradie')
    )

    for operacia in operacie:
        operacia.spracovane_celkom = operacia.kusy_spracovane_celkom()
        operacia.ciel_kusy = operacia.cielove_kusy_na_spracovanie()
        if operacia.ciel_kusy > 0:
            operacia.progres_percent = int(min(100, round((operacia.spracovane_celkom / operacia.ciel_kusy) * 100)))
        else:
            operacia.progres_percent = 0
        operacia.dostupne_kusy = operacia.get_dostupne_kusy_na_vstupe()
        operacia.operatori_unikatni = list(dict.fromkeys(
            operacia.operatori.values_list('operator__username', flat=True)
        ))

    kontroly = (
        zakazka.kontroly
        .select_related('operator')
        .prefetch_related(
            Prefetch('merania', queryset=MeraniePriKontrole.objects.select_related('parameter'))
        )
        .order_by('-cas_kontroly')
    )

    for kontrola in kontroly:
        kontrola.merani_spolu = kontrola.merania.count()
        kontrola.merani_nok = sum(1 for meranie in kontrola.merania.all() if not meranie.je_v_tolerancii())

    kvalita_sumar = kontroly.aggregate(
        pocet_kontrol=Count('id'),
        ok_kusy=Sum('pocet_ok_kusov'),
        nok_kusy=Sum('pocet_nok_kusov'),
        nok_zaznamy=Count('id', filter=Q(vysledok_ok=False)),
    )
    kvalita_sumar['ok_kusy'] = kvalita_sumar['ok_kusy'] or 0
    kvalita_sumar['nok_kusy'] = kvalita_sumar['nok_kusy'] or 0
    kvalita_sumar['pocet_kontrol'] = kvalita_sumar['pocet_kontrol'] or 0
    kvalita_sumar['nok_zaznamy'] = kvalita_sumar['nok_zaznamy'] or 0

    context = {
        'zakazka': zakazka,
        'vyrobne_davky': vyrobne_davky,
        'operacie': operacie,
        'kontroly': kontroly,
        'kvalita_sumar': kvalita_sumar,
        'dnes': timezone.now().date(),
    }

    return render(request, 'core/detail_zakazky.html', context)



@never_cache
def home(request):
    """Home dashboard so štatistikami"""
    kiosk_mode = not request.user.is_authenticated

    if request.user.is_authenticated and not request.user.has_perm('core.view_objednavka'):
        if request.user.has_perm('core.view_produkt'):
            return redirect('zoznam_produktov')
        else:
            return render(request, 'core/no_access.html')
    
    dnes = timezone.now().date()
    celkom_zakazok = Objednavka.objects.exclude(stav='hotovo').count()
    vo_vyrobe = Objednavka.objects.filter(stav='vyroba').count()
    nove_zakazky = Objednavka.objects.filter(stav='nova').count()
    pozastavene = Objednavka.objects.filter(stav='pozastavene').count()
    po_termine = Objednavka.objects.exclude(stav='hotovo').filter(datum_pozadovane__lt=dnes).count()
    dnes_termine = Objednavka.objects.exclude(stav='hotovo').filter(datum_pozadovane=dnes).count()
    tento_tyzden_termine = Objednavka.objects.exclude(stav='hotovo').filter(
        datum_pozadovane__gte=dnes, datum_pozadovane__lt=dnes + timedelta(days=7)
    ).count()
    aktivne_kontrakty = Kontrakt.objects.filter(datum_do__gte=dnes).count()
    kontrakt_exspiruje = Kontrakt.objects.filter(
        datum_do__gte=dnes, datum_do__lt=dnes + timedelta(days=30)
    ).count()
    posledne_zakazky = Objednavka.objects.exclude(stav='hotovo').order_by('-datum_zadania')[:5]
    urgentne_zakazky = Objednavka.objects.exclude(stav='hotovo').filter(
        datum_pozadovane__lte=dnes + timedelta(days=3)
    ).order_by('datum_pozadovane')[:5]
    material_pod_minimum = Material.objects.filter(aktualna_zasoba__lt=F('minimalna_zasoba')).count()

    kiosk_users = []
    if kiosk_mode:
        UserModel = get_user_model()
        pin_users = (
            UserModel.objects
            .filter(is_active=True, profile__pin_hash__gt='')
            .select_related('profile')
            .order_by('first_name', 'last_name', 'username')
        )
        for user in pin_users:
            display_name = user.get_full_name().strip() or user.username
            kiosk_users.append({
                'id': user.id,
                'username': user.username,
                'display_name': display_name,
                'avatar_url': _build_user_avatar_url(user, size=180),
            })
    
    return render(request, 'core/home.html', {
        'celkom_zakazok': celkom_zakazok, 'vo_vyrobe': vo_vyrobe,
        'nove_zakazky': nove_zakazky, 'pozastavene': pozastavene,
        'po_termine': po_termine, 'dnes_termine': dnes_termine,
        'tento_tyzden_termine': tento_tyzden_termine,
        'aktivne_kontrakty': aktivne_kontrakty,
        'kontrakt_exspiruje': kontrakt_exspiruje,
        'posledne_zakazky': posledne_zakazky,
        'urgentne_zakazky': urgentne_zakazky,
        'material_pod_minimum': material_pod_minimum,
        'dnes': dnes,
        'kiosk_mode': kiosk_mode,
        'next_url': request.GET.get('next') or reverse('home'),
        'kiosk_users': kiosk_users,
    })


@require_POST
@never_cache
def touch_login(request):
    username = (request.POST.get('username') or '').strip()
    password = request.POST.get('password') or ''
    next_url = request.POST.get('next')

    user = authenticate(request, username=username, password=password)
    if user is None:
        return JsonResponse({
            'status': 'error',
            'message': 'Neplatné meno alebo heslo.',
        }, status=400)

    login(request, user)
    target = _resolve_post_login_redirect(request, next_url)
    return JsonResponse({
        'status': 'ok',
        'message': 'Prihlásenie úspešné.',
        'redirect_url': target,
    })


@require_POST
@never_cache
def touch_pin_login(request):
    user_id_raw = (request.POST.get('user_id') or '').strip()
    pin = (request.POST.get('pin') or '').strip()
    next_url = request.POST.get('next')

    if not user_id_raw.isdigit():
        return JsonResponse({'status': 'error', 'message': 'Vyberte operátora.'}, status=400)

    if not re.fullmatch(r'\d{6}', pin):
        return JsonResponse({'status': 'error', 'message': 'PIN musí mať 6 číslic.'}, status=400)

    UserModel = get_user_model()
    user = (
        UserModel.objects
        .filter(pk=int(user_id_raw), is_active=True)
        .select_related('profile')
        .first()
    )
    if not user:
        return JsonResponse({'status': 'error', 'message': 'Operátor neexistuje.'}, status=404)

    profile = getattr(user, 'profile', None)
    if not profile or not profile.check_pin(pin):
        return JsonResponse({'status': 'error', 'message': 'Neplatný PIN.'}, status=400)

    login(request, user)
    target = _resolve_post_login_redirect(request, next_url)
    return JsonResponse({
        'status': 'ok',
        'message': 'Prihlásenie úspešné.',
        'redirect_url': target,
    })


@login_required
@never_cache
def avatar_selector(request):
    initials = _extract_user_initials(request.user)
    avatar_options = get_avatar_options(initials)
    return render(request, 'core/avatar_selector.html', {
        'avatar_options': avatar_options,
    })


@login_required
@require_POST
@never_cache
def api_avatar_select(request):
    from django.contrib import messages

    selected_color = str(request.POST.get('selected_color') or '').strip().lower().replace('#', '')
    allowed_colors = {str(item.get('color', '')).strip().lower() for item in AVATAR_COLORS}

    if selected_color not in allowed_colors:
        messages.error(request, 'Invalid avatar choice.')
        return redirect('avatar_selector')

    if not _save_avatar_preset_for_user(request.user, selected_color):
        messages.error(request, 'Avatar could not be saved. Try again.')
        return redirect('avatar_selector')

    messages.success(request, 'Avatar saved.')
    return redirect('operator_dashboard')

@never_cache
def attendance_kiosk(request):
    UserModel = get_user_model()

    posledne_zaznamy = []
    for zaznam in DochadzkovyZaznam.objects.select_related('user').order_by('-cas_udalosti', '-id')[:8]:
        local_dt = timezone.localtime(zaznam.cas_udalosti)
        posledne_zaznamy.append({
            'meno': zaznam.user.get_full_name().strip() or zaznam.user.username,
            'typ': zaznam.get_typ_udalosti_display(),
            'cas': local_dt.strftime('%H:%M'),
            'datum': local_dt.strftime('%d.%m.%Y'),
        })

    aktivne_tokeny = DochadzkovyToken.objects.filter(aktivny=True).count()
    dnes = timezone.localdate()
    start, end = _attendance_range_for_day(dnes)
    today_count = DochadzkovyZaznam.objects.filter(cas_udalosti__gte=start, cas_udalosti__lt=end).count()

    operators_qs = (
        UserModel.objects
        .filter(is_active=True, groups__name__iexact='Operatori')
        .select_related('profile')
        .order_by('first_name', 'last_name', 'username')
        .distinct()
    )
    operators = []
    for operator in operators_qs:
        display_name = operator.get_full_name().strip() or operator.username
        initials_source = display_name.split()
        initials = ''.join(part[0] for part in initials_source[:2]).upper() if initials_source else operator.username[:2].upper()
        avatar_url = None
        try:
            if operator.profile.avatar:
                avatar_url = operator.profile.avatar.url
        except Exception:
            avatar_url = None
        operators.append({
            'username': operator.username,
            'display_name': display_name,
            'initials': initials,
            'avatar_url': avatar_url,
        })

    is_attendance_manager = bool(
        request.user.is_authenticated
        and (
            request.user.is_staff
            or request.user.groups.filter(name='attendance_manager').exists()
            or request.user.has_perm('core.view_dochadzkovyzaznam')
        )
    )

    return render(request, 'core/attendance_kiosk.html', {
        'aktivne_tokeny': aktivne_tokeny,
        'today_count': today_count,
        'dnes': dnes,
        'posledne_zaznamy': posledne_zaznamy,
        'operators': operators,
        'is_attendance_manager': is_attendance_manager,
    })


@require_POST
@never_cache
def attendance_punch(request):
    rate_limit_response = _check_operator_rate_limit(request, 'attendance-punch', limit=10, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    identifikator = _normalize_dochadzka_identifikator(request.POST.get('identifikator'))
    pin = (request.POST.get('pin') or '').strip()

    if not identifikator:
        return _api_error('Zadajte identifikátor karty alebo kód.')

    if not re.fullmatch(r'\d{6}', pin):
        return _api_error('PIN musí mať 6 číslic.')

    # Najprv hľadáme DochadzkovyToken
    token = (
        DochadzkovyToken.objects
        .filter(identifikator=identifikator, aktivny=True, user__is_active=True)
        .select_related('user', 'user__profile')
        .first()
    )

    user = None
    if not token:
        # Ak nenájdeme token, skúsíme hľadať používateľa podľa username (prihlasovacieho mena)
        UserModel = get_user_model()
        user = UserModel.objects.filter(username__iexact=identifikator.lower(), is_active=True).select_related('profile').first()
        if not user:
            return _api_error('Identifikátor alebo prihlasovací meno neexistuje alebo nie je aktívne.')
    else:
        user = token.user

    profile = getattr(user, 'profile', None)
    if not profile or not profile.check_pin(pin):
        return _api_error('Neplatný PIN.')

    now = timezone.now()
    posledny_zaznam = (
        DochadzkovyZaznam.objects
        .filter(user=user)
        .order_by('-cas_udalosti', '-id')
        .first()
    )

    if posledny_zaznam and (now - posledny_zaznam.cas_udalosti) <= timedelta(seconds=10):
        return _api_error('Posledný záznam bol vytvorený pred chvíľou. Skúste to znova o niekoľko sekúnd.')

    typ_udalosti = 'OUT' if posledny_zaznam and posledny_zaznam.typ_udalosti == 'IN' else 'IN'

    with transaction.atomic():
        zaznam = DochadzkovyZaznam.objects.create(
            user=user,
            token=token,
            typ_udalosti=typ_udalosti,
            zdroj='KIOSK',
            cas_udalosti=now,
        )
        # Ak bol použitý token (RFID/NFC), aktualizujeme čas posledného použitia
        if token:
            token.posledne_pouzitie = now
            token.save(update_fields=['posledne_pouzitie', 'updated_at'])

    local_dt = timezone.localtime(zaznam.cas_udalosti)
    display_name = user.get_full_name().strip() or user.username
    return _api_ok(
        f'{display_name}: {zaznam.get_typ_udalosti_display()} zaznamenaný.',
        typ_udalosti=zaznam.typ_udalosti,
        typ_udalosti_label=zaznam.get_typ_udalosti_display(),
        cas_udalosti=local_dt.isoformat(),
        display_time=local_dt.strftime('%H:%M'),
        display_date=local_dt.strftime('%d.%m.%Y'),
        user_name=display_name,
    )


def _attendance_authenticate_by_identifier(identifikator_raw, pin_raw):
    identifikator = _normalize_dochadzka_identifikator(identifikator_raw)
    pin = (pin_raw or '').strip()
    if not identifikator:
        return None, None, 'Zadajte identifikátor.'
    if not re.fullmatch(r'\d{6}', pin):
        return None, None, 'PIN musí mať 6 číslic.'

    token = (
        DochadzkovyToken.objects
        .filter(identifikator=identifikator, aktivny=True, user__is_active=True)
        .select_related('user', 'user__profile')
        .first()
    )
    user = None
    if token:
        user = token.user
    else:
        UserModel = get_user_model()
        user = (
            UserModel.objects
            .filter(username__iexact=identifikator.lower(), is_active=True)
            .select_related('profile')
            .first()
        )
        if not user:
            return None, None, 'Token ani používateľ neexistuje alebo nie je aktívny.'

    profile = getattr(user, 'profile', None)
    if not profile or not profile.check_pin(pin):
        return None, None, 'Neplatný PIN.'

    return user, token, None


@require_POST
@never_cache
def attendance_operator_session(request):
    user, token, error = _attendance_authenticate_by_identifier(
        request.POST.get('identifikator'),
        request.POST.get('pin'),
    )
    if error:
        return _api_error(error)

    rows, month_label = _kiosk_month_rows_for_user(user)

    return _api_ok(
        'Relácia načítaná.',
        username=user.username,
        user_name=user.get_full_name().strip() or user.username,
        token_id=token.id if token else None,
        month=month_label,
        rows=rows,
    )


@require_POST
@never_cache
def attendance_manual_punch(request):
    user, token, error = _attendance_authenticate_by_identifier(
        request.POST.get('identifikator'),
        request.POST.get('pin'),
    )
    if error:
        return _api_error(error)

    event_type = (request.POST.get('event_type') or '').strip().upper()
    if event_type not in {'IN', 'OUT'}:
        return _api_error('Neplatný typ udalosti.')

    now = timezone.now()
    is_valid, message = _validate_attendance_sequence(user, event_type, now)
    if not is_valid:
        return _api_error(message)

    record = DochadzkovyZaznam.objects.create(
        user=user,
        token=token,
        typ_udalosti=event_type,
        zdroj='KIOSK',
        cas_udalosti=now,
    )
    if token:
        token.posledne_pouzitie = now
        token.save(update_fields=['posledne_pouzitie', 'updated_at'])

    rows, month_label = _kiosk_month_rows_for_user(user, timezone.localtime(record.cas_udalosti).date())
    local_dt = timezone.localtime(record.cas_udalosti)

    return _api_ok(
        'Ručný záznam vytvorený.',
        typ_udalosti=record.typ_udalosti,
        typ_udalosti_label=record.get_typ_udalosti_display(),
        display_time=local_dt.strftime('%H:%M'),
        month=month_label,
        rows=rows,
    )


@require_POST
@never_cache
def attendance_vacation_request(request):
    user, _token, error = _attendance_authenticate_by_identifier(
        request.POST.get('identifikator'),
        request.POST.get('pin'),
    )
    if error:
        return _api_error(error)

    date_from = _parse_iso_date(request.POST.get('date_from'), None)
    date_to = _parse_iso_date(request.POST.get('date_to'), None)
    note = (request.POST.get('note') or '').strip()

    if not date_from or not date_to:
        return _api_error('Dátum od/do je povinný.')
    if date_from > date_to:
        return _api_error('Neplatný rozsah dátumu.')

    pracovny_den_existuje = any((date_from + timedelta(days=offset)).weekday() < 5 for offset in range((date_to - date_from).days + 1))
    if not pracovny_den_existuje:
        return _api_error('Rozsah musí obsahovať aspoň jeden pracovný deň.')

    DovolenkaZiadost.objects.create(
        user=user,
        date_from=date_from,
        date_to=date_to,
        note=note,
        status=DovolenkaZiadost.STATUS_PENDING,
    )
    rows, month_label = _kiosk_month_rows_for_user(user)
    return _api_ok('Žiadosť o dovolenku bola odoslaná.', month=month_label, rows=rows)


@login_required
def attendance_vacation_manager_page(request):
    if not (request.user.is_staff or request.user.groups.filter(name='attendance_manager').exists()):
        return HttpResponse('Forbidden', status=403)

    pending_requests = DovolenkaZiadost.objects.filter(
        status=DovolenkaZiadost.STATUS_PENDING,
    ).select_related('user').order_by('requested_at')

    return render(request, 'core/attendance_vacation_manager.html', {
        'pending_requests': pending_requests,
    })


@login_required
@require_POST
def attendance_vacation_decide(request, request_id):
    if not (request.user.is_staff or request.user.groups.filter(name='attendance_manager').exists()):
        return _api_error('Nemáte oprávnenie.')

    request_obj = get_object_or_404(DovolenkaZiadost, pk=request_id)
    decision = (request.POST.get('decision') or '').strip().lower()
    decision_note = (request.POST.get('decision_note') or '').strip()

    if decision == 'approve':
        request_obj.status = DovolenkaZiadost.STATUS_APPROVED
    elif decision == 'reject':
        request_obj.status = DovolenkaZiadost.STATUS_REJECTED
    else:
        return _api_error('Neplatné rozhodnutie.')

    request_obj.decided_by = request.user
    request_obj.decided_at = timezone.now()
    request_obj.decision_note = decision_note
    request_obj.save(update_fields=['status', 'decided_by', 'decided_at', 'decision_note', 'updated_at'])

    return _api_ok('Rozhodnutie uložené.', status_value=request_obj.status)


@login_required
@permission_required('core.view_dochadzkovyzaznam', raise_exception=True)
def attendance_manager_overview(request):
    from django.contrib import messages

    year, month = _parse_iso_month(request.GET.get('month') or request.POST.get('return_month'), timezone.localdate())
    selected_month = f'{year:04d}-{month:02d}'
    selected_user_id = (request.GET.get('user_id') or request.POST.get('return_user_id') or '').strip()

    users = (
        get_user_model().objects
        .filter(is_active=True)
        .filter(Q(dochadzkove_tokeny__isnull=False) | Q(dochadzkove_zaznamy__isnull=False))
        .distinct()
        .order_by('first_name', 'last_name', 'username')
    )

    if request.method == 'POST' and request.user.has_perm('core.add_dochadzkovyzaznam'):
        edit_user_id = (request.POST.get('edit_user_id') or '').strip()
        edit_date = _parse_iso_date(request.POST.get('edit_date'), None)
        action = (request.POST.get('action') or '').strip()
        arrival = (request.POST.get('edit_arrival') or '').strip()
        departure = (request.POST.get('edit_departure') or '').strip()
        note = (request.POST.get('edit_note') or '').strip()

        if not edit_user_id.isdigit() or not edit_date:
            messages.error(request, '❌ Neplatné vstupy pre ručnú úpravu dochádzky.')
        else:
            user_obj = get_user_model().objects.filter(pk=int(edit_user_id), is_active=True).first()
            if not user_obj:
                messages.error(request, '❌ Zamestnanec neexistuje alebo nie je aktívny.')
            else:
                start, end = _attendance_range_for_day(edit_date)
                day_qs = DochadzkovyZaznam.objects.filter(user=user_obj, cas_udalosti__gte=start, cas_udalosti__lt=end)

                if action == 'delete_day':
                    deleted_count, _ = day_qs.delete()
                    messages.success(request, f'✅ Dochádzka bola zmazaná ({deleted_count} záznamov).')
                elif action == 'update_day':
                    tz = timezone.get_current_timezone()
                    day_qs.delete()

                    created = 0
                    if arrival:
                        try:
                            in_dt = datetime.strptime(f'{edit_date.isoformat()} {arrival}', '%Y-%m-%d %H:%M')
                            in_aware = timezone.make_aware(in_dt, tz)
                            DochadzkovyZaznam.objects.create(
                                user=user_obj,
                                typ_udalosti='IN',
                                zdroj='ADMIN',
                                cas_udalosti=in_aware,
                                zaznamenal=request.user,
                                poznamka=note,
                            )
                            created += 1
                        except ValueError:
                            messages.error(request, '❌ Neplatný čas príchodu.')

                    if departure:
                        try:
                            out_dt = datetime.strptime(f'{edit_date.isoformat()} {departure}', '%Y-%m-%d %H:%M')
                            out_aware = timezone.make_aware(out_dt, tz)
                            DochadzkovyZaznam.objects.create(
                                user=user_obj,
                                typ_udalosti='OUT',
                                zdroj='ADMIN',
                                cas_udalosti=out_aware,
                                zaznamenal=request.user,
                                poznamka=note,
                            )
                            created += 1
                        except ValueError:
                            messages.error(request, '❌ Neplatný čas odchodu.')

                    if created:
                        messages.success(request, '✅ Denná dochádzka bola uložená.')
                else:
                    messages.error(request, '❌ Neplatná akcia úpravy.')

        params = {'month': selected_month}
        if selected_user_id:
            params['user_id'] = selected_user_id
        return redirect(f"{reverse('attendance_manager_overview')}?{urlencode(params)}")

    start, end = _attendance_range_for_month(year, month)
    queryset = DochadzkovyZaznam.objects.filter(cas_udalosti__gte=start, cas_udalosti__lt=end)
    if selected_user_id.isdigit():
        queryset = queryset.filter(user_id=int(selected_user_id))

    summary_rows = _build_dochadzka_summary_rows(queryset)

    total_days = len(summary_rows)
    total_open_days = sum(1 for row in summary_rows if row.get('stav') == 'Otvorená dochádzka')
    total_duration_minutes = 0
    for row in summary_rows:
        duration = row.get('trvanie') or ''
        if ':' in duration:
            h_raw, m_raw = duration.split(':', 1)
            if h_raw.isdigit() and m_raw.isdigit():
                total_duration_minutes += int(h_raw) * 60 + int(m_raw)

    duration_hours, duration_minutes = divmod(total_duration_minutes, 60)
    total_duration_text = f'{duration_hours:02d}:{duration_minutes:02d}'

    user_stats = {}
    for row in summary_rows:
        key = row['user_id']
        stat = user_stats.setdefault(key, {
            'name': row['meno'],
            'days': 0,
            'closed_days': 0,
            'open_days': 0,
            'minutes': 0,
        })
        stat['days'] += 1
        if row.get('stav') == 'Uzavreté':
            stat['closed_days'] += 1
        if row.get('stav') == 'Otvorená dochádzka':
            stat['open_days'] += 1
        duration = row.get('trvanie') or ''
        if ':' in duration:
            h_raw, m_raw = duration.split(':', 1)
            if h_raw.isdigit() and m_raw.isdigit():
                stat['minutes'] += int(h_raw) * 60 + int(m_raw)

    user_stats_rows = []
    for stat in user_stats.values():
        h_val, m_val = divmod(stat['minutes'], 60)
        user_stats_rows.append({
            'name': stat['name'],
            'days': stat['days'],
            'closed_days': stat['closed_days'],
            'open_days': stat['open_days'],
            'duration_text': f'{h_val:02d}:{m_val:02d}',
        })
    user_stats_rows.sort(key=lambda item: item['name'])

    vacation_qs = DovolenkaZiadost.objects.filter(
        date_from__lt=end.date(),
        date_to__gte=start.date(),
    ).select_related('user').order_by('date_from', 'user__username')
    if selected_user_id.isdigit():
        vacation_qs = vacation_qs.filter(user_id=int(selected_user_id))

    def _workdays_inclusive(day_from, day_to):
        total = 0
        current = day_from
        while current <= day_to:
            if current.weekday() < 5:
                total += 1
            current += timedelta(days=1)
        return total

    vacation_requests = []
    for item in vacation_qs:
        status_badge = 'warning'
        if item.status == DovolenkaZiadost.STATUS_APPROVED:
            status_badge = 'success'
        elif item.status == DovolenkaZiadost.STATUS_REJECTED:
            status_badge = 'danger'
        vacation_requests.append({
            'user_name': item.user.get_full_name().strip() or item.user.username,
            'date_from': item.date_from,
            'date_to': item.date_to,
            'workdays': _workdays_inclusive(item.date_from, item.date_to),
            'status_label': item.get_status_display(),
            'status_badge': status_badge,
            'note': item.note,
            'decision_note': item.decision_note,
        })

    vacation_approved_count = sum(1 for item in vacation_qs if item.status == DovolenkaZiadost.STATUS_APPROVED)
    vacation_not_approved_count = vacation_qs.count() - vacation_approved_count

    return render(request, 'core/attendance_manager_overview.html', {
        'selected_month': selected_month,
        'selected_user_id': selected_user_id,
        'users': users,
        'summary_rows': summary_rows,
        'total_days': total_days,
        'total_open_days': total_open_days,
        'total_duration_text': total_duration_text,
        'user_stats_rows': user_stats_rows,
        'vacation_requests': vacation_requests,
        'vacation_approved_count': vacation_approved_count,
        'vacation_not_approved_count': vacation_not_approved_count,
    })


@login_required
@permission_required('core.view_dochadzkovyzaznam', raise_exception=True)
def attendance_report(request):
    selected_date = _parse_iso_date(request.GET.get('date') or request.POST.get('return_date'), timezone.localdate())
    selected_user_id = (request.GET.get('user_id') or request.POST.get('return_user_id') or '').strip()

    if request.method == 'POST':
        from django.contrib import messages

        if not request.user.has_perm('core.add_dochadzkovyzaznam'):
            messages.error(request, '❌ Nemáte oprávnenie na ručný záznam dochádzky.')
        else:
            user_id_raw = (request.POST.get('manual_user_id') or '').strip()
            typ_udalosti = (request.POST.get('manual_typ_udalosti') or '').strip().upper()
            cas_raw = (request.POST.get('manual_cas_udalosti') or '').strip()
            poznamka = (request.POST.get('manual_poznamka') or '').strip()

            if not user_id_raw.isdigit():
                messages.error(request, '❌ Vyberte platného zamestnanca.')
            elif typ_udalosti not in {'IN', 'OUT'}:
                messages.error(request, '❌ Neplatný typ udalosti.')
            else:
                UserModel = get_user_model()
                user_obj = UserModel.objects.filter(pk=int(user_id_raw), is_active=True).first()
                if not user_obj:
                    messages.error(request, '❌ Zamestnanec neexistuje alebo nie je aktívny.')
                else:
                    cas_udalosti = timezone.now()
                    if cas_raw:
                        try:
                            parsed = datetime.strptime(cas_raw, '%Y-%m-%dT%H:%M')
                            cas_udalosti = timezone.make_aware(parsed, timezone.get_current_timezone())
                        except ValueError:
                            messages.error(request, '❌ Neplatný dátum/čas. Použite správny formát.')
                            params = {'date': selected_date.isoformat()}
                            if selected_user_id:
                                params['user_id'] = selected_user_id
                            return redirect(f"{reverse('attendance_report')}?{urlencode(params)}")

                    is_valid, validation_message = _validate_attendance_sequence(user_obj, typ_udalosti, cas_udalosti)
                    if not is_valid:
                        messages.error(request, f'❌ {validation_message}')
                        params = {'date': selected_date.isoformat()}
                        if selected_user_id:
                            params['user_id'] = selected_user_id
                        return redirect(f"{reverse('attendance_report')}?{urlencode(params)}")

                    DochadzkovyZaznam.objects.create(
                        user=user_obj,
                        typ_udalosti=typ_udalosti,
                        zdroj='ADMIN',
                        cas_udalosti=cas_udalosti,
                        zaznamenal=request.user,
                        poznamka=poznamka,
                    )
                    messages.success(request, f'✅ Ručný záznam {typ_udalosti} bol vytvorený pre {user_obj.username}.')

        params = {'date': selected_date.isoformat()}
        if selected_user_id:
            params['user_id'] = selected_user_id
        return redirect(f"{reverse('attendance_report')}?{urlencode(params)}")

    start, end = _attendance_range_for_day(selected_date)

    queryset = DochadzkovyZaznam.objects.filter(cas_udalosti__gte=start, cas_udalosti__lt=end)
    if selected_user_id.isdigit():
        queryset = queryset.filter(user_id=int(selected_user_id))

    users = (
        get_user_model().objects
        .filter(is_active=True)
        .filter(Q(dochadzkove_tokeny__isnull=False) | Q(dochadzkove_zaznamy__isnull=False))
        .distinct()
        .order_by('first_name', 'last_name', 'username')
    )

    all_active_users = get_user_model().objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')

    summary_rows = _build_dochadzka_summary_rows(queryset)
    return render(request, 'core/attendance_report.html', {
        'selected_date': selected_date,
        'selected_user_id': selected_user_id,
        'summary_rows': summary_rows,
        'users': users,
        'all_active_users': all_active_users,
        'can_manual_entry': request.user.has_perm('core.add_dochadzkovyzaznam'),
    })


@login_required
@permission_required('core.view_dochadzkovyzaznam', raise_exception=True)
def attendance_export_csv(request):
    year, month = _parse_iso_month(request.GET.get('month'), timezone.localdate())
    selected_user_id = (request.GET.get('user_id') or '').strip()
    start, end = _attendance_range_for_month(year, month)

    queryset = DochadzkovyZaznam.objects.filter(cas_udalosti__gte=start, cas_udalosti__lt=end)
    if selected_user_id.isdigit():
        queryset = queryset.filter(user_id=int(selected_user_id))

    summary_rows = _build_dochadzka_summary_rows(queryset)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="dochadzka-{year:04d}-{month:02d}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Datum', 'Zamestnanec', 'Prichod', 'Odchod', 'Trvanie', 'Stav', 'Pocet udalosti', 'Udalosti'])

    for row in summary_rows:
        writer.writerow([
            row['datum'].strftime('%Y-%m-%d'),
            row['meno'],
            row['prichod_text'],
            row['odchod_text'],
            row['trvanie'],
            row['stav'],
            row['pocet_udalosti'],
            row['udalosti_text'],
        ])

    return response


@login_required
@permission_required('core.view_dochadzkovyzaznam', raise_exception=True)
def attendance_export_manager_pdf(request):
    # Keep the manager UI flow functional even when PDF generator is not available.
    params = {'month': request.GET.get('month') or ''}
    selected_user_id = (request.GET.get('user_id') or '').strip()
    if selected_user_id:
        params['user_id'] = selected_user_id
    return redirect(f"{reverse('attendance_export_csv')}?{urlencode(params)}")


@login_required
@permission_required("core.view_kontrolakvality", raise_exception=True)
def kvalita_dashboard(request):
    """Manažérske vyhodnotenie kvality výroby."""
    dnes = timezone.now().date()
    default_od = dnes - timedelta(days=30)

    datum_od = request.GET.get('od') or default_od.isoformat()
    datum_do = request.GET.get('do') or dnes.isoformat()
    typ_kontroly = request.GET.get('typ', '')
    operator_id = request.GET.get('operator', '')

    kontroly = KontrolaKvality.objects.select_related('objednavka', 'operator').order_by('-cas_kontroly')

    try:
        od_date = datetime.strptime(datum_od, '%Y-%m-%d').date()
        kontroly = kontroly.filter(cas_kontroly__date__gte=od_date)
    except ValueError:
        od_date = default_od
        kontroly = kontroly.filter(cas_kontroly__date__gte=od_date)

    try:
        do_date = datetime.strptime(datum_do, '%Y-%m-%d').date()
        kontroly = kontroly.filter(cas_kontroly__date__lte=do_date)
    except ValueError:
        do_date = dnes
        kontroly = kontroly.filter(cas_kontroly__date__lte=do_date)

    if typ_kontroly in {'PRIEBEZNA', 'FINALNA'}:
        kontroly = kontroly.filter(typ_kontroly=typ_kontroly)

    if operator_id.isdigit():
        kontroly = kontroly.filter(operator_id=int(operator_id))

    suma = kontroly.aggregate(
        ok_sum=Sum('pocet_ok_kusov'),
        nok_sum=Sum('pocet_nok_kusov'),
        nok_zaznamy=Count('id', filter=Q(vysledok_ok=False)),
    )
    ok_sum = suma['ok_sum'] or 0
    nok_sum = suma['nok_sum'] or 0
    spolu_kusy = ok_sum + nok_sum
    nok_percento = round((nok_sum / spolu_kusy) * 100, 2) if spolu_kusy else 0

    top_nok_zakazky = list(
        kontroly.values(
            'objednavka__id',
            'objednavka__cislo_objednavky',
            'objednavka__zakaznik',
        )
        .annotate(
            nok_sum=Sum('pocet_nok_kusov'),
            ok_sum=Sum('pocet_ok_kusov'),
            zaznamy=Count('id'),
        )
        .order_by('-nok_sum', '-zaznamy')[:5]
    )

    trend_data = []
    trend_qs = (
        kontroly.annotate(den=F('pocet_ok_kusov') + F('pocet_nok_kusov'))
        .annotate(den_datum=TruncDate('cas_kontroly'))
        .values('den_datum')
        .annotate(
            ok_sum=Sum('pocet_ok_kusov'),
            nok_sum=Sum('pocet_nok_kusov'),
            spolu=Sum('den'),
        )
        .order_by('den_datum')
    )
    for row in trend_qs:
        den_spolu = row['spolu'] or 0
        trend_data.append({
            'datum': row['den_datum'],
            'ok_sum': row['ok_sum'] or 0,
            'nok_sum': row['nok_sum'] or 0,
            'nok_percento': round(((row['nok_sum'] or 0) / den_spolu) * 100, 2) if den_spolu else 0,
        })

    operatori = (
        KontrolaKvality.objects.exclude(operator__isnull=True)
        .values('operator__id', 'operator__username')
        .distinct()
        .order_by('operator__username')
    )
    context = {
        'kontroly': kontroly[:100],
        'operatori': operatori,
        'top_nok_zakazky': top_nok_zakazky,
        'trend_data': trend_data,
        'filtre': {
            'od': od_date.isoformat(),
            'do': do_date.isoformat(),
            'typ': typ_kontroly,
            'operator': operator_id,
        },
        'kpi': {
            'pocet_kontrol': kontroly.count(),
            'ok_kusy': ok_sum,
            'nok_kusy': nok_sum,
            'nok_zaznamy': suma['nok_zaznamy'] or 0,
            'nok_percento': nok_percento,
        },
    }
    return render(request, 'core/kvalita_dashboard.html', context)


@login_required
@permission_required("core.view_kaliber", raise_exception=True)
def kvalita_kalibre(request):
    dnes = timezone.localdate()
    kalibre_rows = []
    for k in Kaliber.objects.all().order_by('cislo', 'id'):
        if k.kontrola_platna_do:
            zostava_dni = (k.kontrola_platna_do - dnes).days
            if zostava_dni < 0:
                platnost_stav = 'expired'
                platnost_text = 'Po termíne'
                badge_class = 'bg-danger'
            elif zostava_dni <= 30:
                platnost_stav = 'warning'
                platnost_text = f'Končí o {zostava_dni} dní'
                badge_class = 'bg-warning text-dark'
            else:
                platnost_stav = 'ok'
                platnost_text = f'Platný ({zostava_dni} dní)'
                badge_class = 'bg-success'
        else:
            platnost_stav = 'unknown'
            platnost_text = 'Bez termínu platnosti'
            badge_class = 'bg-secondary'

        kalibre_rows.append({
            'obj': k,
            'platnost_stav': platnost_stav,
            'platnost_text': platnost_text,
            'badge_class': badge_class,
        })

    return render(
        request,
        'core/kvalita_kalibre.html',
        {
            'kalibre_rows': kalibre_rows,
        },
    )

@login_required
@permission_required("core.view_stroj", raise_exception=True)
def zoznam_strojov(request):
    stroje = Stroj.objects.all().order_by("nazov")
    from .models import OperaciaVyroby

    interval = request.GET.get('interval') or request.session.get('stroje_interval', 'dni')
    if interval not in {'dni', 'tyzdne', 'mesiace'}:
        interval = 'dni'

    rozsah = request.GET.get('rozsah') or request.session.get('stroje_rozsah', '30')
    if rozsah not in {'7', '30', '90'}:
        rozsah = '30'
    rozsah_dni = int(rozsah)

    request.session['stroje_interval'] = interval
    request.session['stroje_rozsah'] = rozsah

    now = timezone.now()
    today = now.date()

    def add_months(base_date, months):
        year = base_date.year + (base_date.month - 1 + months) // 12
        month = (base_date.month - 1 + months) % 12 + 1
        return date(year, month, 1)

    if interval == 'dni':
        start_day = today - timedelta(days=rozsah_dni - 1)
        bucket_keys = [start_day + timedelta(days=i) for i in range(rozsah_dni)]
        bucket_labels = [day.strftime('%d.%m.') for day in bucket_keys]
        range_start_dt = timezone.make_aware(datetime.combine(start_day, datetime.min.time()))
    elif interval == 'tyzdne':
        pocet_tyzdnov = max(1, (rozsah_dni + 6) // 7)
        current_week_start = today - timedelta(days=today.weekday())
        first_week_start = current_week_start - timedelta(weeks=pocet_tyzdnov - 1)
        bucket_keys = [first_week_start + timedelta(weeks=i) for i in range(pocet_tyzdnov)]
        bucket_labels = [f"{week_start.strftime('%d.%m.')}" for week_start in bucket_keys]
        range_start_dt = timezone.make_aware(datetime.combine(first_week_start, datetime.min.time()))
    else:
        pocet_mesiacov = max(1, (rozsah_dni + 29) // 30)
        current_month_start = today.replace(day=1)
        first_month_start = add_months(current_month_start, -(pocet_mesiacov - 1))
        bucket_keys = [add_months(first_month_start, i) for i in range(pocet_mesiacov)]
        bucket_labels = [month_start.strftime('%m/%Y') for month_start in bucket_keys]
        range_start_dt = timezone.make_aware(datetime.combine(first_month_start, datetime.min.time()))

    range_total_hours = max(1.0, (now - range_start_dt).total_seconds() / 3600)
    machine_count = max(1, stroje.count())

    trend_hours_by_bucket = {key: 0.0 for key in bucket_keys}
    machine_hours = {stroj.nazov: 0.0 for stroj in stroje}

    sessions = OperatorNaOperacii.objects.filter(
        operacia__stroj__isnull=False,
        cas_zaciatku__lte=now,
    ).filter(
        Q(cas_konca__isnull=True) | Q(cas_konca__gte=range_start_dt)
    ).select_related('operacia__stroj')

    for session in sessions:
        start_dt = session.cas_zaciatku if session.cas_zaciatku > range_start_dt else range_start_dt
        end_dt = session.cas_konca if session.cas_konca and session.cas_konca < now else now
        if end_dt <= start_dt:
            continue

        worked_hours = (end_dt - start_dt).total_seconds() / 3600
        stroj_nazov = session.operacia.stroj.nazov
        machine_hours[stroj_nazov] = machine_hours.get(stroj_nazov, 0.0) + worked_hours

        bucket_date = start_dt.date()
        if interval == 'dni':
            bucket_key = bucket_date
        elif interval == 'tyzdne':
            bucket_key = bucket_date - timedelta(days=bucket_date.weekday())
        else:
            bucket_key = bucket_date.replace(day=1)

        if bucket_key in trend_hours_by_bucket:
            trend_hours_by_bucket[bucket_key] += worked_hours

    trend_utilization = []
    for bucket_key in bucket_keys:
        bucket_hours = trend_hours_by_bucket.get(bucket_key, 0.0)
        if interval == 'dni':
            capacity_hours = machine_count * 24
        elif interval == 'tyzdne':
            capacity_hours = machine_count * 168
        else:
            next_month = add_months(bucket_key, 1)
            days_in_month = (next_month - bucket_key).days
            capacity_hours = machine_count * days_in_month * 24

        utilization_pct = (bucket_hours / capacity_hours) * 100 if capacity_hours > 0 else 0
        trend_utilization.append(round(utilization_pct, 1))

    machine_names = []
    machine_utilization = []
    for machine_name, hours in sorted(machine_hours.items(), key=lambda item: item[1], reverse=True):
        machine_names.append(machine_name)
        machine_utilization.append(round((hours / range_total_hours) * 100, 1))

    for s in stroje:
        operacie = OperaciaVyroby.objects.filter(
            stroj=s,
            stav__in=['caka', 'vyroba', 'pozastavena'],
            objednavka__stav__in=['nova', 'vyroba', 'pozastavene']
        )

        rezervacia_min = 0.0
        for op in operacie:
            zostava = op.get_dostupne_kusy_na_vstupe()
            if zostava <= 0:
                continue
            cas_pripravy = 0
            if op.stav == 'caka' and op.vyrobene_kusy == 0 and op.nepodarky == 0:
                cas_pripravy = op.cas_pripravy
            rezervacia_min += (float(op.cas_kus) * zostava) + cas_pripravy

        s.rezervacia_hodin = round(rezervacia_min / 60, 1)
        s.rezervacia_percent = int(min(100, round((s.rezervacia_hodin / 168) * 100)))
        
        # Pridaj absolútnu hodnotu dní do servisu
        if s.dni_do_servisu is not None and s.dni_do_servisu < 0:
            s.dni_po_termine = abs(s.dni_do_servisu)
        else:
            s.dni_po_termine = None

    interval_label = {
        'dni': 'Dni',
        'tyzdne': 'Týždne',
        'mesiace': 'Mesiace',
    }[interval]

    graf_data = {
        'labels': bucket_labels,
        'trend': trend_utilization,
        'machine_labels': machine_names,
        'machine_values': machine_utilization,
        'interval_label': interval_label,
        'range_label': f'Posledných {rozsah_dni} dní',
    }

    return render(request, "core/zoznam_strojov.html", {
        "stroje": stroje,
        "interval": interval,
        "rozsah": rozsah,
        "graf_data_json": json.dumps(graf_data),
    })



@login_required
def operator_dashboard(request):
    def _normalize_unit(value):
        normalized = unicodedata.normalize('NFD', str(value or '').strip().lower())
        return ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')

    def _is_kg_unit(value):
        return _normalize_unit(value) == 'kg'

    def _is_bar_unit(value):
        normalized = _normalize_unit(value)
        return normalized in {'tyc', 'tyce', 'ks'}

    now = timezone.now()
    last_30_days = now - timedelta(days=30)

    rozpracovane = Objednavka.objects.filter(
        stav__in=['vyroba', 'pozastavene']
    ).filter(
        Q(priradeni_operatori=request.user)
        | Q(zaznamy__operator=request.user, zaznamy__typ_udalosti='START')
    ).select_related('produkt').only(
        'id', 'cislo_objednavky', 'mnozstvo', 'stav', 'produkt__nazov'
    ).distinct()

    for obj in rozpracovane:
        total = obj.mnozstvo or 0
        ok_kusy = obj.celkom_ok_kusy or 0
        obj.progress_pct = int(round((ok_kusy / total) * 100)) if total > 0 else 0

    nove_priradene = Objednavka.objects.filter(
        stav='nova', priradeni_operatori=request.user
    ).select_related('produkt').only(
        'id', 'cislo_objednavky', 'mnozstvo', 'datum_pozadovane', 'produkt__nazov'
    ).order_by('datum_pozadovane')

    nove_dostupne = Objednavka.objects.filter(
        stav='nova'
    ).exclude(
        priradeni_operatori__isnull=False
    ).select_related('produkt').only(
        'id', 'cislo_objednavky', 'mnozstvo', 'datum_pozadovane', 'produkt__nazov'
    ).order_by('datum_pozadovane')

    rozpracovane_dostupne = Objednavka.objects.filter(
        stav__in=['vyroba', 'pozastavene'],
        operacie__stav='vyroba',
    ).exclude(
        priradeni_operatori=request.user,
    ).select_related('produkt').only(
        'id', 'cislo_objednavky', 'mnozstvo', 'datum_pozadovane', 'stav', 'produkt__nazov'
    ).distinct().order_by('datum_pozadovane')

    vyrobene_spolu = OperatorNaOperacii.objects.filter(
        operator=request.user
    ).aggregate(total=Sum('vyrobene_kusy'))['total'] or 0

    vyrobene_30 = OperatorNaOperacii.objects.filter(
        operator=request.user,
        cas_zaciatku__gte=last_30_days
    ).aggregate(total=Sum('vyrobene_kusy'))['total'] or 0

    nepodarky_spolu = HlasenieVyroby.objects.filter(
        operator=request.user,
        typ_problemu='NEPODAROK'
    ).aggregate(total=Sum('pocet_kusov_nepodarkov'))['total'] or 0

    celkove_kusy = vyrobene_spolu + nepodarky_spolu
    zmetkovitost_pct = round((nepodarky_spolu / celkove_kusy) * 100, 1) if celkove_kusy > 0 else 0
    vykonnost_ks_den = round(vyrobene_30 / 30, 1)

    dokoncene_zakazky = Objednavka.objects.filter(
        stav='hotovo',
        zaznamy__operator=request.user,
        zaznamy__typ_udalosti='STOP'
    ).distinct().count()

    posledne_ukony = VyrobnyZaznam.objects.filter(
        operator=request.user
    ).select_related('objednavka').only(
        'typ_udalosti', 'cas_zaznamu', 'objednavka__cislo_objednavky'
    ).order_by('-cas_zaznamu')[:8]

    operator_name = request.user.get_full_name().strip() or request.user.username
    initials = ''.join(part[0] for part in operator_name.split()[:2]).upper() or request.user.username[:2].upper()
    operator_avatar_url = None
    try:
        profile = request.user.profile
        if profile.avatar:
            operator_avatar_url = profile.avatar.url
    except Exception:
        profile = None

    if not operator_avatar_url:
        email = (request.user.email or '').strip().lower()
        avatar_hash = hashlib.md5(email.encode('utf-8')).hexdigest() if email else None
        operator_avatar_url = (
            f"https://www.gravatar.com/avatar/{avatar_hash}?d=identicon&s=200"
            if avatar_hash
            else f"https://ui-avatars.com/api/?name={initials}&background=0d6efd&color=fff&size=200"
        )

    from math import ceil
    otvorene_zakazky = Objednavka.objects.exclude(stav='hotovo').select_related('produkt', 'produkt__material_ref').only(
        'id', 'stav', 'mnozstvo',
        'produkt__id', 'produkt__dlzka_na_kus_mm', 'produkt__material_ref__id',
        'produkt__material_ref__nazov', 'produkt__material_ref__kod',
        'produkt__material_ref__jednotka', 'produkt__material_ref__aktualna_zasoba',
        'produkt__material_ref__tyc_dlzka_m', 'produkt__material_ref__kg_na_meter',
    )
    potreby_materialu = {}

    for zakazka in otvorene_zakazky:
        produkt = zakazka.produkt
        material = produkt.material_ref
        if not material:
            continue

        zostava = max(0, int(zakazka.zostava_vyroba() or 0))
        if zostava <= 0:
            continue

        dlzka_na_kus = float(produkt.dlzka_na_kus_mm or 0)
        tyc_dlzka_m = float(material.tyc_dlzka_m or 0)
        kg_na_meter = float(material.kg_na_meter or 0)
        jednotka = material.jednotka or 'kg'

        required = 0.0
        if _is_kg_unit(jednotka):
            if dlzka_na_kus <= 0 or kg_na_meter <= 0:
                continue
            dlzka_m = (dlzka_na_kus * zostava) / 1000
            required = dlzka_m * kg_na_meter
        elif _is_bar_unit(jednotka):
            if dlzka_na_kus <= 0 or tyc_dlzka_m <= 0:
                continue
            dlzka_m = (dlzka_na_kus * zostava) / 1000
            required = float(ceil(dlzka_m / tyc_dlzka_m)) if dlzka_m > 0 else 0.0
        else:
            continue

        if required <= 0:
            continue

        data = potreby_materialu.setdefault(material.id, {
            'nazov': material.nazov,
            'kod': material.kod,
            'jednotka': jednotka,
            'required': 0.0,
            'zasoba': float(material.aktualna_zasoba or 0),
        })
        data['required'] += required

    shortage_materials = []
    for data in potreby_materialu.values():
        missing = data['required'] - data['zasoba']
        if missing <= 0:
            continue

        if _is_kg_unit(data['jednotka']):
            required_display = round(data['required'], 2)
            stock_display = round(data['zasoba'], 2)
            missing_display = round(missing, 2)
        else:
            required_display = int(ceil(data['required']))
            stock_display = int(data['zasoba'])
            missing_display = int(ceil(missing))

        shortage_materials.append({
            'nazov': data['nazov'],
            'kod': data['kod'],
            'jednotka': data['jednotka'],
            'required': required_display,
            'zasoba': stock_display,
            'missing': missing_display,
        })

    shortage_materials.sort(key=lambda item: item['missing'], reverse=True)

    first_active_order = next(iter(rozpracovane), None)
    first_assigned_order = next(iter(nove_priradene), None)
    first_available_order = next(iter(nove_dostupne), None)
    operator_mobile_target_url = request.build_absolute_uri(
        f"{reverse('home')}?next={reverse('operator_dashboard')}"
    )
    operator_mobile_qr_url = f"https://quickchart.io/qr?size=170&margin=1&text={quote(operator_mobile_target_url, safe='')}"

    context = {
        'rozpracovane': rozpracovane,
        'nove_priradene': nove_priradene,
        'nove_dostupne': nove_dostupne,
        'rozpracovane_dostupne': rozpracovane_dostupne,
        'operator_name': operator_name,
        'operator_initials': initials,
        'operator_avatar_url': operator_avatar_url,
        'vyrobene_spolu': vyrobene_spolu,
        'nepodarky_spolu': nepodarky_spolu,
        'zmetkovitost_pct': zmetkovitost_pct,
        'vykonnost_ks_den': vykonnost_ks_den,
        'dokoncene_zakazky': dokoncene_zakazky,
        'aktivne_zakazky': rozpracovane.count(),
        'priradene_cakajuce': nove_priradene.count(),
        'dostupne_nove': nove_dostupne.count(),
        'dostupne_rozpracovane': rozpracovane_dostupne.count(),
        'posledne_ukony': posledne_ukony,
        'shortage_materials': shortage_materials,
        'quick_active_order_id': first_active_order.pk if first_active_order else None,
        'quick_assigned_order_id': first_assigned_order.pk if first_assigned_order else None,
        'quick_available_order_id': first_available_order.pk if first_available_order else None,
        'operator_mobile_target_url': operator_mobile_target_url,
        'operator_mobile_qr_url': operator_mobile_qr_url,
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        response = render(request, 'core/operator/_dashboard_content.html', context)
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['Content-Type'] = 'text/html; charset=utf-8'
        return response

    response = render(request, 'core/operator/dashboard.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    response['Content-Type'] = 'text/html; charset=utf-8'
    return response

@login_required
def operator_zakazka_detail(request, pk):
    """Detail zakázky pre operátora s možnosťou riadenia operácií"""
    from django.contrib import messages
    from django.utils import timezone
    from .models import OperaciaVyroby, OperatorNaOperacii
    
    zakazka = get_object_or_404(Objednavka, pk=pk)
    
    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, zakazka):
        messages.error(request, '⚠️ Nie ste priradený k tejto objednávke!')
        return redirect('operator_dashboard')
    
    # Spracovanie akcií operátora
    if request.method == "POST":
        akcia = request.POST.get('akcia')
        operacia_id = request.POST.get('operacia_id')
        
        if operacia_id:
            operacia = get_object_or_404(OperaciaVyroby, pk=operacia_id, objednavka=zakazka)
            
            if akcia == 'zacat':
                # Validácia - môžem začať?
                if not operacia.moze_zacat():
                    predch = operacia.get_predchadzajuca_operacia()
                    if predch and predch.vyrobene_kusy <= 0:
                        messages.error(
                            request, 
                            f'⚠️ Nemôžete začať operáciu {operacia.nazov_operacie}! '
                            f'Predchádzajúca operácia "{predch.nazov_operacie}" ešte nevyrobila žiadne kusy.'
                        )
                    else:
                        messages.error(request, '⚠️ Nemôžete začať túto operáciu!')
                else:
                    operacia.stav = 'vyroba'
                    operacia.datum_zaciatku = timezone.now()
                    operacia.operator = request.user
                    operacia.save()
                    
                    # Vytvor záznam operátora len ak neexistuje otvorený
                    otvoreny = OperatorNaOperacii.objects.filter(
                        operacia=operacia,
                        operator=request.user,
                        cas_konca__isnull=True
                    ).first()
                    if not otvoreny:
                        OperatorNaOperacii.objects.create(
                            operacia=operacia,
                            operator=request.user,
                            cas_zaciatku=timezone.now()
                        )
                    
                    dostupne = operacia.get_dostupne_kusy_na_vstupe()
                    messages.success(
                        request, 
                        f'✅ Začali ste prácu na operácii: {operacia.nazov_operacie}<br>'
                        f'Dostupné kusy: {dostupne} ks'
                    )
            
            elif akcia == 'pauza':
                operacia.stav = 'pozastavena'
                operacia.save()
                
                # Ukončiť aktuálny záznam operátora
                operator_zaznam = operacia.operatori.filter(
                    operator=request.user,
                    cas_konca__isnull=True
                ).first()
                if operator_zaznam:
                    operator_zaznam.cas_konca = timezone.now()
                    operator_zaznam.save()
                
                messages.warning(request, f'⏸️ Operácia pozastavená: {operacia.nazov_operacie}')
            
            elif akcia == 'pokracovat':
                if not operacia.moze_pokracovat():
                    messages.error(request, '⚠️ Nemôžete pokračovať - nie sú dostupné žiadne kusy!')
                else:
                    operacia.stav = 'vyroba'
                    operacia.save()
                    
                    # Vytvor nový záznam operátora len ak neexistuje otvorený
                    otvoreny = OperatorNaOperacii.objects.filter(
                        operacia=operacia,
                        operator=request.user,
                        cas_konca__isnull=True
                    ).first()
                    if not otvoreny:
                        OperatorNaOperacii.objects.create(
                            operacia=operacia,
                            operator=request.user,
                            cas_zaciatku=timezone.now()
                        )
                    
                    dostupne = operacia.get_dostupne_kusy_na_vstupe()
                    messages.success(
                        request, 
                        f'▶️ Pokračujete v práci na: {operacia.nazov_operacie}<br>'
                        f'Zostávajúce kusy: {dostupne} ks'
                    )
            
            elif akcia == 'ukonci_davku':
                try:
                    vyrobene = int(request.POST.get('vyrobene_kusy', 0))
                    nepodarky = int(request.POST.get('nepodarky', 0))
                except (TypeError, ValueError):
                    messages.error(request, '❌ Chyba: Zadané hodnoty musia byť celé čísla.')
                    return redirect('operator_zakazka_detail', pk=pk)

                try:
                    message = _finish_operation_batch(operacia, request.user, vyrobene, nepodarky)
                    messages.success(request, f'✅ {message}')
                except ValueError as e:
                    messages.error(request, f'❌ Chyba: {str(e)}')
        
        # Uzavretie zakázky
        elif akcia == 'uzavri_zakazku':
            try:
                fotka_balenia = request.FILES.get('fotka_balenia_final')
                poznamka_balenia = request.POST.get('poznamka_balenia_final', '')

                message = _close_order_with_packaging_photo(
                    zakazka,
                    request.user,
                    fotka_balenia,
                    poznamka_balenia,
                )
                messages.success(request, f'✅ {message}')
                return redirect('operator_dashboard')
            except ValueError as e:
                messages.error(request, f'❌ {str(e)}')
        
        return redirect('operator_zakazka_detail', pk=pk)
    
    context = _build_operator_order_detail_context(zakazka, request.user)
    
    return render(request, 'core/operator_zakazka_detail.html', context)


def _build_operation_flow_stats(zakazka, operacie):
    ops = list(operacie)

    def _expected_after_operation(index):
        if index >= len(ops):
            return 0
        downstream_nok = sum(max(0, int(op.nepodarky or 0)) for op in ops[index + 1:])
        return max(0, int(zakazka.mnozstvo or 0) + downstream_nok)

    def _actual_after_operation(index):
        if index >= len(ops):
            return 0
        return max(0, int(ops[index].vyrobene_kusy or 0))

    total_wip = 0
    for i in range(len(ops) - 1):
        current_ok = max(0, int(ops[i].vyrobene_kusy or 0))
        next_processed = max(0, int(ops[i + 1].vyrobene_kusy or 0)) + max(0, int(ops[i + 1].nepodarky or 0))
        total_wip += max(current_ok - next_processed, 0)

    kompletne = max(0, int(zakazka.celkom_ok_kusy or 0))
    nok = max(0, int(zakazka.celkom_nok_kusy or 0))
    nekompletne = max(int(zakazka.potrebne_kusy_celkom or 0) - kompletne, 0)

    return {
        'after_first': {
            'actual': _actual_after_operation(0),
            'expected': _expected_after_operation(0),
        },
        'after_second': {
            'actual': _actual_after_operation(1),
            'expected': _expected_after_operation(1),
        },
        'wip_total': total_wip,
        'kompletne': kompletne,
        'nekompletne': nekompletne,
        'nok': nok,
    }


def _build_operator_order_detail_context(zakazka, current_user=None):
    operacie = zakazka.operacie.all().order_by('poradie')

    for op in operacie:
        op.dostupne_kusy = op.get_dostupne_kusy_na_vstupe()
        op.max_kusy = op.get_max_vyrobitelne_kusy()
        op.operatori_list = op.operatori.all()
        op.aktivny_operator = _get_active_operator_session(op)
        op.aktivny_operator_id = op.aktivny_operator.operator_id if op.aktivny_operator else None
        op.aktivny_operator_username = op.aktivny_operator.operator.username if op.aktivny_operator else None
        op.operatori_unikatni = list(dict.fromkeys(
            op.operatori.select_related('operator').values_list('operator__username', flat=True)
        ))
        op.moze_zacat_teraz = op.moze_zacat()
        op.je_aktivny_operator = bool(
            current_user and op.aktivny_operator_id and op.aktivny_operator_id == current_user.id
        )

    moze_uzavriet, dovod = zakazka.moze_sa_uzavriet()
    flow_stats = _build_operation_flow_stats(zakazka, operacie)

    return {
        'zakazka': zakazka,
        'operacie': operacie,
        'flow_stats': flow_stats,
        'moze_uzavriet': moze_uzavriet,
        'dovod_neuzvretia': dovod if not moze_uzavriet else None,
        'dnes': timezone.now().date(),
        'kontrolne_parametre': zakazka.produkt.kontrolne_parametre.all(),
        'posledne_kontroly': zakazka.kontroly.select_related('operator').order_by('-cas_kontroly')[:10],
    }


def _build_operacie_fragment_etag(operacie):
    snapshot = []
    for op in operacie:
        snapshot.append({
            'id': op.id,
            'stav': op.stav,
            'vyrobene_kusy': op.vyrobene_kusy,
            'nepodarky': op.nepodarky,
            'dostupne_kusy': getattr(op, 'dostupne_kusy', None),
            'operatori_unikatni': getattr(op, 'operatori_unikatni', []),
            'aktivny_operator_id': getattr(op, 'aktivny_operator_id', None),
        })

    digest = hashlib.sha256(
        json.dumps(snapshot, ensure_ascii=False, sort_keys=True, default=str).encode('utf-8')
    ).hexdigest()
    return f'W/"{digest}"'


@login_required
def operator_operacie_fragment(request, pk):
    zakazka = get_object_or_404(Objednavka, pk=pk)

    if not _user_has_operator_access(request.user, zakazka):
        return HttpResponse('Prístup zamietnutý', status=403)

    context = _build_operator_order_detail_context(zakazka, request.user)
    etag = _build_operacie_fragment_etag(context['operacie'])

    if_none_match = request.headers.get('If-None-Match', '')
    request_etags = [item.strip() for item in if_none_match.split(',') if item.strip()]
    if etag in request_etags:
        response = HttpResponse(status=304)
        response['ETag'] = etag
        response['Cache-Control'] = 'private, no-cache'
        return response

    response = render(request, 'core/operator/_operacie_card.html', context)
    response['ETag'] = etag
    response['Cache-Control'] = 'private, no-cache'
    return response

# ========================================
# AJAX AKCIE - TRACKING PER OPERÁCIA
# ========================================

@login_required
@require_POST
def start_operation(request, objednavka_pk, operacia_pk):
    rate_limit_response = _check_operator_rate_limit(request, 'start-operation', limit=18, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=objednavka_pk)
    operacia_vyroby = get_object_or_404(OperaciaVyroby, pk=operacia_pk)
    
    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')
    
    # Kontrola, či operácia patrí k objednávke
    if operacia_vyroby.objednavka != objednavka:
        return _api_error('Operácia nepatrí k tejto objednávke')
    
    # Kontrola, či operácia môže pokračovať (pre pozastavené a hotové operácie)
    if operacia_vyroby.stav in ['pozastavena', 'hotova']:
        if not operacia_vyroby.moze_pokracovat():
            return _api_error('Operácia nemôže pokračovať - nie sú dostupné kusy')
    
    VyrobnyZaznam.objects.create(
        objednavka=objednavka,
        operacia=operacia_vyroby.operacia_sablona,
        operator=request.user,
        typ_udalosti='START'
    )

    _close_other_open_operator_sessions(operacia_vyroby, request.user)
    _get_or_create_open_operator_session(operacia_vyroby, request.user)
    
    # Priradenie operátora k operácii
    operacia_vyroby.operator = request.user
    operacia_vyroby.stav = 'vyroba'
    operacia_vyroby.datum_zaciatku = timezone.now()
    operacia_vyroby.save()
    
    if objednavka.stav in ['nova', 'pozastavene']:
        objednavka.stav = 'vyroba'
        objednavka.save()
    
    return _api_ok(
        f'Operácia {operacia_vyroby.nazov_operacie} začatá',
        stav_operacie=operacia_vyroby.stav,
    )

@login_required
@require_POST
def pause_operation(request, objednavka_pk, operacia_pk):
    rate_limit_response = _check_operator_rate_limit(request, 'pause-operation', limit=18, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=objednavka_pk)
    operacia_vyroby = get_object_or_404(OperaciaVyroby, pk=operacia_pk)
    
    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')
    
    # Kontrola, či operácia patrí k objednávke
    if operacia_vyroby.objednavka != objednavka:
        return _api_error('Operácia nepatrí k tejto objednávke')

    aktivny = _get_active_operator_session(operacia_vyroby)
    if aktivny and aktivny.operator_id != request.user.id:
        return _api_error(
            f'Operáciu aktuálne vykonáva operátor {aktivny.operator.username}. '
            f'Najprv ju prevezmite.'
        )

    data = _get_json_body(request)
    if data is None:
        return _api_error('Neplatný JSON formát požiadavky.')
    dovod = data.get('dovod', '')
    
    VyrobnyZaznam.objects.create(
        objednavka=objednavka,
        operacia=operacia_vyroby.operacia_sablona,
        operator=request.user,
        typ_udalosti='PAUZA',
        dovod_pauzy=dovod
    )
    
    operacia_vyroby.stav = 'pozastavena'
    operacia_vyroby.save()

    _close_other_open_operator_sessions(operacia_vyroby)
    
    objednavka.stav = 'pozastavene'
    objednavka.save()
    
    return _api_ok('Operácia pozastavená', stav_operacie=operacia_vyroby.stav)

@login_required
@require_POST
def end_operation(request, objednavka_pk, operacia_pk):
    rate_limit_response = _check_operator_rate_limit(request, 'end-operation', limit=18, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=objednavka_pk)
    operacia_vyroby = get_object_or_404(OperaciaVyroby, pk=operacia_pk)
    
    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')
    
    # Kontrola, či operácia patrí k objednávke
    if operacia_vyroby.objednavka != objednavka:
        return _api_error('Operácia nepatrí k tejto objednávke')

    aktivny = _get_active_operator_session(operacia_vyroby)
    if aktivny and aktivny.operator_id != request.user.id:
        return _api_error(
            f'Operáciu aktuálne vykonáva operátor {aktivny.operator.username}. '
            f'Najprv ju prevezmite.'
        )
    
    VyrobnyZaznam.objects.create(
        objednavka=objednavka,
        operacia=operacia_vyroby.operacia_sablona,
        operator=request.user,
        typ_udalosti='STOP'
    )
    
    operacia_vyroby.stav = 'hotova'
    operacia_vyroby.datum_ukoncenia = timezone.now()
    operacia_vyroby.save()

    _close_other_open_operator_sessions(operacia_vyroby)
    
    return _api_ok(
        f'Operácia {operacia_vyroby.nazov_operacie} ukončená',
        stav_operacie=operacia_vyroby.stav,
    )


@login_required
@require_POST
def end_batch(request, objednavka_pk, operacia_pk):
    rate_limit_response = _check_operator_rate_limit(request, 'end-batch', limit=14, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=objednavka_pk)
    operacia_vyroby = get_object_or_404(OperaciaVyroby, pk=operacia_pk)

    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')

    if operacia_vyroby.objednavka != objednavka:
        return _api_error('Operácia nepatrí k tejto objednávke')

    aktivny = _get_active_operator_session(operacia_vyroby)
    if aktivny and aktivny.operator_id != request.user.id:
        return _api_error(
            f'Operáciu aktuálne vykonáva operátor {aktivny.operator.username}. '
            f'Najprv ju prevezmite, až potom ukončite dávku.'
        )

    if request.POST:
        vyrobene_raw = request.POST.get('vyrobene_kusy', 0)
        nepodarky_raw = request.POST.get('nepodarky', 0)
    else:
        data = _get_json_body(request)
        if data is None:
            return _api_error('Neplatný JSON formát požiadavky.')
        vyrobene_raw = data.get('vyrobene_kusy', 0)
        nepodarky_raw = data.get('nepodarky', 0)

    try:
        vyrobene = int(vyrobene_raw)
        nepodarky = int(nepodarky_raw)
    except (TypeError, ValueError):
        return _api_error('Zadané hodnoty musia byť celé čísla.')

    try:
        message = _finish_operation_batch(operacia_vyroby, request.user, vyrobene, nepodarky)
    except ValueError as e:
        return _api_error(str(e))

    naskladnene = 0
    if operacia_vyroby.get_nasledujuca_operacia() is None:
        naskladnene = _book_finished_goods_delta(
            objednavka,
            request.user,
            note=f'Priebežná príjemka po dávke – zákazka #{objednavka.cislo_objednavky}',
        )

    if naskladnene > 0:
        message = f'{message}. Naskladnené: {naskladnene} ks'

    return _api_ok(message, stav_operacie=operacia_vyroby.stav, naskladnene_kusy=naskladnene)


@login_required
@require_POST
def take_over_operation(request, objednavka_pk, operacia_pk):
    rate_limit_response = _check_operator_rate_limit(request, 'takeover-operation', limit=18, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=objednavka_pk)
    operacia_vyroby = get_object_or_404(OperaciaVyroby, pk=operacia_pk)

    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')

    if operacia_vyroby.objednavka != objednavka:
        return _api_error('Operácia nepatrí k tejto objednávke')

    if operacia_vyroby.stav != 'vyroba':
        return _api_error('Operáciu je možné prevziať iba keď je v stave "V práci".')

    aktivny = _get_active_operator_session(operacia_vyroby)
    if aktivny and aktivny.operator_id == request.user.id:
        return _api_ok('Operáciu už máte prevzatú.', stav_operacie=operacia_vyroby.stav)

    _close_other_open_operator_sessions(operacia_vyroby, request.user)
    _get_or_create_open_operator_session(operacia_vyroby, request.user)

    operacia_vyroby.operator = request.user
    if not operacia_vyroby.datum_zaciatku:
        operacia_vyroby.datum_zaciatku = timezone.now()
    operacia_vyroby.save(update_fields=['operator', 'datum_zaciatku'])

    VyrobnyZaznam.objects.create(
        objednavka=objednavka,
        operacia=operacia_vyroby.operacia_sablona,
        operator=request.user,
        typ_udalosti='START',
    )

    if objednavka.stav in ['nova', 'pozastavene']:
        objednavka.stav = 'vyroba'
        objednavka.save(update_fields=['stav'])

    return _api_ok(
        f'Operácia {operacia_vyroby.nazov_operacie} bola prevzatá operátorom {request.user.username}.',
        stav_operacie=operacia_vyroby.stav,
    )

@login_required
@require_POST
def end_work(request, pk):
    rate_limit_response = _check_operator_rate_limit(request, 'end-work', limit=10, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=pk)

    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')

    # Kontrola, či sú všetky operácie ukončené
    running_operations = objednavka.operacie.filter(stav='vyroba')
    if running_operations.exists():
        operation_names = ', '.join([op.nazov_operacie for op in running_operations])
        return _api_error(f'Nemôžete ukončiť prácu! Nasledujúce operácie sú stále aktívné: {operation_names}')

    fotka = request.FILES.get('fotka')
    try:
        pocet_ok = int(request.POST.get('pocet_ok', 0))
    except (TypeError, ValueError):
        return _api_error('Počet OK kusov musí byť celé číslo.')
    poznamka = request.POST.get('poznamka', '')

    if pocet_ok < 0:
        return _api_error('Počet OK kusov nemôže byť záporný.')

    KontrolaKvality.objects.create(
        objednavka=objednavka,
        operator=request.user,
        namerana_hodnota=f"OK: {pocet_ok}",
        vysledok_ok=True,
        fotka=fotka,
        poznamka=poznamka
    )

    posledna_operacia = objednavka.operacie.order_by('-poradie').first()
    if posledna_operacia:
        objednavka.vyrobene_mnozstvo = posledna_operacia.vyrobene_kusy
    else:
        objednavka.vyrobene_mnozstvo = max(0, int(objednavka.vyrobene_mnozstvo or 0)) + pocet_ok

    if objednavka.je_dokoncena():
        objednavka.stav = 'hotovo'
    else:
        objednavka.stav = 'vyroba'

    objednavka.save(update_fields=['vyrobene_mnozstvo', 'stav'])

    naskladnene = _book_finished_goods_delta(
        objednavka,
        request.user,
        note=poznamka or f"Priebežná príjemka po smene – zákazka #{objednavka.cislo_objednavky}",
    )

    return _api_ok(
        f'Práca ukončená. Naskladnené: {naskladnene} ks.',
        naskladnene_kusy=naskladnene,
        celkom_ok_kusy=objednavka.celkom_ok_kusy,
    )


@login_required
@require_POST
def close_order(request, pk):
    rate_limit_response = _check_operator_rate_limit(request, 'close-order', limit=10, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=pk)

    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')

    fotka_balenia = request.FILES.get('fotka_balenia_final')
    poznamka_balenia = request.POST.get('poznamka_balenia_final', '')

    try:
        message = _close_order_with_packaging_photo(
            objednavka,
            request.user,
            fotka_balenia,
            poznamka_balenia,
        )
    except ValueError as e:
        return _api_error(str(e))

    return _api_ok(message, redirect_url='/operator/')

@login_required
@require_POST
def uloz_kontrolu_kvality(request, pk):
    """Uloženie záznamu kontroly kvality s meraniami a fotkou"""
    rate_limit_response = _check_operator_rate_limit(request, 'save-quality-check', limit=16, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=pk)

    if request.user not in objednavka.priradeni_operatori.all() and not objednavka.zaznamy.filter(operator=request.user).exists():
        return JsonResponse({'status': 'error', 'message': 'Nie ste priradený k tejto objednávke!'})

    fotka = request.FILES.get('fotka_kontroly')
    typ_kontroly = request.POST.get('typ_kontroly', 'PRIEBEZNA')
    if typ_kontroly not in ['PRIEBEZNA', 'FINALNA']:
        typ_kontroly = 'PRIEBEZNA'

    try:
        pocet_ok_kusov = int(request.POST.get('pocet_ok_kontroly') or 0)
        pocet_nok_kusov = int(request.POST.get('pocet_nok_kontroly') or 0)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Počet kusov musí byť celé číslo.'})

    if pocet_ok_kusov < 0 or pocet_nok_kusov < 0:
        return JsonResponse({'status': 'error', 'message': 'Počet kusov nemôže byť záporný.'})

    poznamka = request.POST.get('poznamka_kontroly', '')

    parametre = objednavka.produkt.kontrolne_parametre.all()
    merania_data = []
    vsetky_ok = True

    for param in parametre:
        hodnota_str = request.POST.get(f'meranie_{param.id}', '').strip()
        if hodnota_str:
            try:
                hodnota = Decimal(hodnota_str)
                min_h = param.hodnota_nominalna - param.tolerancia_minus
                max_h = param.hodnota_nominalna + param.tolerancia_plus
                if not (min_h <= hodnota <= max_h):
                    vsetky_ok = False
                merania_data.append((param, hodnota))
            except InvalidOperation:
                return JsonResponse({'status': 'error', 'message': f'Neplatná hodnota pre parameter "{param.nazov}"'})

    namerana_text = ', '.join([f"{p.nazov}: {v} {p.jednotka}" for p, v in merania_data]) if merania_data else 'Bez meraní'

    kontrola = KontrolaKvality.objects.create(
        objednavka=objednavka,
        operator=request.user,
        typ_kontroly=typ_kontroly,
        pocet_ok_kusov=pocet_ok_kusov,
        pocet_nok_kusov=pocet_nok_kusov,
        namerana_hodnota=namerana_text,
        vysledok_ok=vsetky_ok,
        fotka=fotka,
        poznamka=poznamka,
    )

    for param, hodnota in merania_data:
        MeraniePriKontrole.objects.create(
            kontrola=kontrola,
            parameter=param,
            namerana_hodnota=hodnota,
        )

    vysledok_text = '✅ OK' if vsetky_ok else '❌ NOK – niektoré hodnoty sú mimo tolerancie'
    typ_text = 'Priebežná kontrola' if typ_kontroly == 'PRIEBEZNA' else 'Finálna kontrola'
    return JsonResponse({
        'status': 'ok',
        'message': f'{typ_text} uložená. Výsledok: {vysledok_text}',
        'vysledok_ok': vsetky_ok,
        'kontrola_id': kontrola.id,
    })

@login_required
@require_POST
def report_problem(request, pk):
    rate_limit_response = _check_operator_rate_limit(request, 'report-problem', limit=12, window_seconds=60)
    if rate_limit_response:
        return rate_limit_response

    objednavka = get_object_or_404(Objednavka, pk=pk)

    # Kontrola, či je operátor priradený k objednávke
    if not _user_has_operator_access(request.user, objednavka):
        return _api_error('Nie ste priradený k tejto objednávke!')

    if request.FILES or request.POST:
        typ_problemu = request.POST.get('typ_problemu')
        pocet_kusov = request.POST.get('pocet_kusov', 0)
        popis = request.POST.get('popis', '')
        fotka = request.FILES.get('fotka_problemu')
    else:
        data = _get_json_body(request)
        if data is None:
            return _api_error('Neplatný JSON formát požiadavky.')
        typ_problemu = data.get('typ_problemu')
        pocet_kusov = data.get('pocet_kusov', 0)
        popis = data.get('popis', '')
        fotka = None

    allowed_problem_types = {'NEPODAROK', 'PORUCHA_STROJA', 'POSKODENY_NASTROJ', 'INA_CHYBA'}

    try:
        pocet_kusov = int(pocet_kusov)
    except (TypeError, ValueError):
        return _api_error('Počet zlých kusov musí byť celé číslo.')

    if pocet_kusov < 0:
        return _api_error('Počet zlých kusov nemôže byť záporný.')

    if not typ_problemu:
        return _api_error('Typ problému je povinný.')

    typ_problemu = str(typ_problemu).strip().upper()
    if typ_problemu not in allowed_problem_types:
        return _api_error('Neplatný typ problému.')

    popis = str(popis or '').strip()
    if not popis:
        return _api_error('Popis problému je povinný.')

    if typ_problemu == 'NEPODAROK' and pocet_kusov <= 0:
        return _api_error('Pri type Nepodarok musí byť počet zlých kusov väčší ako 0.')

    HlasenieVyroby.objects.create(
        objednavka=objednavka,
        operator=request.user,
        typ_problemu=typ_problemu,
        pocet_kusov_nepodarkov=pocet_kusov,
        popis_problemu=popis,
        fotka_problemu=fotka,
    )

    return _api_ok('Problém nahlásený')

@login_required
@require_POST
def operator_prevziat_zakazku(request, pk):
    """Operátor prevzme novú objednávku priamo bez sub-batch"""
    objednavka = get_object_or_404(Objednavka, pk=pk)

    if not objednavka.produkt.operacie.exists():
        machines = list(
            Stroj.objects.order_by('nazov').values('id', 'nazov', 'status')
        )
        return JsonResponse({
            'status': 'error',
            'reason': 'no_operation_templates',
            'message': 'Produkt nemá žiadnu operáciu. Pred prevzatím doplňte aspoň jednu operáciu z ponuky.',
            'operations': PREDEFINED_OPERATIONS,
            'machines': machines,
        })

    if objednavka.stav == 'hotovo':
        return JsonResponse({'status': 'error', 'message': 'Hotovú zákazku nie je možné prevziať.'})

    if request.user in objednavka.priradeni_operatori.all():
        return JsonResponse({'status': 'ok', 'message': f'Zakázka #{objednavka.cislo_objednavky} je už priradená.'})

    if objednavka.stav == 'nova':
        if objednavka.priradeni_operatori.exists():
            return JsonResponse({'status': 'error', 'message': 'Objednávka už má priradených operátorov!'})

        objednavka.priradeni_operatori.add(request.user)
        objednavka.stav = 'vyroba'
        objednavka.save()

        VyrobnyZaznam.objects.create(
            objednavka=objednavka,
            operator=request.user,
            typ_udalosti='START'
        )

        return JsonResponse({'status': 'ok', 'message': f'Zakázka #{objednavka.cislo_objednavky} bola prevzatá!'})

    if objednavka.stav in ['vyroba', 'pozastavene']:
        objednavka.priradeni_operatori.add(request.user)
        return JsonResponse({
            'status': 'ok',
            'message': f'Zakázka #{objednavka.cislo_objednavky} bola priradená. Otvorte detail a prevezmite konkrétnu operáciu.',
        })

    return JsonResponse({'status': 'error', 'message': 'Objednávka nie je dostupná na prevzatie!'})


@login_required
@require_POST
def operator_add_operation_template(request, objednavka_pk):
    objednavka = get_object_or_404(Objednavka.objects.select_related('produkt'), pk=objednavka_pk)

    if objednavka.stav == 'hotovo':
        return _api_error('Hotovú zákazku už nie je možné meniť.')

    can_prepare = (
        (objednavka.stav == 'nova' and not objednavka.priradeni_operatori.exists())
        or objednavka.stav in {'vyroba', 'pozastavene'}
        or _user_has_operator_access(request.user, objednavka)
    )
    if not can_prepare:
        return _api_error('Túto zákazku nie je možné pripraviť na prevzatie.')

    data = _get_json_body(request)
    if data is None:
        return _api_error('Neplatný JSON formát požiadavky.')

    skip_now = bool(data.get('skip_now'))
    operation_name = str(data.get('operation_name') or '').strip()
    machine_id = str(data.get('machine_id') or '').strip()

    if skip_now:
        operation_name = PREDEFINED_OPERATIONS[0] if PREDEFINED_OPERATIONS else 'Výroba'
        fallback_machine = Stroj.objects.filter(status='OK').order_by('nazov').first() or Stroj.objects.order_by('nazov').first()
        if not fallback_machine:
            return _api_error('Nie je dostupný žiaden stroj. Najprv založte aspoň jeden stroj.')
        machine_id = str(fallback_machine.pk)

    if operation_name not in PREDEFINED_OPERATIONS:
        return _api_error('Vyberte platnú operáciu z ponuky.')

    if not machine_id.isdigit():
        return _api_error('Vyberte platný stroj.')

    machine = Stroj.objects.filter(pk=int(machine_id)).first()
    if not machine:
        return _api_error('Vybraný stroj neexistuje.')

    produkt = objednavka.produkt
    with transaction.atomic():
        exists = Operacia.objects.filter(produkt=produkt, nazov_operacie__iexact=operation_name).exists()
        if exists:
            return _api_error('Táto operácia už je pre produkt evidovaná.')

        last_template = Operacia.objects.filter(produkt=produkt).order_by('-poradie', '-id').first()
        next_order = (last_template.poradie if last_template else 0) + 1

        created = Operacia.objects.create(
            produkt=produkt,
            stroj=machine,
            poradie=next_order,
            nazov_operacie=operation_name,
            typ_balenia='',
            cas_pripravy=0,
            cas_kus=0,
        )

        if not objednavka.operacie.exists() and objednavka.stav in {'nova', 'vyroba', 'pozastavene'}:
            objednavka._vytvor_operacie_z_kusovnika()

    result_message = 'Operácia bola doplnená. Pokračujte prevzatím zákazky.'
    if skip_now:
        result_message = 'Použitá dočasná predvolená operácia. Pokračujem prevzatím zákazky.'

    return _api_ok(
        result_message,
        operation_id=created.pk,
        operation_name=created.nazov_operacie,
        machine_name=machine.nazov,
    )

@login_required
def download_sprievodka(request, pk):
    """Otvorenie PDF sprievodky v náhľade prehliadača."""
    objednavka = get_object_or_404(Objednavka, pk=pk)
    pdf_buffer = generate_sprievodka_pdf(objednavka, request)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="sprievodka_{objednavka.cislo_objednavky}.pdf"'
    return response

@login_required
def download_sprievodka_davka(request, pk):
    """Otvorenie PDF sprievodky pre výrobnú dávku v náhľade prehliadača."""
    from django.http import Http404
    davka = get_object_or_404(VyrobnaDavka, pk=pk)
    if not davka.objednavka:
        raise Http404
    objednavka = davka.objednavka
    pdf_buffer = generate_sprievodka_pdf(objednavka, request)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="sprievodka_{objednavka.cislo_objednavky}.pdf"'
    return response

@login_required
@permission_required("core.view_objednavka", raise_exception=True)
def vytvor_objednavku_z_davky(request, davka_pk):
    """Vytvorí objednávku z výrobnej dávky"""
    from .models import VyrobnaDavka
    from django.contrib import messages
    
    davka = get_object_or_404(VyrobnaDavka, pk=davka_pk)
    
    if davka.objednavka:
        messages.warning(request, f'Objednávka už existuje: #{davka.objednavka.cislo_objednavky}')
        return redirect('admin:core_objednavka_change', davka.objednavka.pk)
    
    objednavka = davka.vytvor_objednavku()
    messages.success(request, f'✅ Objednávka #{objednavka.cislo_objednavky} bola vytvorená z dávky {davka.cislo_davky}')
    
    return redirect('admin:core_objednavka_change', objednavka.pk)

@login_required
@permission_required("core.view_objednavka", raise_exception=True)
def vytvor_davku_z_kontraktu(request, kontrakt_pk):
    """Rýchle vytvorenie dávky + objednávky z kontraktu"""
    from .models import Kontrakt, VyrobnaDavka
    from django.contrib import messages
    from django.utils.dateparse import parse_date
    
    kontrakt = get_object_or_404(Kontrakt, pk=kontrakt_pk)
    
    if request.method == 'POST':
        mnozstvo = int(request.POST.get('mnozstvo', 0))
        cislo_objednavky_zakaznika = (request.POST.get('cislo_objednavky_zakaznika') or '').strip()
        pozadovany_termin_raw = (request.POST.get('pozadovany_termin') or '').strip()
        index_kontrola = (request.POST.get('index_kontrola') or '').strip()
        index_suhlasi = request.POST.get('index_suhlasi') == 'on'
        index_produktu = (kontrakt.produkt.index or '').strip()
        
        if mnozstvo <= 0:
            messages.error(request, '❌ Množstvo musí byť väčšie ako 0')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)
        
        if mnozstvo > kontrakt.zostavajuce_mnozstvo:
            messages.error(request, f'❌ Množstvo ({mnozstvo} ks) je väčšie ako zostáva dodať ({kontrakt.zostavajuce_mnozstvo} ks)')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)

        if not cislo_objednavky_zakaznika:
            messages.error(request, '❌ Zadajte číslo objednávky zákazníka (PO), aby sa dal overiť index.')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)

        if not index_produktu:
            messages.error(request, '❌ Produkt nemá nastavený index. Doplnte index v karte produktu a skúste znova.')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)

        if not index_kontrola:
            messages.error(request, '❌ Zadajte index z objednávky zákazníka na kontrolu.')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)

        if index_kontrola.upper() != index_produktu.upper():
            messages.error(request, f'❌ Index nesúhlasí. V PO je "{index_kontrola}", produkt má index "{index_produktu}".')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)

        if not index_suhlasi:
            messages.error(request, '❌ Potvrďte, že index podľa objednávky zákazníka súhlasí.')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)
        
        pozadovany_termin = parse_date(pozadovany_termin_raw)
        if not pozadovany_termin:
            messages.error(request, '❌ Zadajte platný termín dodania.')
            return redirect('vytvor_davku_z_kontraktu', kontrakt_pk=kontrakt_pk)
        
        davka = VyrobnaDavka.objects.create(
            kontrakt=kontrakt,
            mnozstvo_davky=mnozstvo,
            pozadovany_termin=pozadovany_termin,
            datum_vytvorenia=timezone.now().date()
        )
        
        objednavka = davka.vytvor_objednavku()
        objednavka.cislo_objednavky_zakaznika = cislo_objednavky_zakaznika
        objednavka.datum_pozadovane = pozadovany_termin
        objednavka.save(update_fields=['cislo_objednavky_zakaznika', 'datum_pozadovane'])
        
        kontrakt.zostavajuce_mnozstvo -= mnozstvo
        kontrakt.save()
        
        messages.success(request, f'✅ Vytvorená dávka {davka.cislo_davky} a objednávka #{objednavka.cislo_objednavky}')
        return redirect('plan_vyroby')
    
    sklad = SkladHotovychDielov.objects.filter(produkt=kontrakt.produkt).first()
    sklad_mnozstvo = sklad.mnozstvo if sklad else 0

    context = {
        'kontrakt': kontrakt,
        'dnes': timezone.now().date(),
        'sklad_mnozstvo': sklad_mnozstvo,
    }
    return render(request, 'core/vytvor_davku_form.html', context)
# ========================================
# SKLAD HOTOVÝCH DIELOV
# ========================================

@login_required
@permission_required("core.view_skladhotovychdielov", raise_exception=True)
def sklad_hotovych_dielov(request):
    """Dashboard skladu hotových dielov"""
    from .models import SkladHotovychDielov
    
    sklady = SkladHotovychDielov.objects.select_related('produkt').all()
    
    # Štatistiky
    celkom_produktov = sklady.count()
    pod_minimom = sklady.filter(mnozstvo__lt=F('minimalna_zasoba')).count()
    nad_optimom = sklady.filter(mnozstvo__gt=F('optimalna_zasoba')).count()
    
    # Posledné pohyby
    from .models import PrijemkaHotovychDielov, VydajkaHotovychDielov
    posledne_prijemky = PrijemkaHotovychDielov.objects.select_related('sklad__produkt', 'objednavka').order_by('-datum')[:10]
    posledne_vydajky = VydajkaHotovychDielov.objects.select_related('sklad__produkt', 'objednavka', 'kontrakt').order_by('-datum')[:10]
    
    context = {
        'sklady': sklady,
        'celkom_produktov': celkom_produktov,
        'pod_minimom': pod_minimom,
        'nad_optimom': nad_optimom,
        'posledne_prijemky': posledne_prijemky,
        'posledne_vydajky': posledne_vydajky,
    }
    
    return render(request, 'core/sklad.html', context)


@login_required
@permission_required("core.view_skladhotovychdielov", raise_exception=True)
def prvotna_inventura_hotovych_dielov(request):
    from .models import Produkt, SkladHotovychDielov, InventuraHotovychDielov

    # Produkty na zadávanie inventúry
    q = (request.GET.get('q') or '').strip()

    produkty = Produkt.objects.all().order_by('cislo_dielu')
    if q:
        produkty = produkty.filter(
            Q(cislo_dielu__icontains=q)
            | Q(nazov__icontains=q)
            | Q(index__icontains=q)
            | Q(cislo_vykresu__icontains=q)
            | Q(material__icontains=q)
        )

    sklady_map = {
        sklad.produkt_id: sklad
        for sklad in SkladHotovychDielov.objects.select_related('produkt').filter(produkt__in=produkty)
    }

    riadky = []
    for produkt in produkty:
        sklad = sklady_map.get(produkt.id)
        riadky.append({
            'produkt': produkt,
            'sklad': sklad,
            'systemova_zasoba': sklad.mnozstvo if sklad else 0,
        })

    # História inventúr
    inventury = InventuraHotovychDielov.objects.select_related('vykonal').order_by('-datum')
    for inv in inventury:
        inv.pocet_poloziek = inv.polozky.count()

    context = {
        'riadky': riadky,
        'q': q,
        'inventury': inventury,
    }
    return render(request, 'core/prvotna_inventura_hotovych_dielov.html', context)


@login_required
@permission_required("core.change_skladhotovychdielov", raise_exception=True)
@require_POST
def prvotna_inventura_hotovych_dielov_zmenit(request, produkt_id):
    from django.contrib import messages
    from .models import Produkt, SkladHotovychDielov
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    produkt = get_object_or_404(Produkt, pk=produkt_id)
    raw_skutocna = (request.POST.get('skutocna_zasoba') or '').strip().replace(',', '.')
    poznamka = (request.POST.get('poznamka') or '').strip()

    if raw_skutocna == '':
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Zadajte skutočnú zásobu pre produkt {produkt.nazov}.'}, status=400)
        messages.error(request, f'❌ Zadajte skutočnú zásobu pre produkt {produkt.nazov}.')
        return redirect('prvotna_inventura_hotovych_dielov')

    try:
        skutocna = int(float(raw_skutocna))
    except (ValueError, TypeError):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Neplatná hodnota pre produkt {produkt.nazov}.'}, status=400)
        messages.error(request, f'❌ Neplatná hodnota pre produkt {produkt.nazov}.')
        return redirect('prvotna_inventura_hotovych_dielov')

    if skutocna < 0:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Zásoba nemôže byť záporná ({produkt.nazov}).'}, status=400)
        messages.error(request, f'❌ Zásoba nemôže byť záporná ({produkt.nazov}).')
        return redirect('prvotna_inventura_hotovych_dielov')

    sklad, created = SkladHotovychDielov.objects.get_or_create(
        produkt=produkt,
        defaults={
            'mnozstvo': skutocna,
            'minimalna_zasoba': 0,
            'optimalna_zasoba': max(100, skutocna),
        },
    )

    if not created:
        sklad.mnozstvo = skutocna
        if not sklad.minimalna_zasoba:
            sklad.minimalna_zasoba = 0
        if not sklad.optimalna_zasoba:
            sklad.optimalna_zasoba = max(100, skutocna)
        sklad.save(update_fields=['mnozstvo', 'minimalna_zasoba', 'optimalna_zasoba'])

    if is_ajax:
        return JsonResponse({
            'status': 'ok',
            'message': f'Inventúra hotových dielov uložená pre {produkt.nazov}.',
            'systemova_zasoba': int(sklad.mnozstvo),
            'stav': 'inventarizovane',
        })

    messages.success(request, f'✅ Inventúra hotových dielov uložená pre {produkt.nazov}.')
    return redirect('prvotna_inventura_hotovych_dielov')


@login_required
@permission_required("core.view_skladhotovychdielov", raise_exception=True)
def export_prvotna_inventura_hotovych_dielov_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from datetime import datetime
    import os

    from .models import Produkt, SkladHotovychDielov

    produkty = Produkt.objects.all().order_by('cislo_dielu')
    sklad_map = {
        sklad.produkt_id: sklad
        for sklad in SkladHotovychDielov.objects.select_related('produkt').filter(produkt__in=produkty)
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=0.8 * cm,
    )

    elements = []
    styles = getSampleStyleSheet()

    normal_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'
    try:
        arial_path = r'C:\Windows\Fonts\arial.ttf'
        arial_bold_path = r'C:\Windows\Fonts\arialbd.ttf'
        if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
            pdfmetrics.registerFont(TTFont('ArialUnicode', arial_path))
            pdfmetrics.registerFont(TTFont('ArialUnicode-Bold', arial_bold_path))
            normal_font = 'ArialUnicode'
            bold_font = 'ArialUnicode-Bold'
    except Exception:
        pass

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=bold_font,
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=1,
    )
    metadata_style = ParagraphStyle(
        'MetadataStyle',
        parent=styles['Normal'],
        fontName=normal_font,
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4,
    )
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName=bold_font, fontSize=10)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName=normal_font, fontSize=8.5)

    elements.append(Paragraph('Inventúra hotových dielov', title_style))
    elements.append(Paragraph('<b>Vyhotovil:</b> ________________________________', metadata_style))
    elements.append(Paragraph('<b>Dátum vyhotovenia:</b> ________________________', metadata_style))
    elements.append(Spacer(1, 0.6 * cm))

    table_data = [[
        Paragraph('Por.č.', header_style),
        Paragraph('Názov položky', header_style),
        Paragraph('Číslo dielu', header_style),
        Paragraph('Aktuálna zásoba v systéme', header_style),
        Paragraph('Skutočná zásoba', header_style),
    ]]

    for idx, produkt in enumerate(produkty, 1):
        sklad = sklad_map.get(produkt.id)
        systemova = sklad.mnozstvo if sklad else 0
        table_data.append([
            Paragraph(f'{idx}', cell_style),
            Paragraph(produkt.nazov, cell_style),
            Paragraph(produkt.cislo_dielu, cell_style),
            Paragraph(f'{systemova} ks', cell_style),
            Paragraph('', cell_style),
        ])

    table = Table(
        table_data,
        colWidths=[1.0 * cm, 6.6 * cm, 3.0 * cm, 4.0 * cm, 4.8 * cm],
        rowHeights=[0.95 * cm] + [0.8 * cm] * len(produkty),
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), normal_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (4, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="prvotna_inventura_hotovych_dielov_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
@permission_required("core.change_skladhotovychdielov", raise_exception=True)
@require_POST
def ulozit_prvotnu_inventuru_hotovych_dielov(request):
    """
    Prijme všetky zadané množstvá z formuláru a uloží inventúru.
    Vrátí JSON bez PDF downloadu.
    Očakávané POST parametre: product_XX: číslo (kde XX je product.id)
    """
    from .models import (
        Produkt,
        SkladHotovychDielov,
        InventuraHotovychDielov,
        InventuraHotovychDielovPolozka,
    )

    # Vytvor inventúru
    inventura = InventuraHotovychDielov.objects.create(
        vykonal=request.user,
        poznamka=(request.POST.get('poznamka') or '').strip(),
    )

    # Spracuj všetky POST parametre vo formáte product_XXX
    polozky = []
    for key, value in request.POST.items():
        if key.startswith('product_'):
            try:
                product_id = int(key.replace('product_', ''))
                mnozstvo = int(value) if value else 0
                
                if mnozstvo > 0:  # Uložiť len nenulové množstvá
                    produkt = Produkt.objects.get(id=product_id)
                    polozky.append(
                        InventuraHotovychDielovPolozka(
                            inventura=inventura,
                            produkt=produkt,
                            nazov_produktu=produkt.nazov,
                            cislo_dielu=produkt.cislo_dielu,
                            mnozstvo_ks=mnozstvo,
                        )
                    )
            except (ValueError, Produkt.DoesNotExist):
                continue

    # Hromadne vytvor všetky položky
    if polozky:
        InventuraHotovychDielovPolozka.objects.bulk_create(polozky)

    return JsonResponse({
        'status': 'ok',
        'message': f'✅ Inventúra #{inventura.id} bola úspešne uložená. Zaznamenaných položiek: {len(polozky)}',
        'inventura_id': inventura.id,
        'pocet_poloziek': len(polozky)
    })

@login_required
@permission_required("core.view_material", raise_exception=True)
def sklad_materialu(request):
    """Dashboard skladu materiálu"""
    from math import ceil
    materialy = Material.objects.all().order_by('nazov')
    otvorene_zakazky = Objednavka.objects.exclude(stav='hotovo').select_related('produkt', 'produkt__material_ref')

    potreby = {}
    for zakazka in otvorene_zakazky:
        produkt = zakazka.produkt
        material = produkt.material_ref
        if not material:
            continue
        if not produkt.dlzka_na_kus_mm or produkt.dlzka_na_kus_mm <= 0:
            continue
        if not material.kg_na_meter or material.kg_na_meter <= 0:
            continue

        dlzka_m = (float(produkt.dlzka_na_kus_mm) * zakazka.mnozstvo) / 1000
        kg = dlzka_m * float(material.kg_na_meter)
        tyc_dlzka_m = float(material.tyc_dlzka_m or 0)
        tyce = ceil(dlzka_m / tyc_dlzka_m) if tyc_dlzka_m > 0 else 0

        data = potreby.setdefault(material.id, {'m': 0.0, 'kg': 0.0, 'tyce': 0})
        data['m'] += dlzka_m
        data['kg'] += kg
        data['tyce'] += tyce

    for material in materialy:
        data = potreby.get(material.id, {'m': 0.0, 'kg': 0.0, 'tyce': 0})
        material.potreba_m = round(data['m'], 2)
        material.potreba_kg = round(data['kg'], 2)
        material.potreba_tyce = data['tyce']
        if str(material.jednotka or '').strip().lower() == 'kg':
            material.nedostatok_kg = round(max(0.0, material.potreba_kg - float(material.aktualna_zasoba)), 2)
        else:
            material.nedostatok_kg = None
    
    # Štatistiky
    celkom = materialy.count()
    pod_minimom = materialy.filter(aktualna_zasoba__lt=F('minimalna_zasoba')).count()
    
    # Posledné pohyby
    posledne_prijemky = PrijemkaNaSklad.objects.select_related('material').order_by('-datum')[:10]
    posledne_vydajky = VydajkaZoSkladu.objects.select_related('material', 'objednavka').order_by('-datum')[:10]
    
    context = {
        'materialy': materialy,
        'celkom': celkom,
        'pod_minimom': pod_minimom,
        'posledne_prijemky': posledne_prijemky,
        'posledne_vydajky': posledne_vydajky,
        'ai_material_enabled': _is_ai_material_enabled(),
        'ai_material_allowed_domains': _get_allowed_material_ai_domains(),
        'ai_material_navrhy': MaterialAINavrh.objects.select_related('created_by', 'material').all()[:5],
    }
    
    return render(request, 'core/sklad_materialu.html', context)


@login_required
@permission_required("core.view_material", raise_exception=True)
def inventura_materialu(request):
    materialy = Material.objects.all().order_by('nazov')
    zaznamy_qs = InventurnyZaznamMaterialu.objects.select_related('material', 'vykonal').order_by('-datum')
    posledne_zaznamy = zaznamy_qs[:50]

    denny_map = defaultdict(lambda: {
        'datum': None,
        'pocet_zaznamov': 0,
        'chybalo_spolu': Decimal('0.00'),
        'strata_spolu': Decimal('0.00'),
    })

    celkove_chybalo = Decimal('0.00')
    celkova_strata = Decimal('0.00')
    for zaznam in zaznamy_qs:
        den = zaznam.datum.date()
        bucket = denny_map[den]
        bucket['datum'] = den
        bucket['pocet_zaznamov'] += 1

        if zaznam.inventurny_rozdiel < 0:
            chybalo = -zaznam.inventurny_rozdiel
            bucket['chybalo_spolu'] += chybalo
            celkove_chybalo += chybalo

        bucket['strata_spolu'] += zaznam.strata_eur
        celkova_strata += zaznam.strata_eur

    denny_prehlad = sorted(denny_map.values(), key=lambda x: x['datum'], reverse=True)

    context = {
        'materialy': materialy,
        'posledne_zaznamy': posledne_zaznamy,
        'denny_prehlad': denny_prehlad,
        'celkove_chybalo': celkove_chybalo,
        'celkova_strata': celkova_strata,
    }
    return render(request, 'core/inventura_materialu.html', context)


@login_required
@permission_required("core.change_material", raise_exception=True)
@require_POST
def inventura_materialu_zmenit(request, material_id):
    from django.contrib import messages
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    material = get_object_or_404(Material, pk=material_id)
    raw_skutocna = (request.POST.get('skutocna_zasoba') or '').strip().replace(',', '.')
    poznamka = (request.POST.get('poznamka') or '').strip()

    if raw_skutocna == '':
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Zadajte skutočnú zásobu pre materiál {material.nazov}.'}, status=400)
        messages.error(request, f'❌ Zadajte skutočnú zásobu pre materiál {material.nazov}.')
        return redirect('inventura_materialu')

    try:
        skutocna = Decimal(raw_skutocna)
    except (InvalidOperation, ValueError):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Neplatná hodnota zásoby pre materiál {material.nazov}.'}, status=400)
        messages.error(request, f'❌ Neplatná hodnota zásoby pre materiál {material.nazov}.')
        return redirect('inventura_materialu')

    if skutocna < 0:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'Skutočná zásoba nemôže byť záporná ({material.nazov}).'}, status=400)
        messages.error(request, f'❌ Skutočná zásoba nemôže byť záporná ({material.nazov}).')
        return redirect('inventura_materialu')

    systemova = material.aktualna_zasoba
    rozdiel = skutocna - systemova
    cena_v_case = material.cena_za_jednotku
    strata_eur = Decimal('0.00')
    if rozdiel < 0:
        strata_eur = (-rozdiel) * cena_v_case

    InventurnyZaznamMaterialu.objects.create(
        material=material,
        systemova_zasoba=systemova,
        skutocna_zasoba=skutocna,
        inventurny_rozdiel=rozdiel,
        cena_za_jednotku_v_case=cena_v_case,
        strata_eur=strata_eur,
        vykonal=request.user,
        poznamka=poznamka,
    )

    material.aktualna_zasoba = skutocna
    material.save(update_fields=['aktualna_zasoba'])

    znamienko = '+' if rozdiel >= 0 else ''
    if is_ajax:
        return JsonResponse({
            'status': 'ok',
            'message': f'Inventúra uložená pre {material.nazov}.',
            'systemova_zasoba': str(skutocna),
            'jednotka': material.jednotka,
            'rozdiel_text': f'{znamienko}{rozdiel} {material.jednotka}',
            'rozdiel_class': 'bg-success' if rozdiel >= 0 else 'bg-danger',
            'strata_eur': f'{strata_eur:.2f}',
        })

    messages.success(
        request,
        f'✅ Inventúra uložená pre {material.nazov}. Rozdiel: {znamienko}{rozdiel} {material.jednotka}, strata: {strata_eur:.2f} €.',
    )
    return redirect('inventura_materialu')


@login_required
@permission_required("core.view_material", raise_exception=True)
def export_inventura_historia_excel(request):
    wb = openpyxl.Workbook()
    ws_sumar = wb.active
    ws_sumar.title = "Denný prehľad"

    zaznamy = list(InventurnyZaznamMaterialu.objects.select_related('material', 'vykonal').order_by('-datum'))

    denny_map = defaultdict(lambda: {
        'datum': None,
        'pocet': 0,
        'chybalo': Decimal('0.00'),
        'strata': Decimal('0.00'),
    })
    for z in zaznamy:
        den = z.datum.date()
        bucket = denny_map[den]
        bucket['datum'] = den
        bucket['pocet'] += 1
        if z.inventurny_rozdiel < 0:
            bucket['chybalo'] += -z.inventurny_rozdiel
        bucket['strata'] += z.strata_eur

    ws_sumar.append(["Dátum", "Počet záznamov", "Chýbalo spolu", "Strata spolu (€)"])
    for c in ws_sumar[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        c.alignment = Alignment(horizontal="center")

    for den in sorted(denny_map.keys(), reverse=True):
        data = denny_map[den]
        ws_sumar.append([den.strftime('%d.%m.%Y'), data['pocet'], float(data['chybalo']), float(data['strata'])])

    for col in ["A", "B", "C", "D"]:
        ws_sumar.column_dimensions[col].width = 22

    ws_detail = wb.create_sheet("Detail inventúr")
    ws_detail.append([
        "Dátum", "Materiál", "Kód", "Jednotka", "Systémová zásoba", "Skutočná zásoba",
        "Inventúrny rozdiel", "Cena v čase", "Strata (€)", "Vykonal", "Poznámka"
    ])
    for c in ws_detail[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        c.alignment = Alignment(horizontal="center")

    for z in zaznamy:
        ws_detail.append([
            z.datum.strftime('%d.%m.%Y %H:%M'),
            z.material.nazov,
            z.material.kod,
            z.material.jednotka,
            float(z.systemova_zasoba),
            float(z.skutocna_zasoba),
            float(z.inventurny_rozdiel),
            float(z.cena_za_jednotku_v_case),
            float(z.strata_eur),
            z.vykonal.username if z.vykonal else '-',
            z.poznamka,
        ])

    detail_widths = {
        "A": 18, "B": 30, "C": 14, "D": 10, "E": 16,
        "F": 16, "G": 16, "H": 14, "I": 12, "J": 14, "K": 24,
    }
    for col, width in detail_widths.items():
        ws_detail.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="inventura_historia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@permission_required("core.view_material", raise_exception=True)
def export_inventura_historia_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    import os

    zaznamy = list(InventurnyZaznamMaterialu.objects.select_related('material', 'vykonal').order_by('-datum'))

    denny_map = defaultdict(lambda: {
        'datum': None,
        'pocet': 0,
        'chybalo': Decimal('0.00'),
        'strata': Decimal('0.00'),
    })
    for z in zaznamy:
        den = z.datum.date()
        bucket = denny_map[den]
        bucket['datum'] = den
        bucket['pocet'] += 1
        if z.inventurny_rozdiel < 0:
            bucket['chybalo'] += -z.inventurny_rozdiel
        bucket['strata'] += z.strata_eur

    normal_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'
    try:
        arial_path = r'C:\Windows\Fonts\arial.ttf'
        arial_bold_path = r'C:\Windows\Fonts\arialbd.ttf'
        if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
            pdfmetrics.registerFont(TTFont('ArialUnicode', arial_path))
            pdfmetrics.registerFont(TTFont('ArialUnicode-Bold', arial_bold_path))
            normal_font = 'ArialUnicode'
            bold_font = 'ArialUnicode-Bold'
    except Exception:
        pass

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontName=bold_font, fontSize=16)
    text_style = ParagraphStyle('text', parent=styles['Normal'], fontName=normal_font, fontSize=9)

    elements = [
        Paragraph('História inventúr materiálu', title_style),
        Paragraph(f'Vygenerované: {datetime.now().strftime("%d.%m.%Y %H:%M")}', text_style),
        Spacer(1, 0.4 * cm),
    ]

    sumar_data = [[
        Paragraph('Dátum', text_style),
        Paragraph('Počet', text_style),
        Paragraph('Chýbalo spolu', text_style),
        Paragraph('Strata spolu (€)', text_style),
    ]]

    for den in sorted(denny_map.keys(), reverse=True):
        d = denny_map[den]
        sumar_data.append([
            Paragraph(den.strftime('%d.%m.%Y'), text_style),
            Paragraph(str(d['pocet']), text_style),
            Paragraph(f"{d['chybalo']:.2f}", text_style),
            Paragraph(f"{d['strata']:.2f}", text_style),
        ])

    sumar_table = Table(sumar_data, colWidths=[4 * cm, 3 * cm, 5 * cm, 5 * cm])
    sumar_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(sumar_table)
    elements.append(Spacer(1, 0.5 * cm))

    detail_data = [[
        Paragraph('Dátum', text_style),
        Paragraph('Položka', text_style),
        Paragraph('Rozdiel', text_style),
        Paragraph('Strata €', text_style),
    ]]
    for z in zaznamy[:150]:
        detail_data.append([
            Paragraph(z.datum.strftime('%d.%m.%Y %H:%M'), text_style),
            Paragraph(z.material.nazov, text_style),
            Paragraph(f"{z.inventurny_rozdiel:+.2f} {z.material.jednotka}", text_style),
            Paragraph(f"{z.strata_eur:.2f}", text_style),
        ])

    detail_table = Table(detail_data, colWidths=[4 * cm, 7.5 * cm, 4 * cm, 3 * cm])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="inventura_historia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )
    return response


@login_required
@permission_required("core.view_material", raise_exception=True)
def export_material_inventory_pdf(request):
    """Export materiálov na inventúru do PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from datetime import datetime
    import os
    
    # Prepare materials data
    materialy = Material.objects.all().order_by('nazov')
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=0.8*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()

    # Register Unicode font to correctly render Slovak diacritics (e.g. "č").
    normal_font = 'Helvetica'
    bold_font = 'Helvetica-Bold'
    try:
        arial_path = r'C:\Windows\Fonts\arial.ttf'
        arial_bold_path = r'C:\Windows\Fonts\arialbd.ttf'
        if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
            pdfmetrics.registerFont(TTFont('ArialUnicode', arial_path))
            pdfmetrics.registerFont(TTFont('ArialUnicode-Bold', arial_bold_path))
            normal_font = 'ArialUnicode'
            bold_font = 'ArialUnicode-Bold'
    except Exception:
        pass
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        fontName=bold_font,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=1  # center
    )
    title = Paragraph('Inventúra majetku - Materiály', title_style)
    elements.append(title)
    
    # Metadata fields for signature
    metadata_style = ParagraphStyle(
        'MetadataStyle',
        parent=styles['Normal'],
        fontName=normal_font,
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4
    )

    table_header_style = ParagraphStyle(
        'TableHeaderStyle',
        parent=styles['Normal'],
        fontName=bold_font,
        fontSize=10,
    )

    table_cell_style = ParagraphStyle(
        'TableCellStyle',
        parent=styles['Normal'],
        fontName=normal_font,
        fontSize=8.5,
    )
    
    elements.append(Paragraph('<b>Vyhotovil:</b> _' + '_'*50, metadata_style))
    elements.append(Paragraph('<b>Dátum vyhotovenia:</b> _' + '_'*40, metadata_style))
    elements.append(Spacer(1, 0.6*cm))
    
    # Prepare table data
    table_data = [
        [
            Paragraph('Por.č.', table_header_style),
            Paragraph('Názov materiálu', table_header_style),
            Paragraph('Ø (mm)', table_header_style),
            Paragraph('J.', table_header_style),
            Paragraph('Skutočný stav', table_header_style),
        ]
    ]
    
    # Add material rows
    for idx, material in enumerate(materialy, 1):
        table_data.append([
            Paragraph(f'{idx}', table_cell_style),
            Paragraph(material.nazov, table_cell_style),
            Paragraph(f'{material.priemer_mm}' if material.priemer_mm else '-', table_cell_style),
            Paragraph(material.jednotka, table_cell_style),
            Paragraph('', table_cell_style),  # Empty field for actual inventory count
        ])
    
    # Create table with wider columns (full page width)
    header_height = 0.95 * cm
    item_height = 0.8 * cm
    row_heights = [header_height] + [item_height] * len(materialy)
    table = Table(
        table_data,
        colWidths=[1.2 * cm, 7.2 * cm, 2 * cm, 1.1 * cm, 6.4 * cm],
        rowHeights=row_heights,
    )
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Center header Č.
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Center sequence numbers
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),  # Center diameter and unit
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), normal_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(table)
    
    # Footer with totals
    elements.append(Spacer(1, 0.4*cm))
    footer_text = f'Spolu materiálov: {materialy.count()}'
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontName=normal_font,
        fontSize=9,
    )
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return as PDF response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventura_materialov_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@require_POST
@login_required
@permission_required("core.add_material", raise_exception=True)
def ai_material_navrh(request):
    if not _is_ai_material_enabled():
        return _api_error('AI návrhy materiálu sú momentálne vypnuté.')

    payload = _get_json_body(request)
    if payload is None:
        return _api_error('Neplatný JSON payload.')

    query = str(payload.get('query') or '').strip()
    source_url = str(payload.get('source_url') or '').strip()

    if len(query) < 3:
        return _api_error('Zadaj aspoň stručný názov alebo dopyt materiálu.')

    is_allowed, domain = _is_allowed_material_ai_url(source_url)
    if not is_allowed:
        domains = ', '.join(_get_allowed_material_ai_domains())
        return _api_error(f'Zdrojová URL nie je povolená. Použi slovenské weby: {domains}.')

    try:
        ai_response = _generate_material_ai_response(query, source_url)
    except ValueError as exc:
        return _api_error(str(exc))
    except Exception as exc:
        logger.exception('AI material navrh failed')
        return _api_error('AI služba momentálne neodpovedá. Skús to znova o chvíľu.')

    ai_data = ai_response.get('data') or {}
    confidence = _safe_decimal(ai_data.get('confidence'), default='0')
    navrh = MaterialAINavrh.objects.create(
        query=query,
        source_url=source_url,
        source_domain=domain,
        ai_model=ai_response.get('model') or '',
        confidence=confidence,
        navrh_data={
            'nazov': str(ai_data.get('nazov') or '').strip(),
            'kod': str(ai_data.get('kod') or '').strip(),
            'typ': str(ai_data.get('typ') or 'SUROVINA').strip().upper() or 'SUROVINA',
            'jednotka': str(ai_data.get('jednotka') or 'kg').strip(),
            'minimalna_zasoba': str(ai_data.get('minimalna_zasoba') if ai_data.get('minimalna_zasoba') is not None else 0),
            'cena_za_jednotku': str(ai_data.get('cena_za_jednotku') if ai_data.get('cena_za_jednotku') is not None else 0),
            'priemer_mm': str(ai_data.get('priemer_mm') if ai_data.get('priemer_mm') is not None else 0),
            'tyc_dlzka_m': str(ai_data.get('tyc_dlzka_m') if ai_data.get('tyc_dlzka_m') is not None else 0),
            'kg_na_meter': str(ai_data.get('kg_na_meter') if ai_data.get('kg_na_meter') is not None else 0),
            'aktualna_zasoba': str(ai_data.get('aktualna_zasoba') if ai_data.get('aktualna_zasoba') is not None else 0),
            'poznamka': str(ai_data.get('poznamka') or '').strip(),
        },
        raw_response=ai_response.get('raw_text') or '',
        created_by=request.user,
    )

    if not navrh.navrh_data.get('nazov') or not navrh.navrh_data.get('kod'):
        return _api_error(
            'AI návrh je neúplný (chýba názov alebo kód). Skús presnejší dopyt alebo konkrétnu URL.',
            navrh=_serialize_material_ai_navrh(navrh),
        )

    return _api_ok('AI návrh pripravený. Skontroluj a potvrď uloženie do skladu.', navrh=_serialize_material_ai_navrh(navrh))


@require_POST
@login_required
@permission_required("core.add_material", raise_exception=True)
def ai_material_navrh_potvrdit(request, pk):
    if not _is_ai_material_enabled():
        return _api_error('AI návrhy materiálu sú momentálne vypnuté.')

    payload = _get_json_body(request)
    if payload is None:
        return _api_error('Neplatný JSON payload.')

    navrh = get_object_or_404(MaterialAINavrh, pk=pk)
    if navrh.stav != 'DRAFT':
        return _api_error('Tento návrh už bol spracovaný.')

    typ = str(payload.get('typ') or 'SUROVINA').strip().upper()
    if typ not in {'SUROVINA', 'POLOTOVAR', 'KOMPONENT'}:
        typ = 'SUROVINA'

    nazov = str(payload.get('nazov') or '').strip()
    kod = str(payload.get('kod') or '').strip()
    if not nazov or not kod:
        return _api_error('Názov a kód materiálu sú povinné.')

    if Material.objects.filter(kod=kod).exists():
        return _api_error(f'Materiál s kódom "{kod}" už existuje. Uprav kód a skús znova.')

    material = Material.objects.create(
        nazov=nazov,
        kod=kod,
        typ=typ,
        jednotka=str(payload.get('jednotka') or 'kg').strip() or 'kg',
        minimalna_zasoba=_safe_decimal(payload.get('minimalna_zasoba')),
        cena_za_jednotku=_safe_decimal(payload.get('cena_za_jednotku')),
        priemer_mm=_safe_decimal(payload.get('priemer_mm')),
        tyc_dlzka_m=_safe_decimal(payload.get('tyc_dlzka_m')),
        kg_na_meter=_safe_decimal(payload.get('kg_na_meter')),
        aktualna_zasoba=_safe_decimal(payload.get('aktualna_zasoba')),
    )

    navrh.stav = 'APPROVED'
    navrh.approved_by = request.user
    navrh.approved_at = timezone.now()
    navrh.material = material
    navrh.navrh_data = {
        'nazov': material.nazov,
        'kod': material.kod,
        'typ': material.typ,
        'jednotka': material.jednotka,
        'minimalna_zasoba': str(material.minimalna_zasoba),
        'cena_za_jednotku': str(material.cena_za_jednotku),
        'priemer_mm': str(material.priemer_mm),
        'tyc_dlzka_m': str(material.tyc_dlzka_m),
        'kg_na_meter': str(material.kg_na_meter),
        'aktualna_zasoba': str(material.aktualna_zasoba),
        'poznamka': str(payload.get('poznamka') or ''),
    }
    navrh.save(update_fields=['stav', 'approved_by', 'approved_at', 'material', 'navrh_data', 'updated_at'])

    return _api_ok(
        f'Materiál "{material.nazov}" bol vytvorený z AI návrhu.',
        material_pk=material.pk,
        edit_url=reverse('upravit_material', kwargs={'pk': material.pk}),
    )


# ========================================
# WEB FORMULÁRE PRE OBJEDNÁVKY A KONTRAKTY
# ========================================

_HANGSLER_IMPORT_FIELD_ALIASES = {
    'zakaznik': {'odberatel', 'odberatel nazov', 'zakaznik', 'customer'},
    'cislo_objednavky_zakaznika': {'cislo', 'cislo zakazky', 'cislo objednavky', 'po', 'po cislo'},
    'produkt_text': {'text polozky', 'polozka', 'produkt', 'nazov produktu', 'part'},
    'datum_pozadovane': {'dat dod', 'datum dodania', 'termin dodania', 'delivery date'},
    'mnozstvo': {'pocet mj', 'mnozstvo', 'ks', 'qty'},
}

_TARGET_CUSTOMER_TOKENS = ('hangsler', 'hengstler')


def _normalize_import_text(value):
    normalized = unicodedata.normalize('NFD', str(value or ''))
    normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = ''.join(ch if ch.isalnum() else ' ' for ch in normalized.lower())
    return re.sub(r'\s+', ' ', normalized).strip()


def _is_target_customer(value):
    normalized = _normalize_import_text(value)
    return any(token in normalized for token in _TARGET_CUSTOMER_TOKENS)


def _normalize_compact_token(value):
    return ''.join(ch for ch in str(value or '') if ch.isalnum()).upper()


def _extract_pdf_order_metadata(uploaded_pdf):
    import pypdf

    original_name = str(getattr(uploaded_pdf, 'name', '') or '')
    raw_bytes = uploaded_pdf.read()
    text_parts = []

    reader = pypdf.PdfReader(io.BytesIO(raw_bytes))
    for page in reader.pages:
        page_text = page.extract_text() or ''
        if page_text:
            text_parts.append(page_text)
    text = '\n'.join(text_parts)

    filename_po = ''
    filename_date = None

    po_match_filename = re.search(r'purchase\s+order\s+nr\.?\s*(\d+)', original_name, re.IGNORECASE)
    if po_match_filename:
        filename_po = po_match_filename.group(1)

    date_match_filename = re.search(r'on\s*(\d{2}\.\d{2}\.\d{4})', original_name, re.IGNORECASE)
    if date_match_filename:
        try:
            filename_date = _parse_import_date(date_match_filename.group(1))
        except ValueError:
            filename_date = None

    po_match_text = re.search(r'\b(450\d{7})\b', text)
    po_number = po_match_text.group(1) if po_match_text else filename_po

    date_text_match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', text)
    order_date = filename_date
    if date_text_match:
        try:
            order_date = _parse_import_date(date_text_match.group(1))
        except ValueError:
            pass

    product_match = re.search(r'00010\s+(\d{4,})', text, re.IGNORECASE)
    product_code = product_match.group(1) if product_match else ''

    qty_match = re.search(r'\n\s*([0-9]+(?:[\.,][0-9]+)?)\s*ks\b', text, re.IGNORECASE)
    quantity = None
    if qty_match:
        try:
            quantity = int(Decimal(qty_match.group(1).replace(',', '.')))
        except (InvalidOperation, ValueError):
            quantity = None

    contract_number = ''
    contract_patterns = [
        r'(?i)(?:č[ií]slo|cislo)\s*kontraktu?\s*[:#\-]?\s*([A-Z0-9\-/]{4,})',
        r'(?i)kontraktu?\s*[:#\-]?\s*([A-Z0-9\-/]{4,})',
    ]
    for pattern in contract_patterns:
        contract_match = re.search(pattern, text)
        if contract_match:
            contract_number = contract_match.group(1).strip().upper()
            break

    return {
        'file_name': original_name,
        'po_number': po_number,
        'order_date': order_date,
        'product_code_norm': _normalize_compact_token(product_code),
        'quantity': quantity,
        'contract_number': contract_number,
    }


def _extract_pdf_entries_from_uploads(pdf_files):
    entries = []
    parse_errors = []

    for uploaded_pdf in pdf_files or []:
        try:
            entry = _extract_pdf_order_metadata(uploaded_pdf)
        except Exception as exc:
            parse_errors.append(f'PDF {getattr(uploaded_pdf, "name", "neznámy")}: {exc}')
            continue

        if not entry.get('order_date') or not entry.get('product_code_norm') or not entry.get('quantity'):
            continue
        entries.append(entry)

    return entries, parse_errors


def _pair_pdf_entries_to_grouped_rows(grouped_rows, pdf_entries):
    by_key = defaultdict(list)
    for entry in pdf_entries:
        key = (
            entry['order_date'],
            entry['product_code_norm'],
            int(entry['quantity']),
        )
        by_key[key].append(entry)

    matched = 0
    for grouped in grouped_rows.values():
        key = (
            grouped['datum_pozadovane'],
            _normalize_compact_token(grouped['produkt'].cislo_dielu),
            int(grouped['mnozstvo']),
        )
        candidates = by_key.get(key) or []
        if not candidates:
            continue
        grouped['matched_pdf'] = candidates.pop(0)
        matched += 1

    unmatched_count = sum(len(items) for items in by_key.values())
    return matched, unmatched_count


def _resolve_kontrakt_from_pdf(pdf_entry):
    contract_number = str((pdf_entry or {}).get('contract_number') or '').strip()
    if not contract_number:
        return None

    exact = Kontrakt.objects.filter(cislo_kontraktu__iexact=contract_number).first()
    if exact:
        return exact

    contract_norm = _normalize_compact_token(contract_number)
    for kontrakt in Kontrakt.objects.only('id', 'cislo_kontraktu'):
        if _normalize_compact_token(kontrakt.cislo_kontraktu) == contract_norm:
            return kontrakt
    return None


def _import_hangsler_orders(uploaded_csv, pdf_files=None):
    payload = _build_hangsler_import_rows(uploaded_csv)
    pdf_entries, pdf_errors = _extract_pdf_entries_from_uploads(pdf_files or [])

    grouped_rows = {}
    for parsed in payload['parsed_rows']:
        key = (
            _normalize_import_text(parsed['zakaznik']),
            parsed['produkt'].pk,
            parsed['datum_pozadovane'],
        )
        if key not in grouped_rows:
            grouped_rows[key] = parsed.copy()
            continue

        grouped_rows[key]['mnozstvo'] += parsed['mnozstvo']
        grouped_rows[key]['is_completed'] = grouped_rows[key].get('is_completed', False) or parsed.get('is_completed', False)
        if parsed['cislo_objednavky_zakaznika']:
            grouped_rows[key]['cislo_objednavky_zakaznika'] = parsed['cislo_objednavky_zakaznika']

    matched_pdf, unmatched_pdf = _pair_pdf_entries_to_grouped_rows(grouped_rows, pdf_entries)

    created_count = 0
    updated_count = 0
    kontrakt_assigned_count = 0

    with transaction.atomic():
        for grouped in grouped_rows.values():
            matched_pdf_entry = grouped.get('matched_pdf') or {}
            kontrakt = _resolve_kontrakt_from_pdf(matched_pdf_entry)

            po_number = str(matched_pdf_entry.get('po_number') or '').strip()
            if po_number:
                grouped['cislo_objednavky_zakaznika'] = po_number

            existing = (
                Objednavka.objects
                .filter(produkt=grouped['produkt'], datum_pozadovane=grouped['datum_pozadovane'])
                .filter(zakaznik__iexact=grouped['zakaznik'])
                .order_by('id')
                .first()
            )

            if existing:
                existing.mnozstvo = grouped['mnozstvo']
                existing.zakaznik = grouped['zakaznik']
                update_fields = ['mnozstvo', 'zakaznik']
                if grouped.get('cislo_objednavky_zakaznika'):
                    existing.cislo_objednavky_zakaznika = grouped['cislo_objednavky_zakaznika']
                    update_fields.append('cislo_objednavky_zakaznika')
                if kontrakt:
                    existing.kontrakt = kontrakt
                    update_fields.append('kontrakt')
                    kontrakt_assigned_count += 1
                if grouped.get('is_completed'):
                    existing.stav = 'hotovo'
                    existing.vyrobene_mnozstvo = grouped['mnozstvo']
                    update_fields.extend(['stav', 'vyrobene_mnozstvo'])
                existing.save(update_fields=update_fields)
                updated_count += 1
                continue

            Objednavka.objects.create(
                zakaznik=grouped['zakaznik'],
                cislo_objednavky_zakaznika=grouped.get('cislo_objednavky_zakaznika', ''),
                produkt=grouped['produkt'],
                mnozstvo=grouped['mnozstvo'],
                datum_pozadovane=grouped['datum_pozadovane'],
                stav='hotovo' if grouped.get('is_completed') else 'nova',
                vyrobene_mnozstvo=grouped['mnozstvo'] if grouped.get('is_completed') else 0,
                kontrakt=kontrakt,
            )
            if kontrakt:
                kontrakt_assigned_count += 1
            created_count += 1

    row_errors = list(payload['row_errors']) + pdf_errors
    return {
        'total_rows': payload['total_rows'],
        'filtered_out': payload['filtered_out'],
        'created_count': created_count,
        'updated_count': updated_count,
        'error_rows': row_errors,
        'matched_pdf': matched_pdf,
        'unmatched_pdf': unmatched_pdf,
        'kontrakt_assigned': kontrakt_assigned_count,
    }


def _decode_csv_upload(uploaded_file):
    payload = uploaded_file.read()
    for encoding in ('utf-8-sig', 'cp1250', 'latin-1'):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError('Súbor sa nepodarilo dekódovať. Uložte export ako CSV UTF-8.')


def _detect_csv_delimiter(csv_text):
    sample = csv_text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t|')
        return dialect.delimiter
    except csv.Error:
        return ';'


def _map_import_headers(fieldnames):
    if not fieldnames:
        raise ValueError('CSV súbor neobsahuje hlavičku.')

    normalized_headers = {
        _normalize_import_text(field_name): field_name
        for field_name in fieldnames
        if str(field_name or '').strip()
    }

    mapped = {}
    missing = []
    for internal_key, aliases in _HANGSLER_IMPORT_FIELD_ALIASES.items():
        source_header = None
        for alias in aliases:
            source_header = normalized_headers.get(_normalize_import_text(alias))
            if source_header:
                break
        if source_header:
            mapped[internal_key] = source_header
        else:
            missing.append(internal_key)

    if missing:
        missing_labels = {
            'zakaznik': 'Odberateľ',
            'cislo_objednavky_zakaznika': 'číslo',
            'produkt_text': 'text položky',
            'datum_pozadovane': 'Dát.dod.',
            'mnozstvo': 'počet mj.',
        }
        missing_text = ', '.join(missing_labels[item] for item in missing)
        raise ValueError(f'CSV hlavička neobsahuje povinné stĺpce: {missing_text}.')

    return mapped


def _parse_import_date(value):
    text_value = str(value or '').strip()
    if not text_value:
        raise ValueError('Chýba dátum dodania.')

    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    try:
        serial = int(Decimal(text_value.replace(',', '.')))
        if serial > 59:
            return date(1899, 12, 30) + timedelta(days=serial)
    except (InvalidOperation, ValueError):
        pass

    raise ValueError(f'Neplatný dátum dodania: {text_value}')


def _parse_import_quantity(value):
    text_value = str(value or '').strip().replace(' ', '').replace(',', '.')
    if not text_value:
        raise ValueError('Chýba množstvo.')

    quantity = Decimal(text_value)
    if quantity <= 0:
        raise ValueError('Množstvo musí byť väčšie ako 0.')
    if quantity != quantity.to_integral_value():
        raise ValueError('Množstvo musí byť celé číslo.')

    return int(quantity)


def _normalize_product_lookup(value):
    normalized = _normalize_import_text(value)
    return ''.join(ch for ch in normalized if ch.isalnum())


def _resolve_product_for_import(raw_text, cache):
    lookup_value = str(raw_text or '').strip()
    if not lookup_value:
        raise ValueError('Chýba hodnota text položky (produkt).')

    cached = cache.get(lookup_value)
    if cached is not None:
        return cached

    exact_matches = list(
        Produkt.objects.filter(
            Q(cislo_dielu__iexact=lookup_value)
            | Q(index__iexact=lookup_value)
            | Q(nazov__iexact=lookup_value)
        ).order_by('cislo_dielu', 'id')
    )
    if len(exact_matches) == 1:
        cache[lookup_value] = exact_matches[0]
        return exact_matches[0]
    if len(exact_matches) > 1:
        raise ValueError(
            f'Nejednoznačný produkt pre "{lookup_value}". Nájdite podľa presného čísla dielu.'
        )

    contains_matches = list(
        Produkt.objects.filter(
            Q(cislo_dielu__icontains=lookup_value)
            | Q(nazov__icontains=lookup_value)
        ).order_by('cislo_dielu', 'id')[:2]
    )
    if len(contains_matches) == 1:
        cache[lookup_value] = contains_matches[0]
        return contains_matches[0]
    if len(contains_matches) > 1:
        raise ValueError(
            f'Viac produktov zodpovedá textu "{lookup_value}". Použite presnejší text položky.'
        )

    # Fallback for MRP exports where product code may include spaces or separators.
    normalized_lookup = _normalize_product_lookup(lookup_value)
    if normalized_lookup:
        normalized_matches = []
        for product in Produkt.objects.only('id', 'cislo_dielu', 'index').order_by('id'):
            code_value = _normalize_product_lookup(product.cislo_dielu)
            index_value = _normalize_product_lookup(product.index)
            if normalized_lookup in {code_value, index_value}:
                normalized_matches.append(product)
                if len(normalized_matches) > 1:
                    break

        if len(normalized_matches) == 1:
            cache[lookup_value] = normalized_matches[0]
            return normalized_matches[0]
        if len(normalized_matches) > 1:
            raise ValueError(
                f'Nejednoznačný produkt pre "{lookup_value}" po normalizácii medzier/znakov.'
            )

    raise ValueError(f'Produkt pre "{lookup_value}" sa v ERP nenašiel.')


def _build_hangsler_import_rows(uploaded_file):
    csv_text = _decode_csv_upload(uploaded_file)
    delimiter = _detect_csv_delimiter(csv_text)
    reader = csv.DictReader(csv_text.splitlines(), delimiter=delimiter)
    try:
        mapped_headers = _map_import_headers(reader.fieldnames)
    except ValueError:
        return _build_hangsler_import_rows_block_csv(csv_text)

    parsed_rows = []
    row_errors = []
    total_rows = 0
    filtered_out = 0
    product_cache = {}

    for row_index, row in enumerate(reader, start=2):
        row_values = [str(value or '').strip() for value in row.values()]
        if not any(row_values):
            continue

        total_rows += 1
        customer_raw = str(row.get(mapped_headers['zakaznik']) or '').strip()
        if not _is_target_customer(customer_raw):
            filtered_out += 1
            continue

        try:
            product = _resolve_product_for_import(row.get(mapped_headers['produkt_text']), product_cache)
            due_date = _parse_import_date(row.get(mapped_headers['datum_pozadovane']))
            quantity = _parse_import_quantity(row.get(mapped_headers['mnozstvo']))
            customer_order_no = str(row.get(mapped_headers['cislo_objednavky_zakaznika']) or '').strip()
        except (ValueError, InvalidOperation) as exc:
            row_errors.append(f'Riadok {row_index}: {exc}')
            continue

        parsed_rows.append({
            'zakaznik': customer_raw,
            'produkt': product,
            'datum_pozadovane': due_date,
            'mnozstvo': quantity,
            'cislo_objednavky_zakaznika': customer_order_no,
        })

    return {
        'parsed_rows': parsed_rows,
        'row_errors': row_errors,
        'total_rows': total_rows,
        'filtered_out': filtered_out,
    }


def _build_hangsler_import_rows_block_csv(csv_text):
    parsed_rows = []
    row_errors = []
    total_rows = 0
    filtered_out = 0
    product_cache = {}

    current_customer = ''
    current_order_no = ''
    current_order_date = None
    current_order_completed = False
    product_col_index = None
    quantity_col_index = None

    for row_index, line in enumerate(csv_text.splitlines(), start=1):
        cols = [str(col or '').strip() for col in line.split(';')]
        if not any(cols):
            continue

        normalized_line = _normalize_import_text(' '.join(cols))
        first_col = cols[0] if cols else ''
        normalized_first = _normalize_import_text(first_col)

        if normalized_first.startswith('objednavka '):
            match = re.search(r'(\d+)', first_col)
            current_order_no = match.group(1) if match else ''
            current_order_completed = False
            continue

        if normalized_first.startswith('odberatel') and ':' in first_col:
            current_customer = first_col.split(':', 1)[1].strip()
            continue

        if normalized_first.startswith('zo dna'):
            date_match = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})', first_col)
            if date_match:
                try:
                    current_order_date = _parse_import_date(date_match.group(1))
                except ValueError:
                    current_order_date = None

            status_text = ''
            for col in cols:
                if _normalize_import_text(col).startswith('stav') and ':' in col:
                    status_text = col.split(':', 1)[1].strip()
                    break
            if status_text:
                current_order_completed = 'vybaven' in _normalize_import_text(status_text)
            continue

        if 'cislo karty' in normalized_line and 'objednanych' in normalized_line:
            normalized_cols = [_normalize_import_text(col) for col in cols]
            for idx, header in enumerate(normalized_cols):
                if header == 'nazov':
                    product_col_index = idx
                if header == 'objednanych':
                    quantity_col_index = idx
            continue

        if product_col_index is None or quantity_col_index is None:
            continue
        if len(cols) <= max(product_col_index, quantity_col_index):
            continue

        product_text = cols[product_col_index]
        quantity_text = cols[quantity_col_index]
        if not product_text or not quantity_text:
            continue

        total_rows += 1
        if not _is_target_customer(current_customer):
            filtered_out += 1
            continue
        if not current_order_date:
            row_errors.append(f'Riadok {row_index}: Chýba dátum objednávky (Zo dňa).')
            continue

        code_match = re.match(r'\s*(\d{4,})', product_text)
        product_lookup = code_match.group(1) if code_match else product_text

        try:
            product = _resolve_product_for_import(product_lookup, product_cache)
            quantity = _parse_import_quantity(quantity_text)
        except (ValueError, InvalidOperation) as exc:
            row_errors.append(f'Riadok {row_index}: {exc}')
            continue

        parsed_rows.append({
            'zakaznik': current_customer,
            'produkt': product,
            'datum_pozadovane': current_order_date,
            'mnozstvo': quantity,
            'cislo_objednavky_zakaznika': current_order_no,
            'is_completed': current_order_completed,
        })

    return {
        'parsed_rows': parsed_rows,
        'row_errors': row_errors,
        'total_rows': total_rows,
        'filtered_out': filtered_out,
    }


@login_required
@require_POST
def import_objednavok_hangsler_csv(request):
    from django.contrib import messages

    if not _can_import_orders(request.user):
        messages.error(request, '❌ Nemáte oprávnenie na import objednávok.')
        return redirect('plan_vyroby')

    uploaded_file = request.FILES.get('subor')
    if not uploaded_file:
        messages.error(request, '❌ Nahrajte platný CSV súbor (UTF-8, separator ;).')
        return redirect('nova_objednavka')

    filename = str(uploaded_file.name or '').lower()
    if not filename.endswith('.csv'):
        messages.error(request, '❌ Nepodporovaný formát. Použite súbor .csv.')
        return redirect('nova_objednavka')

    try:
        summary = _import_hangsler_orders(uploaded_file, request.FILES.getlist('pdf_files'))
    except ValueError as exc:
        messages.error(request, f'❌ Import sa nepodaril: {exc}')
        return redirect('nova_objednavka')

    error_count = len(summary['error_rows'])
    if error_count:
        for row_error in summary['error_rows'][:8]:
            messages.error(request, f'❌ {row_error}')
        if error_count > 8:
            messages.error(request, f'❌ Ďalšie chyby: {error_count - 8}')

    messages.success(
        request,
        (
            '✅ Import Hangsler dokončený. '
            f'Načítané riadky: {summary["total_rows"]}, '
            f'odfiltrované: {summary["filtered_out"]}, '
            f'vytvorené: {summary["created_count"]}, '
            f'aktualizované: {summary["updated_count"]}, '
            f'kontrakty priradené: {summary["kontrakt_assigned"]}, '
            f'spárované PDF: {summary["matched_pdf"]}, '
            f'nespárované PDF: {summary["unmatched_pdf"]}, '
            f'chybné: {error_count}.'
        ),
    )
    return redirect('nova_objednavka')


@login_required
def import_objednavok_mrp_pdf(request):
    from django.contrib import messages

    if not _can_import_orders(request.user):
        messages.error(request, '❌ Nemáte oprávnenie na import objednávok.')
        return redirect('plan_vyroby')

    if request.method == 'POST':
        uploaded_csv = request.FILES.get('csv_file')
        pdf_files = request.FILES.getlist('pdf_files')

        if not uploaded_csv:
            messages.error(request, '❌ Nahrajte CSV export z MRP.')
            return redirect('import_objednavok_mrp_pdf')

        if not str(uploaded_csv.name or '').lower().endswith('.csv'):
            messages.error(request, '❌ Prvý súbor musí byť .csv export z MRP.')
            return redirect('import_objednavok_mrp_pdf')

        try:
            summary = _import_hangsler_orders(uploaded_csv, pdf_files)
        except ValueError as exc:
            messages.error(request, f'❌ Import sa nepodaril: {exc}')
            return redirect('import_objednavok_mrp_pdf')

        error_count = len(summary['error_rows'])
        if error_count:
            for row_error in summary['error_rows'][:8]:
                messages.error(request, f'❌ {row_error}')
            if error_count > 8:
                messages.error(request, f'❌ Ďalšie chyby: {error_count - 8}')

        messages.success(
            request,
            (
                '✅ Import MRP + objednávky dokončený. '
                f'Načítané riadky: {summary["total_rows"]}, '
                f'odfiltrované: {summary["filtered_out"]}, '
                f'vytvorené: {summary["created_count"]}, '
                f'aktualizované: {summary["updated_count"]}, '
                f'kontrakty priradené: {summary["kontrakt_assigned"]}, '
                f'spárované PDF: {summary["matched_pdf"]}, '
                f'nespárované PDF: {summary["unmatched_pdf"]}, '
                f'chybné: {error_count}.'
            ),
        )
        return redirect('import_objednavok_mrp_pdf')

    return render(request, 'core/import_mrp_pdf.html')

@login_required
@permission_required("core.add_objednavka", raise_exception=True)
def nova_objednavka(request):
    """Vytvorenie novej objednávky cez webový formulár"""
    from .forms import ObjednavkaForm, ObjednavkaImportCSVForm
    from django.contrib import messages
    from .models import SkladHotovychDielov
    from .models import Produkt
    
    if request.method == 'POST':
        form = ObjednavkaForm(request.POST)
        if form.is_valid():
            produkt = form.cleaned_data.get('produkt')
            mnozstvo = form.cleaned_data.get('mnozstvo')
            shortage_message = _build_material_shortage_message(produkt, mnozstvo)
            if shortage_message:
                form.add_error(None, shortage_message)
                messages.error(request, f'❌ {shortage_message}')
            else:
                objednavka = form.save(commit=False)
                objednavka.stav = 'nova'
                objednavka.vyrobene_mnozstvo = 0
                objednavka.save()
                
                messages.success(request, f'✅ Objednávka #{objednavka.cislo_objednavky} bola úspešne vytvorená!')
                return redirect('plan_vyroby')
    else:
        form = ObjednavkaForm()
    
    sklad_map = {
        sklad.produkt_id: sklad.mnozstvo
        for sklad in SkladHotovychDielov.objects.all()
    }

    produkt_material_map = {}
    for produkt in Produkt.objects.select_related('material_ref'):
        material = produkt.material_ref
        if not material:
            continue
        produkt_material_map[produkt.id] = {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'jednotka': material.jednotka,
            'zasoba': float(material.aktualna_zasoba),
            'minimalna_zasoba': float(material.minimalna_zasoba),
            'dlzka_na_kus_mm': float(produkt.dlzka_na_kus_mm or 0),
            'tyc_dlzka_m': float(material.tyc_dlzka_m or 0),
            'kg_na_meter': float(material.kg_na_meter or 0),
            'priemer_mm': float(material.priemer_mm or 0),
        }

    context = {
        'form': form,
        'import_form': ObjednavkaImportCSVForm(),
        'title': 'Nová objednávka',
        'submit_text': 'Vytvoriť objednávku',
        'sklad_map': sklad_map,
        'produkt_material_map': produkt_material_map,
    }
    return render(request, 'core/nova_objednavka.html', context)


@login_required
@permission_required("core.add_kontrakt", raise_exception=True)
def novy_kontrakt(request):
    """Vytvorenie nového kontraktu cez webový formulár"""
    from .forms import KontraktForm
    from django.contrib import messages

    if request.method == 'POST':
        form = KontraktForm(request.POST)
        if form.is_valid():
            kontrakt = form.save()
            messages.success(request, f'✅ Kontrakt #{kontrakt.cislo_kontraktu} bol úspešne vytvorený!')
            return redirect('plan_vyroby')
    else:
        form = KontraktForm()

    produkt_material_map = {}
    for produkt in Produkt.objects.select_related('material_ref'):
        material = produkt.material_ref
        if not material:
            continue
        produkt_material_map[produkt.id] = {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'jednotka': material.jednotka,
            'zasoba': float(material.aktualna_zasoba),
            'minimalna_zasoba': float(material.minimalna_zasoba),
            'dlzka_na_kus_mm': float(produkt.dlzka_na_kus_mm or 0),
            'tyc_dlzka_m': float(material.tyc_dlzka_m or 0),
            'kg_na_meter': float(material.kg_na_meter or 0),
            'priemer_mm': float(material.priemer_mm or 0),
        }

    context = {
        'form': form,
        'title': 'Nový kontrakt',
        'submit_text': 'Vytvoriť kontrakt',
        'produkt_material_map': produkt_material_map,
    }
    return render(request, 'core/novy_kontrakt.html', context)


@login_required
@permission_required("core.change_objednavka", raise_exception=True)
def upravit_objednavku(request, pk):
    """Úprava existujúcej objednávky"""
    from .forms import ObjednavkaForm
    from django.contrib import messages
    from .models import SkladHotovychDielov
    from .models import Produkt
    
    objednavka = get_object_or_404(Objednavka, pk=pk)
    
    if request.method == 'POST':
        form = ObjednavkaForm(request.POST, instance=objednavka)
        if form.is_valid():
            produkt = form.cleaned_data.get('produkt')
            mnozstvo = form.cleaned_data.get('mnozstvo')
            shortage_message = _build_material_shortage_message(produkt, mnozstvo)
            if shortage_message:
                form.add_error(None, shortage_message)
                messages.error(request, f'❌ {shortage_message}')
            else:
                form.save()
                messages.success(request, f'✅ Objednávka #{objednavka.cislo_objednavky} bola aktualizovaná!')
                return redirect('detail_zakazky', pk=objednavka.pk)
    else:
        form = ObjednavkaForm(instance=objednavka)
    
    sklad_map = {
        sklad.produkt_id: sklad.mnozstvo
        for sklad in SkladHotovychDielov.objects.all()
    }

    produkt_material_map = {}
    for produkt in Produkt.objects.select_related('material_ref'):
        material = produkt.material_ref
        if not material:
            continue
        produkt_material_map[produkt.id] = {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'jednotka': material.jednotka,
            'zasoba': float(material.aktualna_zasoba),
            'minimalna_zasoba': float(material.minimalna_zasoba),
            'dlzka_na_kus_mm': float(produkt.dlzka_na_kus_mm or 0),
            'tyc_dlzka_m': float(material.tyc_dlzka_m or 0),
            'kg_na_meter': float(material.kg_na_meter or 0),
            'priemer_mm': float(material.priemer_mm or 0),
        }

    context = {
        'form': form,
        'title': f'Upraviť objednávku #{objednavka.cislo_objednavky}',
        'submit_text': 'Uložiť zmeny',
        'objednavka': objednavka,
        'sklad_map': sklad_map,
        'produkt_material_map': produkt_material_map,
    }
    return render(request, 'core/nova_objednavka.html', context)


@login_required
@permission_required("core.change_kontrakt", raise_exception=True)
def upravit_kontrakt(request, pk):
    """Úprava existujúceho kontraktu"""
    from .forms import KontraktForm
    from django.contrib import messages

    kontrakt = get_object_or_404(Kontrakt, pk=pk)

    if request.method == 'POST':
        form = KontraktForm(request.POST, instance=kontrakt)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Kontrakt #{kontrakt.cislo_kontraktu} bol aktualizovaný!')
            return redirect('plan_vyroby')
    else:
        form = KontraktForm(instance=kontrakt)

    produkt_material_map = {}
    for produkt in Produkt.objects.select_related('material_ref'):
        material = produkt.material_ref
        if not material:
            continue
        produkt_material_map[produkt.id] = {
            'material_nazov': material.nazov,
            'material_kod': material.kod,
            'jednotka': material.jednotka,
            'zasoba': float(material.aktualna_zasoba),
            'minimalna_zasoba': float(material.minimalna_zasoba),
            'dlzka_na_kus_mm': float(produkt.dlzka_na_kus_mm or 0),
            'tyc_dlzka_m': float(material.tyc_dlzka_m or 0),
            'kg_na_meter': float(material.kg_na_meter or 0),
            'priemer_mm': float(material.priemer_mm or 0),
        }

    context = {
        'form': form,
        'title': f'Upraviť kontrakt #{kontrakt.cislo_kontraktu}',
        'submit_text': 'Uložiť zmeny',
        'kontrakt': kontrakt,
        'produkt_material_map': produkt_material_map,
    }
    return render(request, 'core/novy_kontrakt.html', context)


@login_required
@permission_required("core.change_kontrakt", raise_exception=True)
def inicializacia_dodavok_kontraktov(request):
    """Jednorazové nastavenie už dodaných kusov pre otvorené kontrakty pred spustením systému."""
    from django.contrib import messages

    otvorene_kontrakty = list(
        Kontrakt.objects
        .filter(zostavajuce_mnozstvo__gt=0)
        .select_related('produkt')
        .order_by('produkt__cislo_dielu', 'cislo_kontraktu')
    )

    riadky = []
    for kontrakt in otvorene_kontrakty:
        uz_dodane = max(kontrakt.pocet_kusov_celkovo - kontrakt.zostavajuce_mnozstvo, 0)
        riadky.append({
            'kontrakt': kontrakt,
            'uz_dodane': uz_dodane,
            'zostava': kontrakt.zostavajuce_mnozstvo,
        })

    if request.method == 'POST':
        chyby = []
        zmeny = []

        for riadok in riadky:
            kontrakt = riadok['kontrakt']
            field_name = f'dodane_{kontrakt.id}'
            raw_value = (request.POST.get(field_name) or '').strip()

            if raw_value == '':
                dodane = riadok['uz_dodane']
            else:
                try:
                    dodane = int(raw_value)
                except ValueError:
                    chyby.append(f'Kontrakt #{kontrakt.cislo_kontraktu}: neplatné číslo.')
                    continue

            if dodane < 0:
                chyby.append(f'Kontrakt #{kontrakt.cislo_kontraktu}: dodané kusy nemôžu byť záporné.')
                continue
            if dodane > kontrakt.pocet_kusov_celkovo:
                chyby.append(
                    f'Kontrakt #{kontrakt.cislo_kontraktu}: dodané ({dodane}) je viac než celkom ({kontrakt.pocet_kusov_celkovo}).'
                )
                continue

            nove_zostava = kontrakt.pocet_kusov_celkovo - dodane
            zmeny.append((kontrakt, nove_zostava))

        if chyby:
            for chyba in chyby:
                messages.error(request, f'❌ {chyba}')
        else:
            with transaction.atomic():
                for kontrakt, nove_zostava in zmeny:
                    kontrakt.zostavajuce_mnozstvo = nove_zostava
                    kontrakt.save(update_fields=['zostavajuce_mnozstvo'])

            messages.success(request, '✅ Už dodané kusy boli uložené a otvorené kontrakty sú zosynchronizované.')
            return redirect('inicializacia_dodavok_kontraktov')

    context = {
        'riadky': riadky,
    }
    return render(request, 'core/inicializacia_dodavok_kontraktov.html', context)


# ========================================
# WEBOVÉ ROZHRANIA PRE STROJE
# ========================================

@login_required
@permission_required("core.add_stroj", raise_exception=True)
def novy_stroj(request):
    """Vytvorenie nového stroja"""
    from .forms import StrojForm
    from django.contrib import messages

    if request.method == "POST":
        form = StrojForm(request.POST, request.FILES)  # DÔLEŽITÉ: request.FILES
        if form.is_valid():
            stroj = form.save()
            messages.success(request, f'✅ Stroj "{stroj.nazov}" bol vytvorený!')
            return redirect("zoznam_strojov")
    else:
        form = StrojForm()

    context = {
        "form": form,
        "title": "Nový stroj",
        "submit_text": "Vytvoriť stroj",
    }

    return render(request, "core/novy_stroj.html", context)


@login_required
@permission_required("core.change_stroj", raise_exception=True)
def upravit_stroj(request, pk):
    """Úprava existujúceho stroja"""
    from .forms import StrojForm
    from django.contrib import messages

    stroj = get_object_or_404(Stroj, pk=pk)

    if request.method == "POST":
        form = StrojForm(request.POST, request.FILES, instance=stroj)  # DÔLEŽITÉ: request.FILES
        if form.is_valid():
            stroj = form.save()
            messages.success(request, f'✅ Stroj "{stroj.nazov}" bol aktualizovaný!')
            return redirect("zoznam_strojov")
    else:
        form = StrojForm(instance=stroj)

    context = {
        "form": form,
        "title": f"Upraviť stroj: {stroj.nazov}",
        "submit_text": "Uložiť zmeny",
        "stroj": stroj,
    }

    return render(request, "core/upravit_stroj.html", context)




# ========================================
# WEBOVÉ ROZHRANIA PRE PRODUKTY
# ========================================

@login_required
@permission_required("core.add_produkt", raise_exception=True)
def novy_produkt(request):
    """Vytvorenie nového produktu"""
    from .forms import ProduktForm
    from django.contrib import messages
    
    if request.method == 'POST':
        form = ProduktForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Produkt "{form.instance.nazov}" bol vytvorený!')
            return redirect('zoznam_produktov')
    else:
        form = ProduktForm()
    
    context = {
        'form': form,
        'title': 'Nový produkt',
        'submit_text': 'Vytvoriť produkt',
    }
    
    return render(request, 'core/form_universal.html', context)

@login_required
@permission_required("core.change_produkt", raise_exception=True)
def upravit_produkt(request, pk):
    """Uprava existujúceho produktu"""
    from .forms import ProduktForm
    from django.contrib import messages
    
    produkt = get_object_or_404(Produkt, pk=pk)
    
    if request.method == 'POST':
        form = ProduktForm(request.POST, request.FILES, instance=produkt)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Produkt "{produkt.nazov}" bol aktualizovaný!')
            return redirect('detail_produkt', pk=produkt.pk)
    else:
        form = ProduktForm(instance=produkt)
    
    context = {
        'form': form,
        'title': f'Upraviť produkt #{produkt.nazov}',
        'submit_text': 'Uložiť zmeny',
        'produkt': produkt,
    }

    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.add_material", raise_exception=True)
def novy_material(request):
    """Vytvorenie nového materiálu"""
    from .forms import MaterialForm
    from django.contrib import messages

    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            messages.success(request, f'✅ Materiál "{material.nazov}" bol vytvorený!')
            return redirect('sklad_materialu')
    else:
        form = MaterialForm()

    context = {
        'form': form,
        'title': 'Nový materiál',
        'submit_text': 'Vytvoriť materiál',
    }

    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.change_material", raise_exception=True)
def upravit_material(request, pk):
    """Úprava existujúceho materiálu"""
    from .forms import MaterialForm
    from django.contrib import messages

    material = get_object_or_404(Material, pk=pk)

    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Materiál "{material.nazov}" bol aktualizovaný!')
            return redirect('sklad_materialu')
    else:
        form = MaterialForm(instance=material)

    context = {
        'form': form,
        'title': f'Upraviť materiál: {material.nazov}',
        'submit_text': 'Uložiť zmeny',
    }

    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.change_skladhotovychdielov", raise_exception=True)
def upravit_sklad_hotovych_dielov(request, pk):
    """Úprava položky skladu hotových dielov"""
    from .forms import SkladHotovychDielovForm
    from django.contrib import messages

    sklad = get_object_or_404(SkladHotovychDielov, pk=pk)

    if request.method == 'POST':
        form = SkladHotovychDielovForm(request.POST, instance=sklad)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Skladová položka "{sklad.produkt.nazov}" bola aktualizovaná!')
            return redirect('sklad_hotovych_dielov')
    else:
        form = SkladHotovychDielovForm(instance=sklad)

    context = {
        'form': form,
        'title': f'Upraviť skladovú položku: {sklad.produkt.nazov}',
        'submit_text': 'Uložiť zmeny',
    }

    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.add_skladhotovychdielov", raise_exception=True)
def novy_sklad_hotovych_dielov(request):
    """Vytvorenie novej položky skladu hotových dielov"""
    from .forms import SkladHotovychDielovForm
    from django.contrib import messages

    if request.method == 'POST':
        form = SkladHotovychDielovForm(request.POST)
        if form.is_valid():
            sklad = form.save()
            messages.success(request, f'✅ Skladová položka "{sklad.produkt.nazov}" bola vytvorená!')
            return redirect('sklad_hotovych_dielov')
    else:
        form = SkladHotovychDielovForm()

    context = {
        'form': form,
        'title': 'Nová skladová položka',
        'submit_text': 'Vytvoriť položku',
    }

    return render(request, 'core/form_universal.html', context)


# ========================================
# WEBOVÉ ROZHRANIA PRE VÝROBNÉ DÁVKY
# ========================================

@login_required
@permission_required("core.add_vyrobnadavka", raise_exception=True)
def nova_vyrobna_davka(request, kontrakt_pk):
    """Vytvorenie novej výrobnej dávky z kontraktu"""
    from .forms import VyrobnaDavkaForm
    from django.contrib import messages
    
    kontrakt = get_object_or_404(Kontrakt, pk=kontrakt_pk)
    
    if request.method == 'POST':
        form = VyrobnaDavkaForm(request.POST)
        if form.is_valid():
            davka = form.save(commit=False)
            davka.kontrakt = kontrakt
            davka.save()
            messages.success(request, f'✅ Výrobná dávka "{davka.cislo_davky}" bola vytvorená!')
            return redirect('detail_kontrakt', pk=kontrakt.pk)
    else:
        form = VyrobnaDavkaForm()
    
    context = {
        'form': form,
        'title': f'Nová výrobná dávka pre kontrakt {kontrakt.cislo_kontraktu}',
        'submit_text': 'Vytvoriť dávku',
        'kontrakt': kontrakt,
    }
    
    return render(request, 'core/form_universal.html', context)

# ========================================
# WEBOVÉ ROZHRANIA PRE SKLAD HOTOVÝCH DIELOV
# ========================================

@login_required
@permission_required("core.add_prijemkahotovychdielov", raise_exception=True)
def nova_prijemka(request):
    """Príjemka hotových dielov na sklad"""
    from .forms import PrijemkaHotovychDielovForm
    from django.contrib import messages
    
    initial = {}
    sklad_id = request.GET.get('sklad')
    if sklad_id:
        initial['sklad'] = sklad_id

    if request.method == 'POST':
        form = PrijemkaHotovychDielovForm(request.POST)
        if form.is_valid():
            prijemka = form.save(commit=False)
            prijemka.operator = request.user
            prijemka.save()
            messages.success(request, f'✅ Príjemka +{prijemka.mnozstvo} ks bola zaznamemaná!')
            return redirect('sklad_hotovych_dielov')
    else:
        form = PrijemkaHotovychDielovForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Nová príjemka',
        'submit_text': 'Naskladniť',
    }
    
    return render(request, 'core/form_universal.html', context)

@login_required
@permission_required("core.add_vydajkahotovychdielov", raise_exception=True)
def nova_vydajka(request):
    """Výdajka hotových dielov zo skladu"""
    from .forms import VydajkaHotovychDielovForm
    from django.contrib import messages
    
    initial = {}
    sklad_id = request.GET.get('sklad')
    if sklad_id:
        initial['sklad'] = sklad_id

    if request.method == 'POST':
        form = VydajkaHotovychDielovForm(request.POST)
        if form.is_valid():
            try:
                vydajka = form.save(commit=False)
                vydajka.operator = request.user
                vydajka.save()
                messages.success(request, f'✅ Výdajka -{vydajka.mnozstvo} ks bola zaznamemaná!')
                return redirect('sklad_hotovych_dielov')
            except ValueError as e:
                messages.error(request, f'❌ {str(e)}')
    else:
        form = VydajkaHotovychDielovForm(initial=initial)
    
    context = {
        'form': form,
        'title': 'Nová výdajka',
        'submit_text': 'Vydať',
    }
    
    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.add_prijemkanasklad", raise_exception=True)
def nova_prijemka_materialu(request):
    """Príjemka materiálu na sklad"""
    from .forms import PrijemkaNaSkladForm
    from django.contrib import messages

    initial = {}
    material_id = request.GET.get('material')
    if material_id:
        initial['material'] = material_id

    if request.method == 'POST':
        form = PrijemkaNaSkladForm(request.POST)
        if form.is_valid():
            prijemka = form.save()
            messages.success(request, f'✅ Príjemka +{prijemka.mnozstvo} {prijemka.material.jednotka} bola zaznamenaná!')
            return redirect('sklad_materialu')
    else:
        form = PrijemkaNaSkladForm(initial=initial)

    context = {
        'form': form,
        'title': 'Nová príjemka materiálu',
        'submit_text': 'Naskladniť',
    }

    return render(request, 'core/form_universal.html', context)


@login_required
@permission_required("core.add_vydajkazoskladu", raise_exception=True)
def nova_vydajka_materialu(request):
    """Výdajka materiálu zo skladu"""
    from .forms import VydajkaZoSkladuForm
    from django.contrib import messages

    initial = {}
    material_id = request.GET.get('material')
    if material_id:
        initial['material'] = material_id

    if request.method == 'POST':
        form = VydajkaZoSkladuForm(request.POST)
        if form.is_valid():
            vydajka = form.save(commit=False)
            vydajka.operator = request.user

            if vydajka.mnozstvo > vydajka.material.aktualna_zasoba:
                messages.error(
                    request,
                    f'❌ Nedostatok materiálu na sklade. Dostupné: {vydajka.material.aktualna_zasoba} {vydajka.material.jednotka}.',
                )
            else:
                vydajka.save()
                messages.success(request, f'✅ Výdajka -{vydajka.mnozstvo} {vydajka.material.jednotka} bola zaznamenaná!')
                return redirect('sklad_materialu')
    else:
        form = VydajkaZoSkladuForm(initial=initial)

    context = {
        'form': form,
        'title': 'Nová výdajka materiálu',
        'submit_text': 'Vydať',
    }

    return render(request, 'core/form_universal.html', context)


# =============================================================================
# ERP DOCUMENTS MODULE
# =============================================================================

import os
import shutil
import mimetypes
from pathlib import Path

from .docs_utils import (
    docs_root, trash_root, tmp_root,
    safe_resolve, to_rel, is_extension_blocked,
    is_docs_admin, resolve_collision, safe_filename,
)


def _docs_require_admin(user):
    """Return JsonResponse 403 or None."""
    if not is_docs_admin(user):
        return JsonResponse({'status': 'error', 'message': 'Nemáte oprávnenie (docs_admin).'}, status=403)
    return None


# Max number of trash entries returned by the trash list endpoint
TRASH_LIST_LIMIT = 200


def _log_doc_action(user, produkt, action, src='', dest='', size=None):
    DocumentAuditLog.objects.create(
        user=user,
        produkt=produkt,
        action=action,
        src_rel_path=src,
        dest_rel_path=dest,
        file_size=size,
    )


# ─── Folder picker tree ─────────────────────────────────────────────────────

@login_required
def docs_tree(request):
    """Return JSON list of child folders for lazy folder-picker tree.

    GET /api/docs/tree/?path=<rel_path>
    rel_path is relative to ERP_DOCS_ROOT; empty = list root children.
    """
    rel_path = request.GET.get('path', '')
    root = docs_root()

    if not root.is_dir():
        return JsonResponse({'status': 'error', 'message': f'ERP_DOCS_ROOT neexistuje: {root}'}, status=500)

    try:
        target = safe_resolve(rel_path, root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná cesta (zakázaný prístup).'}, status=400)

    if not target.is_dir():
        return JsonResponse({'status': 'error', 'message': 'Nie je priečinok.'}, status=400)

    children = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: e.name.lower()):
            if entry.is_dir():
                has_children = any(e.is_dir() for e in entry.iterdir())
                children.append({
                    'name': entry.name,
                    'path': to_rel(entry, root),
                    'has_children': has_children,
                })
    except PermissionError:
        return JsonResponse({'status': 'error', 'message': 'Nemáte prístup k priečinku.'}, status=403)

    return JsonResponse({'status': 'ok', 'path': to_rel(target, root) if target != root else '', 'children': children})


# ─── Set documents_path on product ──────────────────────────────────────────

@login_required
@require_POST
def docs_set_path(request, pk):
    """POST /api/docs/<pk>/set-path/ — save documents_path for a product."""
    err = _docs_require_admin(request.user)
    if err:
        return err

    produkt = get_object_or_404(Produkt, pk=pk)
    root = docs_root()

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': 'error', 'message': 'Neplatný JSON.'}, status=400)

    rel_path = data.get('path', '').strip()

    if rel_path:
        # Validate the path exists and is within root
        try:
            target = safe_resolve(rel_path, root)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Neplatná cesta (zakázaný prístup).'}, status=400)
        if not target.is_dir():
            return JsonResponse({'status': 'error', 'message': 'Zvolená cesta nie je priečinok.'}, status=400)
        # Store normalised (forward-slash) relative path
        stored = to_rel(target, root)
    else:
        stored = ''

    old_path = produkt.documents_path
    produkt.documents_path = stored
    produkt.save(update_fields=['documents_path'])

    _log_doc_action(request.user, produkt, DocumentAuditLog.ACTION_SET_PATH,
                    src=old_path, dest=stored)

    return JsonResponse({'status': 'ok', 'message': 'Cesta k dokumentom bola uložená.', 'path': stored})


# ─── Browse documents ────────────────────────────────────────────────────────

@login_required
def docs_list(request, pk):
    """Return folder listing for a product's documents directory.

    GET /api/docs/<pk>/list/?subpath=<rel_within_product_folder>
    """
    produkt = get_object_or_404(Produkt, pk=pk)
    root = docs_root()

    if not produkt.documents_path:
        return JsonResponse({'status': 'error', 'message': 'Produkt nemá nastavenú cestu k dokumentom.'}, status=400)

    # Resolve product base folder
    try:
        base = safe_resolve(produkt.documents_path, root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná uložená cesta dokumentov.'}, status=400)

    # Optional sub-navigation inside the product folder
    subpath = request.GET.get('subpath', '')
    try:
        target = safe_resolve(subpath, base) if subpath else base
        # Double-check target is still under root (belt-and-suspenders)
        safe_resolve(to_rel(target, root), root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná cesta (zakázaný prístup).'}, status=400)

    if not target.is_dir():
        return JsonResponse({'status': 'error', 'message': 'Priečinok neexistuje.'}, status=404)

    folders = []
    files = []
    try:
        raw_entries = list(target.iterdir())
        raw_entries.sort(key=lambda e: e.name.lower())
        for entry in raw_entries:
            rel_from_base = to_rel(entry, base)
            if entry.is_dir():
                folders.append({'name': entry.name, 'subpath': rel_from_base})
            else:
                stat = entry.stat()
                mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
                files.append({
                    'name': entry.name,
                    'subpath': rel_from_base,
                    'size': stat.st_size,
                    'modified': mtime.astimezone(timezone.get_default_timezone()).strftime('%d.%m.%Y %H:%M'),
                })
    except PermissionError:
        return JsonResponse({'status': 'error', 'message': 'Nemáte prístup k priečinku.'}, status=403)

    # Build breadcrumb: list of {name, subpath} from base to current
    breadcrumb = []
    if subpath:
        parts = Path(subpath.replace('\\', '/')).parts
        accumulated = ''
        for part in parts:
            accumulated = f'{accumulated}/{part}'.lstrip('/')
            breadcrumb.append({'name': part, 'subpath': accumulated})

    return JsonResponse({
        'status': 'ok',
        'base_path': produkt.documents_path,
        'subpath': subpath,
        'breadcrumb': breadcrumb,
        'folders': folders,
        'files': files,
        'is_docs_admin': is_docs_admin(request.user),
    })


# ─── Download file ───────────────────────────────────────────────────────────

@login_required
def docs_download(request, pk):
    """Stream a file download.

    GET /api/docs/<pk>/download/?subpath=<path_within_product_folder>
    """
    produkt = get_object_or_404(Produkt, pk=pk)
    root = docs_root()

    if not produkt.documents_path:
        return HttpResponse('Produkt nemá nastavenú cestu k dokumentom.', status=400)

    subpath = request.GET.get('subpath', '')
    if not subpath:
        return HttpResponse('Chýba subpath.', status=400)

    try:
        base = safe_resolve(produkt.documents_path, root)
        target = safe_resolve(subpath, base)
        safe_resolve(to_rel(target, root), root)  # belt-and-suspenders
    except ValueError:
        return HttpResponse('Neplatná cesta (zakázaný prístup).', status=400)

    if not target.is_file():
        return HttpResponse('Súbor neexistuje.', status=404)

    content_type, _ = mimetypes.guess_type(target.name)
    content_type = content_type or 'application/octet-stream'

    def file_iterator(file_path, chunk_size=8192):
        with open(file_path, 'rb') as f:
            while True:
                chunk_data = f.read(chunk_size)
                if not chunk_data:
                    break
                yield chunk_data

    response = StreamingHttpResponse(file_iterator(target), content_type=content_type)
    from urllib.parse import quote
    encoded_name = quote(target.name, safe='')
    response['Content-Disposition'] = (
        f"attachment; filename=\"{target.name}\"; filename*=UTF-8''{encoded_name}"
    )
    response['Content-Length'] = target.stat().st_size
    return response


# ─── Upload files ────────────────────────────────────────────────────────────

@login_required
@require_POST
def docs_upload(request, pk):
    """Multi-file upload into a product documents subfolder.

    POST /api/docs/<pk>/upload/?subpath=<path_within_product_folder>
    Body: multipart/form-data with files[] field.
    """
    err = _docs_require_admin(request.user)
    if err:
        return err

    produkt = get_object_or_404(Produkt, pk=pk)
    root = docs_root()

    if not produkt.documents_path:
        return JsonResponse({'status': 'error', 'message': 'Produkt nemá nastavenú cestu k dokumentom.'}, status=400)

    subpath = request.GET.get('subpath', '')
    try:
        base = safe_resolve(produkt.documents_path, root)
        target_dir = safe_resolve(subpath, base) if subpath else base
        safe_resolve(to_rel(target_dir, root), root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná cesta (zakázaný prístup).'}, status=400)

    uploaded_files = request.FILES.getlist('files')
    if not uploaded_files:
        return JsonResponse({'status': 'error', 'message': 'Žiadne súbory neboli nahrané.'}, status=400)

    tmp = tmp_root()
    tmp.mkdir(parents=True, exist_ok=True)
    target_dir.mkdir(parents=True, exist_ok=True)

    results = []
    for f in uploaded_files:
        original_name = safe_filename(f.name)

        if is_extension_blocked(original_name):
            results.append({'name': original_name, 'status': 'error', 'message': 'Zakázaná prípona súboru.'})
            continue

        # Write to tmp atomically
        tmp_file = resolve_collision(tmp / original_name)
        try:
            with open(tmp_file, 'wb') as out:
                for chunk in f.chunks():
                    out.write(chunk)

            # Move to destination
            dest_file = resolve_collision(target_dir / original_name)
            shutil.move(str(tmp_file), str(dest_file))

            rel_dest = to_rel(dest_file, root)
            _log_doc_action(request.user, produkt, DocumentAuditLog.ACTION_UPLOAD,
                            src='', dest=rel_dest, size=dest_file.stat().st_size)

            results.append({'name': dest_file.name, 'status': 'ok'})
        except OSError:
            if tmp_file.exists():
                tmp_file.unlink(missing_ok=True)
            results.append({'name': original_name, 'status': 'error', 'message': 'Nahrávanie súboru zlyhalo.'})

    all_ok = all(r['status'] == 'ok' for r in results)
    return JsonResponse({
        'status': 'ok' if all_ok else 'partial',
        'results': results,
    })


# ─── Delete (move to trash) ───────────────────────────────────────────────────

@login_required
@require_POST
def docs_delete(request, pk):
    """Move a file or folder to trash.

    POST /api/docs/<pk>/delete/
    Body JSON: {"subpath": "<path_within_product_folder>"}
    """
    err = _docs_require_admin(request.user)
    if err:
        return err

    produkt = get_object_or_404(Produkt, pk=pk)
    root = docs_root()

    if not produkt.documents_path:
        return JsonResponse({'status': 'error', 'message': 'Produkt nemá nastavenú cestu k dokumentom.'}, status=400)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': 'error', 'message': 'Neplatný JSON.'}, status=400)

    subpath = data.get('subpath', '').strip()
    if not subpath:
        return JsonResponse({'status': 'error', 'message': 'Chýba subpath.'}, status=400)

    try:
        base = safe_resolve(produkt.documents_path, root)
        target = safe_resolve(subpath, base)
        safe_resolve(to_rel(target, root), root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná cesta (zakázaný prístup).'}, status=400)

    if not target.exists():
        return JsonResponse({'status': 'error', 'message': 'Súbor/priečinok neexistuje.'}, status=404)

    # Build trash destination: trash_root / <timestamp>__<user>__<rel_from_root>
    src_rel = to_rel(target, root)
    ts = timezone.now().strftime('%Y%m%d_%H%M%S')
    username = request.user.username
    trash_prefix = f"{ts}__{username}"
    trash_dest = trash_root() / trash_prefix / src_rel

    trash_dest.parent.mkdir(parents=True, exist_ok=True)

    # Handle collision at trash destination
    trash_dest = resolve_collision(trash_dest)

    size = None
    if target.is_file():
        size = target.stat().st_size

    shutil.move(str(target), str(trash_dest))

    dest_rel = to_rel(trash_dest, trash_root())
    _log_doc_action(request.user, produkt, DocumentAuditLog.ACTION_DELETE,
                    src=src_rel, dest=dest_rel, size=size)

    return JsonResponse({'status': 'ok', 'message': 'Presunuté do koša.', 'trash_path': dest_rel})


# ─── Trash list ──────────────────────────────────────────────────────────────

@login_required
def docs_trash_list(request):
    """Return trash audit log entries.

    GET /api/docs/trash/?produkt_pk=<optional>
    """
    qs = DocumentAuditLog.objects.filter(
        action=DocumentAuditLog.ACTION_DELETE
    ).select_related('user', 'produkt').order_by('-timestamp')

    produkt_pk = request.GET.get('produkt_pk')
    if produkt_pk:
        qs = qs.filter(produkt_id=produkt_pk)

    items = [
        {
            'id': log.id,
            'src_rel_path': log.src_rel_path,
            'dest_rel_path': log.dest_rel_path,
            'user': log.user.username if log.user else '–',
            'timestamp': log.timestamp.strftime('%d.%m.%Y %H:%M'),
            'produkt': str(log.produkt) if log.produkt else '–',
            'produkt_pk': log.produkt_id,
            'size': log.file_size,
        }
        for log in qs[:TRASH_LIST_LIMIT]
    ]
    return JsonResponse({'status': 'ok', 'items': items})


# ─── Trash restore ───────────────────────────────────────────────────────────

@login_required
@require_POST
def docs_trash_restore(request):
    """Restore a trash item to its original location.

    POST /api/docs/trash/restore/
    Body JSON: {"log_id": <int>}
    """
    err = _docs_require_admin(request.user)
    if err:
        return err

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': 'error', 'message': 'Neplatný JSON.'}, status=400)

    log_id = data.get('log_id')
    try:
        log_entry = DocumentAuditLog.objects.get(
            pk=log_id, action=DocumentAuditLog.ACTION_DELETE
        )
    except DocumentAuditLog.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Záznam v koši nenájdený.'}, status=404)

    root = docs_root()
    t_root = trash_root()

    # Resolve trash source
    try:
        trash_src = safe_resolve(log_entry.dest_rel_path, t_root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná cesta v koši.'}, status=400)

    if not trash_src.exists():
        return JsonResponse({'status': 'error', 'message': 'Súbor v koši neexistuje (možno bol vymazaný).'}, status=404)

    # Resolve restore destination
    try:
        restore_dest = safe_resolve(log_entry.src_rel_path, root)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Neplatná pôvodná cesta pre obnovenie.'}, status=400)

    restore_dest.parent.mkdir(parents=True, exist_ok=True)
    restore_dest = resolve_collision(restore_dest)

    shutil.move(str(trash_src), str(restore_dest))

    dest_rel = to_rel(restore_dest, root)
    _log_doc_action(request.user, log_entry.produkt, DocumentAuditLog.ACTION_RESTORE,
                    src=log_entry.dest_rel_path, dest=dest_rel)

    return JsonResponse({'status': 'ok', 'message': 'Súbor bol obnovený.', 'restored_to': dest_rel})





